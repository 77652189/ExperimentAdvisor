"""Excel workbook -> structured schema conversion for legacy fermenter run exports.

See migration_audit.py for the separate old-vs-new data migration audit tooling
that used to live in this file.
"""
from __future__ import annotations

import hashlib
import re
import warnings
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

PRODUCT_COLUMNS = ["code", "full_name", "created_at"]
STRAIN_COLUMNS = ["name", "created_at"]
EXPERIMENT_COLUMNS = [
    "id",
    "product_code",
    "experiment_date",
    "file_name",
    "recorder",
    "source_file_md5",
    "created_at",
]
FERMENTER_RUN_COLUMNS = [
    "id",
    "experiment_id",
    "fermenter_label",
    "sheet_name",
    "strain_name",
    "batch_number",
    "inoculum_ratio",
    "seed_culture_time",
    "seed_od_value",
    "inoculation_time",
    "fermentation_end_time",
    "condition_notes",
    "created_at",
]
TIME_SERIES_COLUMNS = [
    "id",
    "fermenter_run_id",
    "fermentation_time_h",
    "temperature_c",
    "ph",
    "feed1_ml",
    "feed2_ml",
    "base_ml",
    "lactose_ml",
    "volume_ml",
    "od600",
    "yield_g_per_l",
    "lactose_g_per_l",
    "remarks",
    "created_at",
]
HPLC_COLUMNS = [
    "id",
    "fermenter_run_id",
    "sample_time_h",
    "extracellular_yield_g_per_l",
    "inactivated_yield_g_per_l",
    "extracellular_lactose_g_per_l",
    "extracellular_lactose_peak_area",
    "inactivated_lactose_g_per_l",
    "inactivated_lactose_peak_area",
    "extracellular_acetate_g_per_l",
    "inactivated_acetate_g_per_l",
    "created_at",
]
EXCEL_CELL_COLUMNS = [
    "file_name",
    "sheet_name",
    "row",
    "column",
    "cell",
    "value",
    "formula",
    "is_formula",
]
SUPPLEMENTAL_CELL_COLUMNS = [
    "file_name",
    "sheet_name",
    "fermenter_run_id",
    "row",
    "column",
    "cell",
    "value",
    "formula",
]
LIQUID_LONG_COLUMNS = [
    "id",
    "experiment_id",
    "file_name",
    "sheet_name",
    "section",
    "sample_label",
    "sample_time_h",
    "value",
    "formula",
    "source_cell",
    "created_at",
]

_KNOWN_METADATA_LABELS = (
    "文件名称",
    "发酵罐编号",
    "发酵批次",
    "页码",
    "发酵菌株名称",
    "种子液接种比例",
    "种子液培养时间",
    "种子液最终 OD 值",
    "接种上罐时间",
    "发酵结束时间",
    "发酵条件优化操作",
)


@dataclass(frozen=True)
class ConversionResult:
    output_dir: Path
    tables: dict[str, pd.DataFrame]
    skipped_sheets: list[dict[str, str]]


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\s_（）()/-]+", "", _text(value)).casefold()


def _number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _parse_date_from_filename(path: Path) -> str:
    match = re.match(r"(\d{4})-(\d{2})-(\d{2})", path.name)
    if not match:
        raise ValueError(f"Cannot parse experiment date from filename: {path.name}")
    return "-".join(match.groups())


def _parse_product_code(path: Path) -> str:
    match = re.search(r"(\d+)[-\s_]*(FL|SL)", path.name, flags=re.IGNORECASE)
    if match:
        return f"{match.group(1)}-{match.group(2).upper()}"
    for code in ("LNnT", "LNFP", "LNT", "SA"):
        if code.casefold() in path.name.casefold():
            return code
    return "UNKNOWN"


def _experiment_id(path: Path, product_code: str) -> str:
    date = datetime.strptime(_parse_date_from_filename(path), "%Y-%m-%d")
    compact_product = product_code.replace("-", "")
    return f"EXP-{date:%y%m%d}-{compact_product}-01"


def _find_labeled_value(ws: Worksheet, label_keyword: str, max_row: int = 8) -> str:
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, ws.max_column + 1):
            value = _text(ws.cell(row, col).value)
            if label_keyword in value:
                after_colon = re.split(r"[:：]", value, maxsplit=1)
                if len(after_colon) == 2 and after_colon[1].strip():
                    return after_colon[1].strip()
                max_offset = 3 if col <= 2 else 4
                for offset in range(1, max_offset + 1):
                    candidate = _text(ws.cell(row, col + offset).value)
                    if candidate and not any(label in candidate for label in _KNOWN_METADATA_LABELS):
                        return candidate
    return ""


def _find_labeled_values(ws: Worksheet, label_keyword: str, max_row: int = 8) -> list[str]:
    values: list[str] = []
    for row in range(1, min(ws.max_row, max_row) + 1):
        for col in range(1, ws.max_column + 1):
            value = _text(ws.cell(row, col).value)
            if label_keyword not in value:
                continue
            max_offset = 3 if col <= 2 else 4
            for offset in range(1, max_offset + 1):
                candidate = _text(ws.cell(row, col + offset).value)
                if candidate and not any(label in candidate for label in _KNOWN_METADATA_LABELS):
                    values.append(candidate)
                    break
    return values


def _batch_number(ws: Worksheet) -> str:
    values = _find_labeled_values(ws, "发酵批次")
    if not values:
        return ""
    values = sorted(set(values), key=lambda value: (bool(re.search(r"-\d{1,3}$", value)), len(value)), reverse=True)
    return values[0]


def _condition_notes(ws: Worksheet) -> str:
    for row in range(1, min(ws.max_row, 8) + 1):
        for col in range(1, ws.max_column + 1):
            value = _text(ws.cell(row, col).value)
            if "发酵条件优化操作" in value:
                parts = re.split(r"[:：]", value, maxsplit=1)
                return parts[1].strip() if len(parts) == 2 else value
    return ""


def _strain_name(raw: str) -> str:
    match = re.search(r"[（(]([^）)]+)[）)]", raw)
    value = match.group(1) if match else raw
    value = value.replace(" ", "").strip()
    if value.endswith("S") and value[:-1]:
        return value[:-1]
    return value or "UNKNOWN"


def _id_suffix(value: str) -> str:
    suffix = re.sub(r"\s+", "_", value.strip())
    suffix = re.sub(r"[^0-9A-Za-z_\-\u4e00-\u9fff]+", "", suffix)
    return suffix or "RUN"


def _header_map(ws: Worksheet, header_row: int = 9) -> dict[str, int]:
    result: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        raw = _text(ws.cell(header_row, col).value)
        header = _normalize_header(raw)
        if "发酵时长" in raw:
            result["fermentation_time_h"] = col
        elif "温度" in raw:
            result["temperature_c"] = col
        elif header == "ph":
            result["ph"] = col
        elif "补料1" in raw:
            result["feed1_ml"] = col
        elif "补料2" in raw:
            result["feed2_ml"] = col
        elif "补碱" in raw:
            result["base_ml"] = col
        elif "乳糖" in raw and ("ml" in raw.casefold() or "ｍｌ" in raw.casefold()):
            result["lactose_ml"] = col
        elif ("实时体积" in raw or raw == "体积") and "volume_ml" not in result:
            result["volume_ml"] = col
        elif "OD 600" in raw or header == "od600":
            result["od600"] = col
        elif "产量" in raw and "总产量" not in raw and "yield_g_per_l" not in result:
            result["yield_g_per_l"] = col
        elif "乳糖" in raw and "g/L" in raw and "lactose_g_per_l" not in result:
            result["lactose_g_per_l"] = col
        elif "备注" in raw and "remarks" not in result:
            result["remarks"] = col
    return result


def _is_fermentation_sheet(ws: Worksheet) -> bool:
    headers = _header_map(ws)
    required = {"fermentation_time_h", "temperature_c", "ph"}
    return required.issubset(headers)


def _read_time_series(ws: Worksheet, run_id: str, created_at: str) -> list[dict[str, Any]]:
    headers = _header_map(ws)
    rows: list[dict[str, Any]] = []
    blank_streak = 0
    row_number = 1
    for row in range(10, ws.max_row + 1):
        time_col = headers.get("fermentation_time_h")
        time_value = _number(ws.cell(row, time_col).value) if time_col else None
        if time_value is None:
            if rows:
                blank_streak += 1
                if blank_streak >= 3:
                    break
            continue
        blank_streak = 0
        item: dict[str, Any] = {
            "id": f"{run_id}-{row_number:04d}",
            "fermenter_run_id": run_id,
            "fermentation_time_h": time_value,
            "created_at": created_at,
        }
        for name in TIME_SERIES_COLUMNS:
            if name in {"id", "fermenter_run_id", "fermentation_time_h", "created_at"}:
                continue
            col = headers.get(name)
            value = ws.cell(row, col).value if col else None
            item[name] = _number(value) if name not in {"remarks"} else _text(value)
        rows.append(item)
        row_number += 1
    return rows


def _main_table_row_set(ws: Worksheet) -> set[int]:
    headers = _header_map(ws)
    time_col = headers.get("fermentation_time_h")
    if not time_col:
        return set()
    rows: set[int] = set()
    blank_streak = 0
    for row in range(10, ws.max_row + 1):
        time_value = _number(ws.cell(row, time_col).value)
        if time_value is None:
            if rows:
                blank_streak += 1
                if blank_streak >= 3:
                    break
            continue
        blank_streak = 0
        rows.add(row)
    return rows


def _dump_workbook_cells(workbook_path: Path, values_workbook: Any, formulas_workbook: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for values_ws in values_workbook.worksheets:
        formulas_ws = formulas_workbook[values_ws.title]
        for row in range(1, values_ws.max_row + 1):
            for col in range(1, values_ws.max_column + 1):
                value = values_ws.cell(row, col).value
                formula_or_raw = formulas_ws.cell(row, col).value
                if value is None and formula_or_raw is None:
                    continue
                formula = formula_or_raw if isinstance(formula_or_raw, str) and formula_or_raw.startswith("=") else ""
                rows.append(
                    {
                        "file_name": workbook_path.name,
                        "sheet_name": values_ws.title,
                        "row": row,
                        "column": col,
                        "cell": f"{get_column_letter(col)}{row}",
                        "value": _cell_value(value),
                        "formula": formula,
                        "is_formula": bool(formula),
                    }
                )
    return rows


def _supplemental_cells(workbook_path: Path, values_ws: Worksheet, formulas_ws: Worksheet, run_id: str) -> list[dict[str, Any]]:
    main_rows = _main_table_row_set(values_ws)
    rows: list[dict[str, Any]] = []
    for row in range(10, values_ws.max_row + 1):
        if row in main_rows:
            continue
        for col in range(1, values_ws.max_column + 1):
            value = values_ws.cell(row, col).value
            formula_or_raw = formulas_ws.cell(row, col).value
            if value is None and formula_or_raw is None:
                continue
            formula = formula_or_raw if isinstance(formula_or_raw, str) and formula_or_raw.startswith("=") else ""
            rows.append(
                {
                    "file_name": workbook_path.name,
                    "sheet_name": values_ws.title,
                    "fermenter_run_id": run_id,
                    "row": row,
                    "column": col,
                    "cell": f"{get_column_letter(col)}{row}",
                    "value": _cell_value(value),
                    "formula": formula,
                }
            )
    return rows


def _looks_like_liquid_sheet(ws: Worksheet) -> bool:
    title = ws.title.casefold()
    if "hplc" in title:
        return True
    if "\u6db2" in ws.title:
        return True
    for row in range(1, min(ws.max_row, 90) + 1):
        for col in range(1, min(ws.max_column, 8) + 1):
            value = _text(ws.cell(row, col).value)
            if "\u80de\u5916\u4ea7\u91cf" in value or "\u706d\u6d3b\u4ea7\u91cf" in value or "\u4e59\u9178" in value:
                return True
    return False


def _section_name(value: str, current: str) -> str:
    if not value:
        return current
    if "OD600" in value or value.upper() == "OD600":
        return "od600"
    if "\u80de\u5916\u4ea7\u91cf" in value:
        return "extracellular_yield_g_per_l"
    if "\u706d\u6d3b\u4ea7\u91cf" in value:
        return "inactivated_yield_g_per_l"
    if "\u4e73\u7cd6" in value:
        return "lactose_g_per_l"
    if "\u4e59\u9178" in value:
        return "acetate_g_per_l"
    if "\u4ea7\u7269\u5dee\u503c" in value:
        return "product_delta"
    if "\u65f6\u7a7a\u4ea7\u7387" in value:
        return "space_time_yield"
    if "\u4f53\u79ef" in value:
        return "volume_ml"
    return current


def _parse_liquid_long(
    workbook_path: Path,
    values_ws: Worksheet,
    formulas_ws: Worksheet,
    experiment_id: str,
    created_at: str,
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    item_index = 1

    def label_at(row: int, col: int) -> str:
        value = _text(values_ws.cell(row, col).value)
        if value.startswith("#"):
            value = _text(formulas_ws.cell(row, col).value)
        return value

    def add_value(row: int, col: int, section: str, sample_label: str, sample_time: float) -> None:
        nonlocal item_index
        value = values_ws.cell(row, col).value
        formula_or_raw = formulas_ws.cell(row, col).value
        if value is None and formula_or_raw is None:
            return
        formula = formula_or_raw if isinstance(formula_or_raw, str) and formula_or_raw.startswith("=") else ""
        rows.append(
            {
                "id": f"{experiment_id}-{values_ws.title}-L{item_index:04d}",
                "experiment_id": experiment_id,
                "file_name": workbook_path.name,
                "sheet_name": values_ws.title,
                "section": section,
                "sample_label": sample_label,
                "sample_time_h": sample_time,
                "value": _number(value),
                "formula": formula,
                "source_cell": f"{get_column_letter(col)}{row}",
                "created_at": created_at,
            }
        )
        item_index += 1

    def numeric_rows(start_row: int) -> list[int]:
        result: list[int] = []
        for row in range(start_row, values_ws.max_row + 1):
            if _number(values_ws.cell(row, 1).value) is None:
                if result:
                    break
                continue
            result.append(row)
        return result

    def row_labels(label_row: int) -> dict[int, str]:
        return {
            col: label_at(label_row, col)
            for col in range(2, values_ws.max_column + 1)
            if label_at(label_row, col) and _number(label_at(label_row, col)) is None
        }

    def parse_wide_block(label_row: int, start_row: int, section: str, section_by_col: dict[int, str] | None = None) -> None:
        labels = row_labels(label_row)
        for row in numeric_rows(start_row):
            sample_time = _number(values_ws.cell(row, 1).value)
            if sample_time is None:
                continue
            for col, sample_label in labels.items():
                add_value(row, col, section_by_col.get(col, section) if section_by_col else section, sample_label, sample_time)

    if values_ws.max_row >= 2 and _number(values_ws.cell(2, 1).value) is not None:
        parse_wide_block(1, 2, "od600")

    if values_ws.max_row >= 22 and _number(values_ws.cell(22, 1).value) is not None:
        section_by_col = {}
        for idx, col in enumerate(col for col in range(2, values_ws.max_column + 1) if label_at(21, col)):
            section_by_col[col] = "extracellular_yield_g_per_l" if idx % 2 == 0 else "extracellular_lactose_g_per_l"
        parse_wide_block(21, 22, "extracellular_yield_g_per_l", section_by_col)

    for row in range(1, values_ws.max_row):
        sample_label = _text(values_ws.cell(row, 1).value)
        if not sample_label or _number(sample_label) is not None:
            continue
        headers = {}
        for col in range(2, values_ws.max_column + 1):
            header_text = _text(values_ws.cell(row + 1, col).value)
            if not any(token in header_text for token in ("\u4ea7\u91cf", "\u4e73\u7cd6\u91cf", "\u4e59\u9178\u91cf")):
                continue
            section = _section_name(header_text, "")
            if section:
                headers[col] = section
        if not headers:
            continue
        for data_row in numeric_rows(row + 2):
            sample_time = _number(values_ws.cell(data_row, 1).value)
            if sample_time is None:
                continue
            for col, section in headers.items():
                add_value(data_row, col, section, sample_label, sample_time)

    inactivated_header = next(
        (row for row in range(1, values_ws.max_row + 1) if "\u706d\u6d3b\u4ea7\u91cf" in _text(values_ws.cell(row, 1).value)),
        None,
    )
    if inactivated_header:
        label_row = inactivated_header + 1
        start_row = label_row + 1
        if _number(values_ws.cell(label_row, 1).value) is not None:
            start_row = label_row
        parse_wide_block(label_row, start_row, "inactivated_yield_g_per_l")

        product_delta_row = next(
            (row for row in range(start_row, values_ws.max_row + 1) if "\u4ea7\u7269\u5dee\u503c" in _text(values_ws.cell(row, 1).value)),
            None,
        )
        if product_delta_row:
            volume_rows = [row for row in range(start_row, product_delta_row) if _number(values_ws.cell(row, 1).value) is not None]
            if len(volume_rows) > 4:
                for row in volume_rows[-4:]:
                    sample_time = _number(values_ws.cell(row, 1).value)
                    if sample_time is None:
                        continue
                    for col, sample_label in row_labels(label_row).items():
                        add_value(row, col, "volume_ml", sample_label, sample_time)

    for marker, section in {
        "\u4ea7\u7269\u5dee\u503c": "product_delta",
        "\u65f6\u95f4\u5dee\u503c": "time_delta",
    }.items():
        marker_row = next((row for row in range(1, values_ws.max_row + 1) if marker in _text(values_ws.cell(row, 1).value)), None)
        if not marker_row:
            continue
        labels = row_labels(marker_row - 1) or row_labels(marker_row)
        for row in numeric_rows(marker_row + 1):
            sample_time = _number(values_ws.cell(row, 1).value)
            if sample_time is None:
                continue
            for col, sample_label in labels.items():
                add_value(row, col, section, sample_label, sample_time)

    for row in range(1, values_ws.max_row + 1):
        if "h~" not in _text(values_ws.cell(row + 1, 1).value):
            continue
        labels = row_labels(row)
        for data_row in range(row + 1, values_ws.max_row + 1):
            interval = _text(values_ws.cell(data_row, 1).value)
            if "h~" not in interval:
                break
            for col, sample_label in labels.items():
                add_value(data_row, col, "space_time_yield", sample_label, float(data_row))
        break
    return rows


def convert_excel_directory(
    excel_dir: str | Path = "data/excel",
    output_dir: str | Path = "data/csv_from_excel",
) -> ConversionResult:
    excel_path = Path(excel_dir)
    output_path = Path(output_dir)
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    products: dict[str, dict[str, Any]] = {}
    strains: dict[str, dict[str, Any]] = {}
    experiments: list[dict[str, Any]] = []
    fermenter_runs: list[dict[str, Any]] = []
    time_series: list[dict[str, Any]] = []
    hplc_rows: list[dict[str, Any]] = []
    excel_cells: list[dict[str, Any]] = []
    supplemental_cells: list[dict[str, Any]] = []
    liquid_long: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, str]] = []
    used_run_ids: set[str] = set()

    for workbook_path in sorted(excel_path.glob("*.xlsx")):
        product_code = _parse_product_code(workbook_path)
        experiment_id = _experiment_id(workbook_path, product_code)
        experiment_date = _parse_date_from_filename(workbook_path)
        products.setdefault(product_code, {"code": product_code, "full_name": "", "created_at": created_at})
        experiments.append(
            {
                "id": experiment_id,
                "product_code": product_code,
                "experiment_date": experiment_date,
                "file_name": workbook_path.name,
                "recorder": "",
                "source_file_md5": _md5(workbook_path),
                "created_at": created_at,
            }
        )

        with warnings.catch_warnings():
            warnings.simplefilter("ignore", UserWarning)
            workbook = load_workbook(workbook_path, read_only=False, data_only=True)
            formula_workbook = load_workbook(workbook_path, read_only=False, data_only=False)
        excel_cells.extend(_dump_workbook_cells(workbook_path, workbook, formula_workbook))

        fermentation_sheets: list[tuple[Worksheet, str]] = []
        for ws in workbook.worksheets:
            if not _is_fermentation_sheet(ws):
                if _looks_like_liquid_sheet(ws):
                    liquid_long.extend(_parse_liquid_long(workbook_path, ws, formula_workbook[ws.title], experiment_id, created_at))
                else:
                    skipped_sheets.append({"file_name": workbook_path.name, "sheet_name": ws.title, "reason": "not fermentation main table"})
                continue
            fermenter_label = _find_labeled_value(ws, "发酵罐编号") or ws.title
            fermentation_sheets.append((ws, fermenter_label))

        label_counts = pd.Series([label for _, label in fermentation_sheets]).value_counts().to_dict()
        for ws, fermenter_label in fermentation_sheets:
            batch_number = _batch_number(ws)
            strain = _strain_name(_find_labeled_value(ws, "发酵菌株名称"))
            run_suffix = _id_suffix(ws.title) if label_counts.get(fermenter_label, 0) > 1 else fermenter_label
            run_id = f"{experiment_id}-{run_suffix}"
            if run_id in used_run_ids:
                run_id = f"{experiment_id}-{_id_suffix(ws.title)}"
            used_run_ids.add(run_id)
            strains.setdefault(strain, {"name": strain, "created_at": created_at})
            fermenter_runs.append(
                {
                    "id": run_id,
                    "experiment_id": experiment_id,
                    "fermenter_label": fermenter_label,
                    "sheet_name": ws.title,
                    "strain_name": strain,
                    "batch_number": batch_number,
                    "inoculum_ratio": _find_labeled_value(ws, "种子液接种比例"),
                    "seed_culture_time": _find_labeled_value(ws, "种子液培养时间"),
                    "seed_od_value": _find_labeled_value(ws, "种子液最终 OD 值"),
                    "inoculation_time": _find_labeled_value(ws, "接种上罐时间"),
                    "fermentation_end_time": _find_labeled_value(ws, "发酵结束时间"),
                    "condition_notes": _condition_notes(ws),
                    "created_at": created_at,
                }
            )
            time_series.extend(_read_time_series(ws, run_id, created_at))
            supplemental_cells.extend(_supplemental_cells(workbook_path, ws, formula_workbook[ws.title], run_id))

    tables = {
        "product": pd.DataFrame(products.values(), columns=PRODUCT_COLUMNS),
        "strain": pd.DataFrame(strains.values(), columns=STRAIN_COLUMNS),
        "experiment": pd.DataFrame(experiments, columns=EXPERIMENT_COLUMNS),
        "fermenter_run": pd.DataFrame(fermenter_runs, columns=FERMENTER_RUN_COLUMNS),
        "time_series_data": pd.DataFrame(time_series, columns=TIME_SERIES_COLUMNS),
        "hplc_data": pd.DataFrame(hplc_rows, columns=HPLC_COLUMNS),
        "liquid_long_data": pd.DataFrame(liquid_long, columns=LIQUID_LONG_COLUMNS),
        "supplemental_cells": pd.DataFrame(supplemental_cells, columns=SUPPLEMENTAL_CELL_COLUMNS),
        "excel_cells": pd.DataFrame(excel_cells, columns=EXCEL_CELL_COLUMNS),
    }

    output_path.mkdir(parents=True, exist_ok=True)
    for table_name, frame in tables.items():
        frame.to_csv(output_path / f"{table_name}.csv", index=False, encoding="utf-8-sig")

    pd.DataFrame(skipped_sheets).to_csv(output_path / "skipped_sheets.csv", index=False, encoding="utf-8-sig")
    return ConversionResult(output_dir=output_path, tables=tables, skipped_sheets=skipped_sheets)


