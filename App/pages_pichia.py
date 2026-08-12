"""Pichia hLF shake-flask page: Round 1 / Round 2 / History tabs.

Split out of App/app.py.
"""
from __future__ import annotations

import hashlib
import itertools
import re
from io import BytesIO
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from experiment_advisor.recommendation.round1_design import (
    BASELINE as PICHIA_ROUND1_BASELINE,
    OFAT_LEVELS as PICHIA_OFAT_LEVELS,
    count_round1_rows,
    generate_round1_design,
    round1_template_columns,
)
from experiment_advisor.recommendation.round2_design import (
    ALL_VARIABLES as PICHIA_VARIABLES,
    CONTINUOUS_BOUNDS as PICHIA_CONTINUOUS_BOUNDS,
    FIXED_LEVELS as PICHIA_FIXED_LEVELS,
    OD_COL as PICHIA_OD_COL,
    TARGET_COL as PICHIA_TARGET_COL,
    FactorEffect,
    Round2Plan,
    analyze_interval_interaction,
    assemble_round2_design,
    canonical_analysis,
    classify_response_surface_case,
    evaluate_response_surface,
    fit_ccd_response_surface,
    gp_leave_one_out_cv,
    gp_partial_dependence,
    optimize_joint_desirability,
    plan_round2,
    predict_with_confidence_interval,
    recommend_round2_bo_batch,
    response_surface_grid,
    sensitivity_analysis,
    summarize_bo_recommendation,
    summarize_response_surface,
)
from App.ui_shared import _num, _clear_ui_cache, _remember_ui_cache

PICHIA_DATA_DIR = PROJECT_ROOT / "data" / "pichia"
PICHIA_UPLOAD_DIR = PICHIA_DATA_DIR / "uploads"
PICHIA_FINAL_DIR = PICHIA_DATA_DIR / "final"
PICHIA_TEMPLATE_DIR = PICHIA_DATA_DIR / "templates"
PICHIA_DEFAULT_DATASET_PATH = PICHIA_FINAL_DIR / "pichia_run_level_dataset.csv"
PICHIA_ROUND2_DATASET_PATH = PICHIA_FINAL_DIR / "pichia_round2_dataset.csv"
PICHIA_TEMPLATE_PATH = PICHIA_TEMPLATE_DIR / "pichia_run_level_template.csv"
PICHIA_VARIABLE_LABELS = {
    "seed_od": "种子初始 OD600",
    "glucose_pct": "补料葡萄糖浓度 (%)",
    "ph": "培养基 pH",
    "volume_ml": "初始装液量 (mL)",
    "temp_c": "发酵温度 (℃)",
    "interval_h": "补料时间间隔 (h)",
}

# dataviz palette (dark-mode slots) -- see .claude/plans note on chart conventions:
# fixed categorical order for run_type identity, one diverging pair for correlation,
# accent-vs-muted "emphasis" pair for the effect-magnitude significance cutoff.
PICHIA_RUN_TYPE_COLORS = {
    "baseline": "#3987e5",
    "ofat": "#d95926",
    "combo": "#199e70",
    "ccd": "#8e6fce",
    "interval_interaction": "#c9a227",
    "noise_reference": "#e0607e",
    "lhs": "#4fb8af",
}
PICHIA_RUN_TYPE_LABELS = {
    "baseline": "基线重复",
    "ofat": "单变量(OFAT)",
    "combo": "联合探索",
    "ccd": "响应面(CCD)",
    "interval_interaction": "补料间隔交互",
    "noise_reference": "新区域噪声参考",
    "lhs": "拉丁超立方(LHS)",
}
PICHIA_DIVERGING_COLORSCALE = [[0, "#3987e5"], [0.5, "#383835"], [1, "#e66767"]]
# sequential (dark->bright) for magnitude data (predicted yield on a 3D
# surface/contour) -- several user-selectable options rather than one fixed
# scale: a 3-stop single hue (the original choice) reads "on brand" but
# compresses most of the value range into similar-looking mid blues, which
# made adjacent CCD contour bands hard to tell apart in practice. The default
# widens that same blue hue's lightness/chroma span instead of changing hue
# family; the other options trade "on brand" for more perceptual travel
# (still never a full hue-wheel rainbow -- see dataviz skill's sequential
# palette guidance).
PICHIA_SEQUENTIAL_COLORSCALE_OPTIONS: dict[str, list[list[Any]]] = {
    "扩展单色蓝（推荐）": [[0, "#0a1420"], [0.2, "#163a6b"], [0.4, "#2c5fa8"], [0.6, "#5ba0e0"], [0.8, "#a8d4f5"], [1, "#eaf4fd"]],
    "单色蓝（原配色）": [[0, "#1b2838"], [0.5, "#2c5fa8"], [1, "#7fb8f5"]],
    "蓝→青双色": [[0, "#0d1b2e"], [0.25, "#1a5276"], [0.5, "#17a398"], [0.75, "#7ee8d8"], [1, "#eafff9"]],
    "Viridis（多色，辨识度最高）": [[0, "#440154"], [0.25, "#3b528b"], [0.5, "#21918c"], [0.75, "#5ec962"], [1, "#fde725"]],
}
PICHIA_SEQUENTIAL_COLORSCALE_DEFAULT = "扩展单色蓝（推荐）"
PICHIA_ACCENT_COLOR = "#3987e5"
PICHIA_MUTED_COLOR = "#6b6a64"

def _ensure_pichia_data_area() -> None:
    for directory in [PICHIA_UPLOAD_DIR, PICHIA_FINAL_DIR, PICHIA_TEMPLATE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
    if not PICHIA_TEMPLATE_PATH.exists():
        pd.DataFrame(columns=round1_template_columns()).to_csv(PICHIA_TEMPLATE_PATH, index=False, encoding="utf-8-sig")

def _pichia_restore_persisted_dataset(session_key: str, path: Path) -> bool:
    """Loads path into st.session_state[session_key] if that key isn't
    already set and the file exists -- st.session_state is per Streamlit
    process, so a server restart (not just a browser refresh) wipes it even
    though "save to data/pichia/final/" already wrote the real data to disk.
    Both save buttons literally promise "供...下次打开使用" (usable next time
    the app opens); without this, that promise only holds across a browser
    refresh, not a process restart, which is the more likely reason someone
    actually reopens the app later. Returns whether a restore happened, so
    the caller can show a one-time notice instead of staying silent about it."""
    if session_key in st.session_state:
        return False
    if not path.exists():
        return False
    try:
        df = pd.read_csv(path)
    except Exception:
        return False
    if df.empty:
        return False
    st.session_state[session_key] = df
    return True

def _pichia_numeric_results(df: pd.DataFrame) -> pd.DataFrame:
    numeric = df.copy()
    for column in (PICHIA_TARGET_COL, PICHIA_OD_COL):
        if column in numeric.columns:
            numeric[column] = pd.to_numeric(numeric[column], errors="coerce")
    return numeric

def _pichia_ui_records() -> list[dict[str, Any]]:
    records = st.session_state.setdefault("pichia_ui_design_records", [])
    if not isinstance(records, list):
        records = []
        st.session_state["pichia_ui_design_records"] = records
    return records

def _pichia_variable_display(row: dict[str, Any]) -> dict[str, Any]:
    return {PICHIA_VARIABLE_LABELS.get(name, name): _num(row.get(name)) for name in PICHIA_VARIABLES if name in row}

def _pichia_round1_builder() -> None:
    st.markdown("#### 方案配置")
    st.caption(
        "基线重复、单变量法(OFAT)、联合探索(LHS) 三个模块可以独立开关和配置——某个数量填 0，"
        "或不选任何变量，就相当于关掉那个模块，可以拼出纯 LHS、纯 OFAT 或任意组合。"
        "表单打开时的默认值就是已经和研发组确认过的方案（基线×3 + OFAT×11 + 联合探索×4，共 18 个样本）。"
    )

    st.markdown("##### 基线取值")
    baseline_config: dict[str, float] = {}
    continuous_vars = list(PICHIA_CONTINUOUS_BOUNDS)
    baseline_cols = st.columns(len(continuous_vars))
    for index, variable in enumerate(continuous_vars):
        lower, upper = PICHIA_CONTINUOUS_BOUNDS[variable]
        with baseline_cols[index]:
            baseline_config[variable] = float(
                st.number_input(
                    PICHIA_VARIABLE_LABELS.get(variable, variable),
                    min_value=float(lower),
                    max_value=float(upper),
                    value=float(PICHIA_ROUND1_BASELINE[variable]),
                    key=f"round1_builder_baseline_{variable}",
                )
            )
    fixed_vars = list(PICHIA_FIXED_LEVELS)
    fixed_cols = st.columns(len(fixed_vars))
    for index, variable in enumerate(fixed_vars):
        options = PICHIA_FIXED_LEVELS[variable]
        default_value = PICHIA_ROUND1_BASELINE[variable]
        with fixed_cols[index]:
            baseline_config[variable] = float(
                st.selectbox(
                    PICHIA_VARIABLE_LABELS.get(variable, variable),
                    options,
                    index=options.index(default_value) if default_value in options else 0,
                    key=f"round1_builder_baseline_{variable}",
                )
            )

    st.markdown("##### 基线重复")
    n_baseline = st.number_input(
        "基线重复次数（0 = 不生成基线重复行）",
        min_value=0,
        max_value=10,
        value=3,
        key="round1_builder_n_baseline",
    )

    st.markdown("##### 单变量法 (OFAT)")
    st.caption(
        "连续变量的测试水平不是固定选项——推荐值只是预填，勾掉即可移除，下面还有专门的输入框+按钮可以增加新水平；"
        "温度和补料间隔受设备限制，只能在 20/25/30℃、12/24h 这几个物理档位里选，不能新增。"
    )
    ofat_vars = st.multiselect(
        "参与 OFAT 的变量（不选 = 不生成 OFAT 行）",
        list(PICHIA_VARIABLES),
        default=list(PICHIA_VARIABLES),
        format_func=lambda v: PICHIA_VARIABLE_LABELS.get(v, v),
        key="round1_builder_ofat_vars",
    )
    ofat_levels_config: dict[str, list[float]] = {}
    ofat_errors: list[str] = []
    for variable in ofat_vars:
        default_levels = PICHIA_OFAT_LEVELS.get(variable, [])
        label = PICHIA_VARIABLE_LABELS.get(variable, variable)
        is_fixed_level = variable in PICHIA_FIXED_LEVELS
        if is_fixed_level:
            selected = st.multiselect(
                f"{label} 的测试水平",
                options=PICHIA_FIXED_LEVELS[variable],
                default=default_levels,
                help="设备只支持这几个固定档位，不能输入自定义数值。",
                key=f"round1_builder_ofat_levels_{variable}",
            )
        else:
            # explicit pool + add-button instead of relying on multiselect's
            # accept_new_options=True: that only reveals its "type to add" UI
            # once the box is clicked/focused, so it looked identical to a
            # plain fixed-choice dropdown in a screenshot -- not discoverable.
            # The add-row must run BEFORE the multiselect below: Streamlit
            # forbids writing to a widget's session_state key after that
            # widget has already been instantiated in the same script run, so
            # mutating pool/levels has to happen ahead of the multiselect call
            # that owns `levels_key`, not in a button handler placed after it.
            pool_key = f"round1_builder_ofat_pool_{variable}"
            levels_key = f"round1_builder_ofat_levels_{variable}"
            st.session_state.setdefault(pool_key, list(default_levels))
            st.session_state.setdefault(levels_key, list(default_levels))

            lower, upper = PICHIA_CONTINUOUS_BOUNDS[variable]
            st.caption(f"新增「{label}」测试水平")
            add_cols = st.columns([3, 1])
            with add_cols[0]:
                new_level = st.number_input(
                    f"新增{label}水平",
                    min_value=float(lower),
                    max_value=float(upper),
                    value=None,
                    step=0.1,
                    placeholder="输入数值",
                    label_visibility="collapsed",
                    key=f"round1_builder_ofat_newval_{variable}",
                )
            with add_cols[1]:
                add_clicked = st.button(
                    "+ 添加", width="stretch", key=f"round1_builder_ofat_addbtn_{variable}"
                )
            if add_clicked:
                if new_level is None:
                    st.warning("请先在左侧输入一个数值。")
                else:
                    pool = st.session_state[pool_key]
                    if new_level not in pool:
                        st.session_state[pool_key] = sorted(pool + [new_level])
                    levels = st.session_state[levels_key]
                    if new_level not in levels:
                        st.session_state[levels_key] = sorted(levels + [new_level])

            selected = st.multiselect(
                f"{label} 的测试水平（勾掉即可移除，上方可新增）",
                options=st.session_state[pool_key],
                key=levels_key,
            )
        try:
            ofat_levels_config[variable] = sorted({float(value) for value in selected})
        except (TypeError, ValueError):
            ofat_errors.append(label)
    if ofat_errors:
        st.error(f"以下变量输入了无法识别成数字的自定义水平，请检查：{', '.join(ofat_errors)}")

    st.markdown("##### 联合探索 (LHS)")
    n_combo = st.number_input(
        "联合探索点数（0 = 不生成联合探索行）",
        min_value=0,
        max_value=50,
        value=4,
        key="round1_builder_n_combo",
    )
    combo_vars = st.multiselect(
        "参与联合探索的连续变量",
        continuous_vars,
        default=continuous_vars,
        format_func=lambda v: PICHIA_VARIABLE_LABELS.get(v, v),
        key="round1_builder_combo_vars",
    )

    counts = count_round1_rows(ofat_levels_config, int(n_baseline), int(n_combo), baseline=baseline_config)
    st.caption(
        f"预计生成：基线 {counts['baseline']} + OFAT {counts['ofat']} + 联合探索 {counts['combo']} "
        f"= 共 {counts['total']} 行"
    )

    generate_col, reset_col = st.columns(2)
    with generate_col:
        if st.button(
            "生成 Round 1 设计",
            type="primary",
            width="stretch",
            disabled=bool(ofat_errors),
            key="generate_round1_design_button",
        ):
            st.session_state["round1_results_df"] = generate_round1_design(
                baseline=baseline_config,
                ofat_levels=ofat_levels_config or None,
                n_baseline_replicates=int(n_baseline),
                n_combo_points=int(n_combo),
                combo_variables=combo_vars or None,
            )
            st.session_state.pop("round2_bo_result", None)
            st.success(f"已生成 {counts['total']} 行 Round 1 设计。")
    with reset_col:
        if st.button("直接使用已验证方案（18 样本，不改上面的表单）", width="stretch", key="round1_use_preset"):
            st.session_state["round1_results_df"] = generate_round1_design()
            st.session_state.pop("round2_bo_result", None)
            st.success("已生成已验证的 18 样本方案。")

def _pichia_yield_scatter_chart(df: pd.DataFrame, value_col: str, title: str) -> go.Figure:
    """Box+jitter chart of `value_col` grouped by run_type: shows both the
    per-group distribution/spread and every individual point (run_id on hover),
    covering the "distribution" and "error/noise" asks in one figure."""

    numeric = df.copy()
    numeric[value_col] = pd.to_numeric(numeric[value_col], errors="coerce")
    plot_df = numeric.dropna(subset=[value_col])

    fig = go.Figure()
    for run_type in ["baseline", "ofat", "combo"]:
        subset = plot_df[plot_df["run_type"] == run_type]
        if subset.empty:
            continue
        fig.add_trace(
            go.Box(
                y=subset[value_col],
                x=[PICHIA_RUN_TYPE_LABELS[run_type]] * len(subset),
                boxpoints="all",
                jitter=0.5,
                pointpos=0,
                marker_color=PICHIA_RUN_TYPE_COLORS[run_type],
                line_color=PICHIA_RUN_TYPE_COLORS[run_type],
                text=subset["run_id"],
                hovertemplate="%{text}<br>" + value_col + "=%{y}<extra></extra>",
            )
        )
    fig.update_layout(
        title=title,
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        showlegend=False,
        yaxis_title=value_col,
        margin=dict(t=40, b=30, l=50, r=20),
        height=380,
    )
    return fig

def _pichia_correlation_heatmap(df: pd.DataFrame) -> go.Figure:
    columns = [*PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL]
    numeric = df[columns].apply(pd.to_numeric, errors="coerce")
    corr = numeric.corr(method="spearman", min_periods=3)
    labels = [PICHIA_VARIABLE_LABELS.get(column, column) for column in columns]
    fig = go.Figure(
        data=go.Heatmap(
            z=corr.to_numpy(),
            x=labels,
            y=labels,
            colorscale=PICHIA_DIVERGING_COLORSCALE,
            zmid=0,
            zmin=-1,
            zmax=1,
            hovertemplate="%{y} vs %{x}: %{z:.2f}<extra></extra>",
            colorbar=dict(title="Spearman r"),
        )
    )
    fig.update_layout(
        title="变量与产量/OD600 相关性（Spearman，仅 18 个样本，供参考，非结论性）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=40, b=40, l=110, r=20),
        height=450,
    )
    return fig

def _pichia_baseline_lookup(df: pd.DataFrame) -> dict[str, Any]:
    """Most-common value per variable column. Stands in for "the baseline"
    even when baseline-replicate rows are disabled (n_baseline_replicates=0),
    since every row that doesn't deliberately vary a given variable leaves it
    at that value -- so the mode across the whole design is still correct."""
    lookup: dict[str, Any] = {}
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            mode = df[variable].mode()
            if not mode.empty:
                lookup[variable] = mode.iloc[0]
    return lookup

def _pichia_format_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)

def _pichia_type_label(run_type: str, changed_variable: str | None) -> str:
    """"类型" column text. OFAT rows name the specific variable being swept
    ("单变量-种子初始OD600") rather than a generic "单变量(OFAT)" label shared
    by all 11 OFAT rows -- otherwise the type column alone can't tell them
    apart at a glance. Baseline/combo stay generic since every row of that
    type really is the same thing."""
    if run_type == "ofat":
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable or "")
        return f"单变量-{label}"
    if run_type == "interval_interaction" and changed_variable:
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable)
        return f"补料间隔交互-{label}"
    return PICHIA_RUN_TYPE_LABELS.get(run_type, run_type)

def _pichia_row_note(
    run_type: str,
    changed_variable: str | None,
    row: pd.Series | None = None,
    baseline_lookup: dict[str, Any] | None = None,
) -> str:
    if run_type == "baseline":
        return "当前摇瓶标准条件重复，用于估计批次噪声（纯误差）"
    if run_type == "ofat":
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable or "")
        if row is not None and baseline_lookup and changed_variable in baseline_lookup:
            old = _pichia_format_value(baseline_lookup[changed_variable])
            new = _pichia_format_value(row.get(changed_variable))
            return f"单变量法：「{label}」由基线 {old} 改为 {new}，其余变量维持基线不变"
        return f"单变量法：只改变「{label}」，其余变量维持基线不变"
    if run_type == "combo":
        return "联合探索：连续变量按拉丁超立方(LHS)联合取值，温度/补料间隔取固定档位组合"
    if run_type == "ccd":
        reused = row.get("reused_from_round1") if row is not None else None
        if reused:
            return f"响应面(CCD)：条件和 round 1 的 {reused} 完全重合，产量/OD600 已按该样本回填，不需要重新做"
        return "响应面(CCD)：round 1 显著变量的角点/轴点/中心点，用于拟合响应面方程"
    if run_type == "interval_interaction":
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable or "")
        return f"补料间隔交互：新补料间隔 x 「{label}」，检验间隔的效应是否随「{label}」变化"
    if run_type == "noise_reference":
        return "新区域噪声参考：新补料间隔下的重复测量，用于估计该区域此前没有的噪声水平"
    if run_type == "lhs":
        return "拉丁超立方(LHS)：随机空间填充点，喂给贝叶斯优化的GP模型，不用于响应面拟合"
    return ""

def _pichia_design_display_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Chinese-labeled, note-annotated preview of the design -- used for the
    plain in-app st.dataframe. No row coloring here: st.dataframe's Styler
    support renders inconsistently for this table in practice, so the actual
    "pretty, colored" output is the Excel workbook below instead of trying to
    force parity inside Streamlit's canvas-rendered grid."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    display = pd.DataFrame(index=df.index)
    display["编号"] = df["run_id"]
    display["类型"] = [
        _pichia_type_label(run_type.loc[idx], changed_variable.loc[idx]) for idx in df.index
    ]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            display[PICHIA_VARIABLE_LABELS.get(variable, variable)] = df[variable]
    display["收获时OD600"] = df.get(PICHIA_OD_COL)
    display["hLF产量"] = df.get(PICHIA_TARGET_COL)
    # only present after a multi-replicate upload (_average_duplicate_run_ids) --
    # kept visible here, not just in the one-time upload toast, since this table
    # is also what "保存到 final" persists and what gets looked at again later.
    if f"{PICHIA_TARGET_COL}_n" in df.columns:
        display["重复次数"] = df.get(f"{PICHIA_TARGET_COL}_n")
        display["产量重复间差值"] = df.get(f"{PICHIA_TARGET_COL}_spread")
    if f"{PICHIA_OD_COL}_spread" in df.columns:
        display["OD600重复间差值"] = df.get(f"{PICHIA_OD_COL}_spread")
    display["备注/目的"] = [
        _pichia_row_note(run_type.loc[idx], changed_variable.loc[idx], df.loc[idx], baseline_lookup)
        for idx in df.index
    ]
    return display

def _pichia_round1_workbook_bytes(df: pd.DataFrame) -> bytes:
    """Polished .xlsx twin of the design: colored rows by run_type, Chinese
    headers, an auto-generated notes column, a legend, and a frozen header --
    the same look already approved from an earlier one-off script, now built
    dynamically from whatever design was actually generated (any mix of
    baseline/OFAT/LHS) instead of a fixed example."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    font_name = "Calibri"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    baseline_fill = PatternFill("solid", fgColor="E2EFDA")
    combo_fill = PatternFill("solid", fgColor="DDEBF7")
    fillin_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # only "type" and "note" hold variable-length sentences -- everything else
    # is a short id/number, so it stays single-line and the table stays compact
    wrap_keys = {"run_type", "note"}

    columns: list[tuple[str, str, float]] = [("run_id", "编号", 9), ("run_type", "类型", 18)]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            columns.append((variable, PICHIA_VARIABLE_LABELS.get(variable, variable), 13))
    columns.append((PICHIA_OD_COL, "收获时OD600(待填)", 12))
    columns.append((PICHIA_TARGET_COL, "hLF产量(待填)", 12))
    columns.append(("note", "备注/目的", 34))

    wb = Workbook()
    ws = wb.active
    ws.title = "Round1设计"
    ws.sheet_view.showGridLines = False

    for col_idx, (_key, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # deliberately no explicit row_dimensions[1].height: leaving customHeight
    # unset lets Excel auto-fit each row to its own wrapped content on open,
    # which is what actually works when note length varies this much between
    # row types -- any single fixed height is either too tall for short
    # baseline notes or clipped for the longer combo/ofat notes.
    ws.freeze_panes = "A2"

    row_fill_by_type = {"baseline": baseline_fill, "combo": combo_fill}

    for offset, idx in enumerate(df.index):
        row_idx = offset + 2
        rt = run_type.loc[idx]
        row = df.loc[idx]
        note = _pichia_row_note(rt, changed_variable.loc[idx], row, baseline_lookup)
        values: dict[str, Any] = {
            "run_id": row.get("run_id"),
            "run_type": _pichia_type_label(rt, changed_variable.loc[idx]),
            PICHIA_OD_COL: row.get(PICHIA_OD_COL),
            PICHIA_TARGET_COL: row.get(PICHIA_TARGET_COL),
            "note": note,
        }
        for variable in PICHIA_VARIABLES:
            if variable in df.columns:
                values[variable] = row.get(variable)

        row_fill = row_fill_by_type.get(rt)
        for col_idx, (key, _label, _width) in enumerate(columns, start=1):
            value = values.get(key)
            if isinstance(value, float) and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            if key == "note":
                cell.alignment = wrap_left
            elif key in wrap_keys:
                cell.alignment = wrap_center
            else:
                cell.alignment = center
            if key in (PICHIA_OD_COL, PICHIA_TARGET_COL):
                cell.fill = fillin_fill
            elif row_fill is not None:
                cell.fill = row_fill

    # legend lives on its own sheet, not appended below the data table --
    # otherwise pd.read_excel() on reimport reads it as extra data rows
    # (confirmed: it silently inflated an 18-row design to 23 rows).
    legend_ws = wb.create_sheet("图例说明")
    legend_ws.sheet_view.showGridLines = False
    legend_ws.column_dimensions["A"].width = 4
    legend_ws.column_dimensions["B"].width = 40
    legend_items = [
        (baseline_fill, "基线重复（估计批次噪声）"),
        (combo_fill, "联合探索点（LHS，多变量同时变化）"),
        (fillin_fill, "需要填写的结果列"),
    ]
    for row_idx, (fill, text) in enumerate(legend_items, start=1):
        marker = legend_ws.cell(row=row_idx, column=1, value=" ")
        marker.fill = fill
        marker.border = border
        label_cell = legend_ws.cell(row=row_idx, column=2, value=text)
        label_cell.font = Font(name=font_name, size=10)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def _pichia_round2_workbook_bytes(df: pd.DataFrame) -> bytes:
    """Round 2 twin of _pichia_round1_workbook_bytes: same visual style
    (colored rows by run_type, Chinese headers, notes column, legend), but
    two differences instead of a from-scratch rewrite. (1) run_type/
    changed_variable ride along as hidden columns rather than being encoded
    into "类型" text and decoded back on reimport -- round 1's "类型" decoder
    only recognizes 基线重复/联合探索/单变量-X, and teaching it four more
    prefixes for a one-off Round 2 sheet isn't worth the risk of a subtly
    wrong decode; keeping the machine-readable columns literally present
    means _pichia_remap_uploaded_columns needs no Round-2-specific case at
    all. (2) row colors come from the now 7-entry PICHIA_RUN_TYPE_COLORS
    rather than the 2-entry round-1 subset."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    font_name = "Calibri"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    fillin_fill = PatternFill("solid", fgColor="FFF2CC")
    type_fills = {run_type_key: PatternFill("solid", fgColor=color.lstrip("#")) for run_type_key, color in PICHIA_RUN_TYPE_COLORS.items()}
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wrap_keys = {"run_type", "note"}

    columns: list[tuple[str, str, float]] = [("run_id", "编号", 9), ("run_type", "类型", 20)]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            columns.append((variable, PICHIA_VARIABLE_LABELS.get(variable, variable), 13))
    columns.append((PICHIA_OD_COL, "收获时OD600(待填)", 12))
    columns.append((PICHIA_TARGET_COL, "hLF产量(待填)", 12))
    columns.append(("note", "备注/目的", 36))
    hidden_columns: list[tuple[str, str, float]] = [("_run_type_raw", "run_type", 12), ("_changed_variable_raw", "changed_variable", 16)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Round2设计"
    ws.sheet_view.showGridLines = False

    all_columns = columns + hidden_columns
    for col_idx, (_key, label, width) in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(len(columns) + 1, len(all_columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    ws.freeze_panes = "A2"

    for offset, idx in enumerate(df.index):
        row_idx = offset + 2
        rt = run_type.loc[idx]
        cv = changed_variable.loc[idx]
        row = df.loc[idx]
        note = _pichia_row_note(rt, cv, row, baseline_lookup)
        values: dict[str, Any] = {
            "run_id": row.get("run_id"),
            "run_type": _pichia_type_label(rt, cv),
            PICHIA_OD_COL: row.get(PICHIA_OD_COL),
            PICHIA_TARGET_COL: row.get(PICHIA_TARGET_COL),
            "note": note,
            "_run_type_raw": rt,
            "_changed_variable_raw": None if pd.isna(cv) else cv,
        }
        for variable in PICHIA_VARIABLES:
            if variable in df.columns:
                values[variable] = row.get(variable)

        row_fill = type_fills.get(rt)
        for col_idx, (key, _label, _width) in enumerate(all_columns, start=1):
            value = values.get(key)
            if isinstance(value, float) and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            if key == "note":
                cell.alignment = wrap_left
            elif key in wrap_keys:
                cell.alignment = wrap_center
            else:
                cell.alignment = center
            if key in (PICHIA_OD_COL, PICHIA_TARGET_COL):
                cell.fill = fillin_fill
            elif row_fill is not None:
                cell.fill = row_fill

    legend_ws = wb.create_sheet("图例说明")
    legend_ws.sheet_view.showGridLines = False
    legend_ws.column_dimensions["A"].width = 4
    legend_ws.column_dimensions["B"].width = 40
    present_types = [key for key in PICHIA_RUN_TYPE_COLORS if key in set(run_type.dropna())]
    legend_items = [(type_fills[key], PICHIA_RUN_TYPE_LABELS.get(key, key)) for key in present_types]
    legend_items.append((fillin_fill, "需要填写的结果列"))
    for row_idx, (fill, text) in enumerate(legend_items, start=1):
        marker = legend_ws.cell(row=row_idx, column=1, value=" ")
        marker.fill = fill
        marker.border = border
        label_cell = legend_ws.cell(row=row_idx, column=2, value=text)
        label_cell.font = Font(name=font_name, size=10)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

_YIELD_HEADER_HINT = "产量"
_MG_PER_L_PATTERN = re.compile(r"mg\s*/\s*l", re.IGNORECASE)
_NOT_DETECTED_TOKENS = {"未检测到", "未检出", "未测出", "nd", "n.d.", "n/a"}


def _looks_not_detected(value: Any) -> bool:
    return isinstance(value, str) and value.strip().lower() in _NOT_DETECTED_TOKENS


def _coerce_result_numeric(series: pd.Series) -> pd.Series:
    """Real Round 1 results can hold a qualitative "below detection limit"
    token instead of a number in a result cell -- treated as 0 (a censored-low
    reading is evidence of "very little/none", not a missing measurement),
    everything else parsed as usual so a genuinely blank cell stays NaN."""
    cleaned = series.apply(lambda value: 0.0 if _looks_not_detected(value) else value)
    return pd.to_numeric(cleaned, errors="coerce")


def _average_duplicate_run_ids(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Collapse 2+ rows sharing a run_id (technical replicate measurements)
    into one row per run_id.

    The 6 design variables are expected to be *identical* across one run_id's
    replicate rows -- that's what makes them replicates. A mismatch there
    (data-entry typo, or a row mislabeled with the wrong run_id) is reported
    as a warning string rather than silently resolved by "whichever row came
    first", since silently picking one would hide a real data problem.

    Every other numeric column (yield, od600, and any future response
    variable such as UPR) is averaged, but the raw values are not discarded:
    "<col>_n" (replicate count), "<col>_spread" (max-min across replicates),
    and "<col>_reps" (the raw values, "|"-joined) ride alongside the mean --
    a large spread on a specific run is exactly the kind of thing averaging
    would otherwise hide, and the original numbers stay recoverable from this
    same table without reopening the raw upload archive. Single-replicate
    run_ids get the same three columns (n=1, spread/reps=None) so the schema
    is uniform whether or not that particular run happened to repeat.
    """
    order = df["run_id"].drop_duplicates().tolist()
    design_columns = [column for column in PICHIA_VARIABLES if column in df.columns]
    passthrough_columns = [column for column in ("run_type", "changed_variable") if column in df.columns]
    numeric_columns = [
        column
        for column in df.columns
        if column not in ("run_id", *design_columns, *passthrough_columns) and pd.api.types.is_numeric_dtype(df[column])
    ]
    leftover_columns = [
        column
        for column in df.columns
        if column not in ("run_id", *design_columns, *passthrough_columns, *numeric_columns)
    ]

    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    for run_id, group in df.groupby("run_id", sort=False):
        row: dict[str, Any] = {"run_id": run_id}
        for column in (*design_columns, *passthrough_columns, *leftover_columns):
            distinct = group[column].dropna()
            if column in design_columns:
                distinct = distinct.round(6)
            distinct = distinct.unique()
            if column in design_columns and len(distinct) > 1:
                warnings.append(
                    f"编号 {run_id} 的「{PICHIA_VARIABLE_LABELS.get(column, column)}」在重复测量行之间不一致："
                    f"{list(distinct)}，已取第一行的值，请核对是否录入有误"
                )
            row[column] = group[column].iloc[0]
        for column in numeric_columns:
            values = group[column].dropna()
            row[column] = float(values.mean()) if len(values) else float("nan")
            row[f"{column}_n"] = int(len(values))
            row[f"{column}_spread"] = float(values.max() - values.min()) if len(values) > 1 else None
            row[f"{column}_reps"] = "|".join(f"{value:g}" for value in values) if len(values) > 1 else None
        rows.append(row)

    averaged = pd.DataFrame(rows)
    averaged["_order"] = averaged["run_id"].apply(order.index)
    averaged = averaged.sort_values("_order").drop(columns="_order").reset_index(drop=True)
    return averaged, warnings


def _pichia_replicate_spread_report(df: pd.DataFrame) -> pd.DataFrame:
    """Display-friendly extract of the "<col>_n"/"<col>_spread" columns
    _average_duplicate_run_ids attaches, limited to run_ids that actually had
    2+ replicates for the yield or od600 column -- for the upload-time "look
    at this now" callout, sorted by the caller worst-spread-first."""
    records: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        counts = [row.get(f"{column}_n") for column in (PICHIA_TARGET_COL, PICHIA_OD_COL) if f"{column}_n" in df.columns]
        if not any(pd.notna(count) and count > 1 for count in counts):
            continue
        record: dict[str, Any] = {"run_id": row["run_id"], "重复次数": int(max(count for count in counts if pd.notna(count)))}
        for column, label in ((PICHIA_TARGET_COL, "hLF产量"), (PICHIA_OD_COL, "OD600")):
            spread = row.get(f"{column}_spread")
            if pd.isna(spread):
                continue
            mean = row.get(column)
            record[f"{label}均值"] = _num(mean)
            record[f"{label}重复间差值"] = _num(spread)
            record[f"{label}重复原始值"] = row.get(f"{column}_reps")
        records.append(record)
    return pd.DataFrame(records)


def _pichia_pooled_technical_noise(df: pd.DataFrame, column: str = PICHIA_TARGET_COL) -> dict[str, Any] | None:
    """Pool within-run_id replicate variance across every run that has 2+
    technical replicates (parses "<column>_reps", see _average_duplicate_run_ids),
    rather than trusting any single run's own 1-degree-of-freedom spread.

    This is a genuinely different noise source from plan_round2's baseline_sd:
    that one is *between-run* (batch) noise, estimated from how much the 3
    baseline runs' means disagree with each other. This one is *within-run*
    (technical/assay) noise, estimated from how much one run's own repeated
    measurement wobbles. Standard pooled-variance treatment across k groups:
    s_pooled^2 = sum(SS_i) / sum(n_i - 1), which is why pooling across every
    run_id (not just baseline) matters here -- 16 runs x 1 df each gives a far
    more stable estimate than the 2 df available from 3 baseline runs alone.

    Deliberately kept separate from estimate_baseline_noise's significance
    threshold, not just left unfinished: every round 1 sample is measured with
    the same 2 technical replicates, so the between-run (batch) SD already IS
    the right noise scale for comparing one run's mean against another's --
    folding this in wouldn't change which variables clear the threshold (for
    2026-08 Y103 data, this pooled SD is smaller than baseline_sd anyway, so a
    "use whichever is stricter" rule would be a no-op), and the threshold only
    gates a soft, self-correcting choice (which variable gets a round-2 CCD
    slot) where the added rigor isn't worth the small-sample estimation risk a
    formal variance-components combination would carry. This function's value
    is as a data-quality flag on individual runs, independent of that decision.
    Returns None if the column has no replicate-detail data at all.
    """
    reps_column = f"{column}_reps"
    if reps_column not in df.columns:
        return None

    sum_of_squares = 0.0
    degrees_of_freedom = 0
    per_run_sd: list[tuple[str, float]] = []
    for _, row in df.iterrows():
        raw = row.get(reps_column)
        if not isinstance(raw, str) or not raw:
            continue
        values = [float(token) for token in raw.split("|")]
        if len(values) < 2:
            continue
        mean = sum(values) / len(values)
        run_ss = sum((value - mean) ** 2 for value in values)
        sum_of_squares += run_ss
        degrees_of_freedom += len(values) - 1
        per_run_sd.append((str(row["run_id"]), (run_ss / (len(values) - 1)) ** 0.5))

    if degrees_of_freedom == 0:
        return None

    pooled_sd = (sum_of_squares / degrees_of_freedom) ** 0.5
    outliers = sorted(
        ((run_id, sd) for run_id, sd in per_run_sd if sd > 2 * pooled_sd),
        key=lambda item: item[1],
        reverse=True,
    )
    return {
        "pooled_sd": pooled_sd,
        "dof": degrees_of_freedom,
        "n_runs": len(per_run_sd),
        "outliers": outliers,
    }


def _pichia_remap_uploaded_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Recognizes the polished Excel export's Chinese headers, in addition to
    the plain English round1_template_columns() schema, so a design that was
    downloaded, filled in by hand in Excel, and re-uploaded round-trips
    correctly. Also reconstructs run_type/changed_variable from the combined
    "类型" column (e.g. "单变量-发酵温度 (℃)") -- round2's significance
    analysis and the result charts both key off those two columns, so losing
    them on reimport would silently break Round 2, not just cosmetics.

    Two real-world variations seen in returned lab data (2026-08 Y103 round 1)
    are also handled: (1) the yield header can carry a unit annotation this
    app's own template doesn't include (e.g. "hLF产量（mg/L）(待填)") -- matched
    by "产量" appearing anywhere in a header rather than an exact string, and
    auto-converted to g/L when "mg/L" is present, since every yield_g_per_l
    consumer in this codebase assumes g/L; (2) a "未检测到" token in a result
    cell instead of a number.

    Returns (cleaned_df, replicate_spread_report, consistency_warnings).
    cleaned_df keeps one row per run_id -- analyze_round1_effects/plan_round2
    are built on that shape, and feeding raw un-collapsed replicate rows
    through them would silently double-count each run_id's evidence and
    conflate technical (within-run) noise with the between-run noise the
    baseline-replicate significance threshold is meant to estimate. Nothing
    from the original rows is discarded in exchange for that shape, though:
    see _average_duplicate_run_ids for the "<col>_n/_spread/_reps" columns
    that keep the raw values reachable from cleaned_df itself, and
    replicate_spread_report/consistency_warnings for what's worth a human
    looking at right away. Both are empty/[] unless some run_id had 2+ rows.
    """
    label_to_variable = {label: variable for variable, label in PICHIA_VARIABLE_LABELS.items()}
    rename_map = {"编号": "run_id", "收获时OD600(待填)": PICHIA_OD_COL}
    rename_map.update(label_to_variable)
    renamed = df.rename(columns=rename_map)

    yield_source = next(
        (
            column
            for column in renamed.columns
            if isinstance(column, str) and _YIELD_HEADER_HINT in column and column != PICHIA_TARGET_COL
        ),
        None,
    )
    if yield_source is not None:
        scale = 0.001 if _MG_PER_L_PATTERN.search(yield_source) else 1.0
        renamed[PICHIA_TARGET_COL] = _coerce_result_numeric(renamed[yield_source]) * scale
        renamed = renamed.drop(columns=[yield_source])
    elif PICHIA_TARGET_COL in renamed.columns:
        renamed[PICHIA_TARGET_COL] = _coerce_result_numeric(renamed[PICHIA_TARGET_COL])
    if PICHIA_OD_COL in renamed.columns:
        renamed[PICHIA_OD_COL] = _coerce_result_numeric(renamed[PICHIA_OD_COL])

    if "类型" in renamed.columns and "run_type" not in renamed.columns:
        run_types: list[str | None] = []
        changed_vars: list[str | None] = []
        for raw in renamed["类型"]:
            text = str(raw).strip()
            if text == PICHIA_RUN_TYPE_LABELS.get("baseline"):
                run_types.append("baseline")
                changed_vars.append(None)
            elif text == PICHIA_RUN_TYPE_LABELS.get("combo"):
                run_types.append("combo")
                changed_vars.append(None)
            elif text.startswith("单变量-"):
                run_types.append("ofat")
                changed_vars.append(label_to_variable.get(text[len("单变量-"):]))
            else:
                run_types.append(None)
                changed_vars.append(None)
        renamed["run_type"] = run_types
        renamed["changed_variable"] = changed_vars
        renamed = renamed.drop(columns=["类型"])

    renamed = renamed.drop(columns=["备注/目的"], errors="ignore")

    # defensive: drop any row missing a variable value -- every real design
    # row has all 6 populated, so this filters out a legend/blank row picked
    # up from an older export that still had the legend below the table (or
    # any stray manual annotation row) without needing to special-case it.
    variable_columns = [variable for variable in PICHIA_VARIABLES if variable in renamed.columns]
    if variable_columns:
        renamed = renamed[renamed[variable_columns].notna().all(axis=1)].reset_index(drop=True)

    if "run_id" in renamed.columns and renamed["run_id"].duplicated().any():
        renamed, consistency_warnings = _average_duplicate_run_ids(renamed)
        spread_report = _pichia_replicate_spread_report(renamed)
    else:
        spread_report = pd.DataFrame()
        consistency_warnings = []

    return renamed, spread_report, consistency_warnings

def _pichia_round1_tab() -> None:
    _ensure_pichia_data_area()
    if _pichia_restore_persisted_dataset("round1_results_df", PICHIA_DEFAULT_DATASET_PATH):
        st.toast(f"已从 {PICHIA_DEFAULT_DATASET_PATH.name} 恢复上次保存的 Round 1 结果。", icon="✅")
    st.markdown("### Round 1：实验设计构建器")
    st.caption(
        "温度只能是 20/25/30℃，补料间隔只能是 12/24h——这两个变量的档位一旦被单变量法测全，"
        "就已经穷尽了离散选项，Round 2 不会再对它们做响应面细化。"
    )

    with st.expander("方案配置", expanded="round1_results_df" not in st.session_state):
        _pichia_round1_builder()

    if "round1_results_df" not in st.session_state:
        st.info("配置好方案后点击「生成 Round 1 设计」，或直接用「直接使用已验证方案」一键生成。")
        return

    st.markdown("#### 上传已回填结果（可选）")
    uploaded = st.file_uploader(
        "上传已回填 yield_g_per_l / od600 的 CSV，或者上面下载的 Excel 填完后原样传回来",
        type=["csv", "xlsx"],
        key="round1_results_upload",
    )
    if uploaded is not None:
        try:
            payload = uploaded.getvalue()
            if uploaded.name.lower().endswith(".xlsx"):
                uploaded_df = pd.read_excel(BytesIO(payload))
            else:
                uploaded_df = pd.read_csv(BytesIO(payload))
            uploaded_df, replicate_spread, consistency_warnings = _pichia_remap_uploaded_columns(uploaded_df)
        except Exception as exc:
            st.error(f"读取失败：{exc}")
        else:
            missing = [column for column in ["run_id", *PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL] if column not in uploaded_df.columns]
            if missing:
                st.error(f"上传文件缺少字段：{', '.join(missing)}")
            else:
                _ensure_pichia_data_area()
                archive_name = Path(uploaded.name).name or "pichia_round1_uploaded.csv"
                (PICHIA_UPLOAD_DIR / archive_name).write_bytes(payload)
                st.session_state["round1_results_df"] = uploaded_df
                st.success("已加载并归档上传的 Round 1 结果。")
                for warning_text in consistency_warnings:
                    st.warning(warning_text)
                if not replicate_spread.empty:
                    st.markdown("##### 同一编号的重复测量")
                    st.caption(
                        "检测到部分编号有 2 次及以上重复测量，后续显著性/响应面分析用的是均值；"
                        "原始的每次重复数值和差值没有被丢弃，一直跟着这一行数据（存到 final 时也在），"
                        "下表按 hLF产量的重复间差值从大到小排列，供你判断重复间是否足够一致。"
                    )
                    sort_column = next(
                        (column for column in replicate_spread.columns if column.endswith("重复间差值")), None
                    )
                    if sort_column:
                        replicate_spread = replicate_spread.sort_values(sort_column, ascending=False)
                    st.dataframe(replicate_spread, width="stretch", hide_index=True)

    working_df = st.session_state["round1_results_df"]

    st.markdown("#### 设计总览")
    excel_bytes = _pichia_round1_workbook_bytes(working_df)
    st.download_button(
        "下载设计表格（Excel，配色+备注，和示例一致）",
        data=excel_bytes,
        file_name="pichia_round1_design.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_round1_excel",
    )
    st.caption("配色、图例和备注列在 Excel 里最完整；下面是网页内的快速预览（不带颜色）。")
    st.dataframe(_pichia_design_display_frame(working_df), width="stretch", hide_index=True)

    st.markdown("#### 回填实测结果")
    st.caption("按编号对应到上面总览表里的样本，这里只需要填产量和 OD600 这两列。")
    entry_view = working_df[["run_id", PICHIA_TARGET_COL, PICHIA_OD_COL]].reset_index(drop=True)
    edited_entries = st.data_editor(
        entry_view,
        width="stretch",
        num_rows="fixed",
        hide_index=True,
        key="round1_data_editor",
        disabled=["run_id"],
        column_config={
            "run_id": "编号",
            PICHIA_TARGET_COL: "hLF产量",
            PICHIA_OD_COL: "收获时OD600",
        },
    )
    edited = working_df.copy()
    edited[PICHIA_TARGET_COL] = edited_entries[PICHIA_TARGET_COL].to_numpy()
    edited[PICHIA_OD_COL] = edited_entries[PICHIA_OD_COL].to_numpy()
    st.session_state["round1_results_df"] = edited

    numeric = _pichia_numeric_results(edited)
    filled = int(numeric[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1).sum())
    st.caption(f"已回填 {filled}/{len(edited)} 行完整结果（yield_g_per_l 和 od600 都有值才算完成）。")

    action_cols = st.columns(2)
    with action_cols[0]:
        csv_bytes = edited.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "下载当前表格",
            data=csv_bytes,
            file_name="pichia_round1_results.csv",
            mime="text/csv",
            key="download_round1_results",
        )
    with action_cols[1]:
        if st.button("保存到 data/pichia/final/（供 Round 2 和下次打开使用）", key="save_round1_to_final"):
            _ensure_pichia_data_area()
            edited.to_csv(PICHIA_DEFAULT_DATASET_PATH, index=False, encoding="utf-8-sig")
            st.success(f"已保存到 {PICHIA_DEFAULT_DATASET_PATH}")

    if int(numeric[PICHIA_TARGET_COL].notna().sum()) >= 3:
        st.markdown("#### 结果可视化")
        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.plotly_chart(
                _pichia_yield_scatter_chart(numeric, PICHIA_TARGET_COL, "产量分布（按样本类型）"),
                width="stretch",
            )
        with chart_cols[1]:
            st.plotly_chart(
                _pichia_yield_scatter_chart(numeric, PICHIA_OD_COL, "OD600 分布（按样本类型）"),
                width="stretch",
            )
        st.plotly_chart(_pichia_correlation_heatmap(numeric), width="stretch")
    else:
        st.caption("至少回填 3 行产量数据后，这里会显示分布图和相关性热力图。")

def _pichia_effect_magnitude_chart(effects: dict[str, FactorEffect], threshold: float) -> go.Figure:
    """Horizontal bar of each variable's effect_magnitude, descending, with a
    confidence-interval error bar and a reference line at the significance
    threshold -- this is what actually decides K (how many variables become
    "active" in resolve_round2_variables). Emphasis coloring (accent vs muted)
    rather than a full categorical palette, since the only distinction that
    matters here is "crossed the line" vs "didn't"."""

    ordered = sorted(effects.values(), key=lambda effect: effect.effect_magnitude, reverse=True)
    labels = [PICHIA_VARIABLE_LABELS.get(effect.variable, effect.variable) for effect in ordered]
    magnitudes = [effect.effect_magnitude for effect in ordered]
    colors = [PICHIA_ACCENT_COLOR if effect.significant else PICHIA_MUTED_COLOR for effect in ordered]
    error_plus = [max(effect.ci_high - effect.effect_magnitude, 0.0) for effect in ordered]
    error_minus = [max(effect.effect_magnitude - effect.ci_low, 0.0) for effect in ordered]

    fig = go.Figure(
        go.Bar(
            x=magnitudes,
            y=labels,
            orientation="h",
            marker_color=colors,
            error_x=dict(type="data", symmetric=False, array=error_plus, arrayminus=error_minus, color="#c3c2b7"),
            hovertemplate="%{y}: %{x:.3g}<extra></extra>",
        )
    )
    fig.add_vline(
        x=threshold,
        line_dash="dash",
        line_color="#c3c2b7",
        annotation_text=f"显著性阈值 {threshold:.3g}",
        annotation_position="top",
    )
    fig.update_layout(
        title="各变量效应量（误差棒=置信区间，仅 3 次基线重复，df=2，区间偏宽属实情）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title="效应量 |偏离基线|",
        margin=dict(t=50, b=40, l=140, r=30),
        height=340,
    )
    return fig

def _pichia_simulate_round2_results(design: pd.DataFrame, seed: int = 42) -> pd.DataFrame:
    """Fills currently-blank yield/od600 cells with deterministic, plausible
    (not random-garbage) synthetic values -- same formula and seed used to
    build the one-off demo Excel generated earlier this project (a clean
    quadratic in glucose/ph/volume centered near round 1's own best region,
    a modest fixed penalty for the untested 36h interval, small Gaussian
    noise). Never touches a cell that already has a real value: this exists
    so someone can exercise the CCD-fit/interaction/BO-visualization pipeline
    before real Round 2 results exist, not to backfill Round 2 for them."""
    result = design.copy()
    rng = np.random.default_rng(seed)
    center = {"glucose_pct": 1.0, "ph": 6.2, "volume_ml": 50.0}
    still_blank = result[PICHIA_TARGET_COL].isna()
    if not still_blank.any():
        return result

    glucose = result.loc[still_blank, "glucose_pct"]
    ph = result.loc[still_blank, "ph"]
    volume = result.loc[still_blank, "volume_ml"]
    interval = result.loc[still_blank, "interval_h"]
    yield_value = (
        0.013
        - 0.04 * (glucose - center["glucose_pct"]) ** 2
        - 0.015 * (ph - center["ph"]) ** 2
        + 0.00015 * (volume - center["volume_ml"])
        - 0.0000005 * (volume - center["volume_ml"]) ** 2
    )
    yield_value = yield_value.where(interval != 36.0, yield_value * 0.85)
    yield_value = (yield_value + rng.normal(0.0, 0.0003, int(still_blank.sum()))).clip(lower=0.00005)
    od_base = pd.Series(np.where(interval.to_numpy() == 24.0, 32.0, 26.0), index=interval.index)
    od_value = (od_base + rng.normal(0.0, 1.0, int(still_blank.sum()))).clip(lower=5.0)

    result.loc[still_blank, PICHIA_TARGET_COL] = yield_value
    result.loc[still_blank, PICHIA_OD_COL] = od_value
    return result

def _pichia_round2_design_and_backfill_section(plan: Round2Plan, round1_df: pd.DataFrame) -> None:
    """Extends plan's CCD table with two blocks generate_ccd can never
    produce on its own -- a feed-interval level round 1 never tried, crossed
    with one continuous variable to check for interaction, plus Latin
    Hypercube space-filling points for whichever BO round consumes the
    combined dataset next (see round2_design.generate_round2_extension_design
    for why both are structurally separate from the CCD track) -- then gives
    the combined sheet the same download / upload-backfill / save-to-final
    loop Round 1's design has, so results collected under it are reproducible
    and don't dead-end in a chat transcript."""
    _ensure_pichia_data_area()
    if _pichia_restore_persisted_dataset("round2_full_design_df", PICHIA_ROUND2_DATASET_PATH):
        st.toast(f"已从 {PICHIA_ROUND2_DATASET_PATH.name} 恢复上次保存的 Round 2 结果。", icon="✅")
    st.markdown("#### 完整 Round 2 设计表（CCD + 补料间隔交互 + LHS）")
    st.caption(
        "在上面的响应面(CCD)之外，再加两块：一是测试一个 round 1 没试过的补料间隔水平，"
        "看它的效应是否跟某个连续变量有交互；二是拉丁超立方(LHS)随机点，专门填补 CCD 网格之间的空隙，"
        "供后续贝叶斯优化的 GP 模型使用。同样的输入（round 1 数据 + 下面这些参数）永远生成同一张表，方便留档核对。"
    )

    interaction_choices = plan.active_variables or list(PICHIA_CONTINUOUS_BOUNDS)
    param_cols = st.columns(4)
    with param_cols[0]:
        extra_interval = st.number_input(
            "新补料间隔 (h)",
            min_value=0.0,
            value=36.0,
            step=1.0,
            help="round 1 从未测过的一个补料间隔水平，例如比现有 24h 更长的档位；具体数值需先跟研发组确认设备/排期可行。",
            key="round2_extra_interval",
        )
    with param_cols[1]:
        interaction_variable = st.selectbox(
            "跟新间隔做交互的变量",
            options=interaction_choices,
            format_func=lambda name: PICHIA_VARIABLE_LABELS.get(name, name),
            help="通常选一个 round 1 已经显著的变量，交互测试才有决策价值。",
            key="round2_interaction_variable",
        )
    with param_cols[2]:
        n_noise_reference = st.number_input(
            "新间隔噪声参考重复数",
            min_value=0,
            max_value=6,
            value=2,
            help="新补料间隔区域没有任何历史噪声数据，这里的重复测量专门给它一个噪声估计，不是可选的锦上添花。",
            key="round2_n_noise_reference",
        )
    with param_cols[3]:
        n_lhs = st.number_input(
            "LHS 点数",
            min_value=0,
            max_value=30,
            value=10,
            help="纯随机空间填充点，不进入响应面拟合，只用于让后续贝叶斯优化的 GP 模型看到更多样的条件组合。",
            key="round2_n_lhs",
        )

    if st.button("生成完整 Round 2 设计", type="primary", key="generate_round2_full_design"):
        full_design = assemble_round2_design(
            plan,
            PICHIA_ROUND1_BASELINE,
            round1_df,
            extra_interval_levels=[float(extra_interval)],
            interaction_variable=interaction_variable,
            n_noise_reference=int(n_noise_reference),
            n_lhs=int(n_lhs),
            seed=42,
        )
        st.session_state["round2_full_design_df"] = full_design
        st.success(f"已生成 {len(full_design)} 行设计（种子固定为 42，同样的参数下次重跑结果完全一致）。")

    full_design = st.session_state.get("round2_full_design_df")
    if full_design is None or full_design.empty:
        return

    st.caption("各模块行数：" + "、".join(f"{PICHIA_RUN_TYPE_LABELS.get(rt, rt)} {count}" for rt, count in full_design["run_type"].value_counts().items()))

    excel_bytes = _pichia_round2_workbook_bytes(full_design)
    st.download_button(
        "下载 Round 2 设计表格（Excel）",
        data=excel_bytes,
        file_name="pichia_round2_design.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_round2_excel",
    )

    st.markdown("##### 上传已回填结果（可选）")
    uploaded = st.file_uploader(
        "上传已回填 yield_g_per_l / od600 的 CSV，或者上面下载的 Excel 填完后原样传回来",
        type=["csv", "xlsx"],
        key="round2_results_upload",
    )
    if uploaded is not None:
        try:
            payload = uploaded.getvalue()
            if uploaded.name.lower().endswith(".xlsx"):
                uploaded_df = pd.read_excel(BytesIO(payload))
            else:
                uploaded_df = pd.read_csv(BytesIO(payload))
            uploaded_df, replicate_spread, consistency_warnings = _pichia_remap_uploaded_columns(uploaded_df)
        except Exception as exc:
            st.error(f"读取失败：{exc}")
        else:
            missing = [column for column in ["run_id", *PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL] if column not in uploaded_df.columns]
            if missing:
                st.error(f"上传文件缺少字段：{', '.join(missing)}")
            else:
                _ensure_pichia_data_area()
                archive_name = Path(uploaded.name).name or "pichia_round2_uploaded.csv"
                (PICHIA_UPLOAD_DIR / archive_name).write_bytes(payload)
                st.session_state["round2_full_design_df"] = uploaded_df
                full_design = uploaded_df
                st.success("已加载并归档上传的 Round 2 结果。")
                for warning_text in consistency_warnings:
                    st.warning(warning_text)
                if not replicate_spread.empty:
                    sort_column = next((column for column in replicate_spread.columns if column.endswith("重复间差值")), None)
                    if sort_column:
                        replicate_spread = replicate_spread.sort_values(sort_column, ascending=False)
                    st.dataframe(replicate_spread, width="stretch", hide_index=True)

    with st.expander("🧪 开发/测试用：填入模拟数据", expanded=False):
        st.caption(
            "只填当前还是空的产量/OD600格子，编造的示意数值，不是真实实验结果——已经回填的真实数据不会被覆盖。"
            "用于在真实 Round 2 结果回来之前，先试一遍下面响应面拟合/交互检验/贝叶斯优化这几块的界面是否正常。"
        )
        confirm_simulate = st.checkbox("我知道这是假数据，只是用来测试界面", key="confirm_simulate_round2")
        # the explicit "and confirm_simulate" doesn't just duplicate disabled=
        # for show -- it's what actually gates the fill; disabled= only stops
        # a mouse click in a real browser, not a widget state set some other
        # way, and the one thing that must never happen is filling fake data
        # in without the checkbox truly being on.
        if st.button("填入模拟数据", key="simulate_round2_results", disabled=not confirm_simulate) and confirm_simulate:
            already_filled = int(full_design[PICHIA_TARGET_COL].notna().sum())
            full_design = _pichia_simulate_round2_results(full_design)
            st.session_state["round2_full_design_df"] = full_design
            newly_filled = int(full_design[PICHIA_TARGET_COL].notna().sum()) - already_filled
            st.success(f"已给 {newly_filled} 行空白结果填入模拟数据；原有 {already_filled} 行真实回填数据未改动。")

    st.markdown("##### 回填实测结果")
    entry_view = full_design[["run_id", PICHIA_TARGET_COL, PICHIA_OD_COL]].reset_index(drop=True)
    edited_entries = st.data_editor(
        entry_view,
        width="stretch",
        num_rows="fixed",
        hide_index=True,
        key="round2_data_editor",
        disabled=["run_id"],
        column_config={"run_id": "编号", PICHIA_TARGET_COL: "hLF产量", PICHIA_OD_COL: "收获时OD600"},
    )
    edited = full_design.copy()
    edited[PICHIA_TARGET_COL] = edited_entries[PICHIA_TARGET_COL].to_numpy()
    edited[PICHIA_OD_COL] = edited_entries[PICHIA_OD_COL].to_numpy()
    st.session_state["round2_full_design_df"] = edited

    numeric_edited = _pichia_numeric_results(edited)
    filled = int(numeric_edited[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1).sum())
    st.caption(f"已回填 {filled}/{len(edited)} 行完整结果。")

    if st.button("保存到 data/pichia/final/（供下一轮和历史查阅使用）", key="save_round2_to_final"):
        _ensure_pichia_data_area()
        edited.to_csv(PICHIA_ROUND2_DATASET_PATH, index=False, encoding="utf-8-sig")
        st.success(f"已保存到 {PICHIA_ROUND2_DATASET_PATH}")

def _pichia_response_surface_term_label(term: str) -> str:
    """Chinese-friendly label for a fitted response-surface term name
    ("ph" / "ph^2" / "ph*glucose_pct") -- reuses PICHIA_VARIABLE_LABELS so it
    stays in sync with every other Chinese label in this file rather than
    hand-maintaining a second translation table."""
    if term == "intercept":
        return "截距"
    if "*" in term:
        first, second = term.split("*")
        return f"{PICHIA_VARIABLE_LABELS.get(first, first)} × {PICHIA_VARIABLE_LABELS.get(second, second)}"
    if term.endswith("^2"):
        base = term[:-2]
        return f"{PICHIA_VARIABLE_LABELS.get(base, base)}²"
    return PICHIA_VARIABLE_LABELS.get(term, term)

def _pichia_significance_stars(p_value: float | None) -> str:
    """Standard significance-star convention (R/statsmodels-style) -- a
    single p-value column is precise but a binary 是/否 throws away exactly
    how close to the cutoff a term is; stars keep both the exact number and
    a quick-scan read."""
    if p_value is None or pd.isna(p_value):
        return ""
    if p_value < 0.001:
        return "***"
    if p_value < 0.01:
        return "**"
    if p_value < 0.05:
        return "*"
    if p_value < 0.1:
        return "·"
    return ""

# help text for st.dataframe(column_config=...) -- centralized here (rather
# than inline at each call site) since several tables below share the same
# jargon (p值/显著性/满意度 etc.) and a hover tooltip is the only place this
# explanation lives; keeping it in one dict makes it easy to keep the wording
# consistent across tables instead of drifting per call site.
PICHIA_COEFFICIENT_HELP: dict[str, str] = {
    "项": "拟合模型里的一项：截距、某个变量的一次项、二次项（²，弯曲/封顶效应），或两个变量的交互项（×）。",
    "系数": "这一项对预测产量的贡献方向和大小；正负号表示增加该变量取值是让预测产量升高还是降低，不能跨项直接比较大小（单位不同）。",
    "标准误": "这个系数估计值本身的不确定度——数值越大，说明换一批数据重新拟合，这个系数可能差很多，估计不稳。",
    "p值": "这一项是否只是噪声的统计检验：p 值越小，这一项真实存在（不是噪声）的把握越大；一般 p<0.05 认为显著。",
    "显著性": "p 值的快速参考：*** p<0.001，** p<0.01，* p<0.05，· p<0.1，空白=不显著（p≥0.1）。",
}


def _pichia_coefficient_column_config() -> dict[str, Any]:
    return {column: st.column_config.Column(help=text) for column, text in PICHIA_COEFFICIENT_HELP.items()}


def _pichia_coefficient_table(fit: Any) -> pd.DataFrame:
    rows = []
    for term in fit.term_names:
        stats_dict = fit.coefficient_significance.get(term)
        p_value = stats_dict["p_value"] if stats_dict else None
        rows.append(
            {
                "项": _pichia_response_surface_term_label(term),
                "系数": _num(fit.coefficients[term]),
                "标准误": _num(stats_dict["se"]) if stats_dict else None,
                "p值": _num(p_value) if stats_dict else None,
                "显著性": _pichia_significance_stars(p_value) if stats_dict else "样本不足",
            }
        )
    return pd.DataFrame(rows)

PICHIA_VERDICT_RENDERERS = {"success": st.success, "info": st.info, "warning": st.warning}

def _pichia_render_verdicts(verdicts: list[Any], translate: dict[str, str] | None = None) -> None:
    """Shared severity->st.* dispatch for any verdict list shaped like
    ResponseSurfaceVerdict/BoRecommendationVerdict (a .severity/.message
    pair). `translate`, when given, swaps bare「name」placeholders for
    Chinese labels -- round2_design.py's response-surface verdicts need
    this (it stays UI-agnostic, no label dict of its own); its BO verdicts
    don't, since summarize_bo_recommendation already writes fully Chinese
    messages with no placeholders to translate."""
    for verdict in verdicts:
        message = verdict.message
        if translate:
            for name, label in translate.items():
                message = message.replace(f"「{name}」", f"「{label}」")
        PICHIA_VERDICT_RENDERERS.get(verdict.severity, st.info)(message)

def _pichia_render_response_surface_verdicts(verdicts: list[Any]) -> None:
    _pichia_render_verdicts(verdicts, translate=PICHIA_VARIABLE_LABELS)

PICHIA_SENSITIVITY_HELP: dict[str, str] = {
    "变量": "该活跃变量。",
    "峰值点": "其余活跃变量固定在联合最优点时，单独扫这一个变量得到的预测产量最高点——K≥2 时一般很接近联合最优点，但不必完全相同。",
    "峰值预测产量": "峰值点对应的预测产量。",
    "容许范围(±5%)": "峰值点附近、预测产量仍不低于峰值 95% 的取值区间——区间越窄，说明这个变量需要控制得越精确才能拿到接近最优的产量；区间越宽，说明这个变量比较「皮实」，微调影响不大。",
    "范围宽度": "容许范围的绝对宽度（上限-下限），单位与该变量本身一致。",
    "占测试范围比例": "容许范围宽度 ÷ 本轮该变量实际测试范围的宽度，例如 0.3 表示容许范围只占本轮测试跨度的 30%。",
    "触及测试范围边界": "容许范围是否顶到了本轮实际测试范围的边缘——如果是，真实的容许范围可能比这里算出的更宽，只是被测试范围本身卡住了，不代表这个变量真的这么敏感。",
}

def _pichia_sensitivity_table(sensitivity: dict[str, Any]) -> pd.DataFrame:
    rows = []
    for variable, result in sensitivity.items():
        touches = []
        if result.touches_lower_bound:
            touches.append("下限")
        if result.touches_upper_bound:
            touches.append("上限")
        rows.append(
            {
                "变量": PICHIA_VARIABLE_LABELS.get(variable, variable),
                "峰值点": _num(result.peak_x),
                "峰值预测产量": _num(result.peak_value),
                "容许范围(±5%)": f"[{_num(result.plateau_low)}, {_num(result.plateau_high)}]",
                "范围宽度": _num(result.plateau_width),
                "占测试范围比例": _num(result.plateau_width_fraction),
                "触及测试范围边界": "、".join(touches) if touches else "否",
            }
        )
    return pd.DataFrame(rows)

_PICHIA_CANONICAL_LABELS: dict[str, str] = {
    "maximum": "真极大值（有明确的单一峰值）",
    "minimum": "真极小值（有明确的单一谷值——如果目标是最大化产量，这个方向不是好消息）",
    "saddle": "鞍点（沿一个方向升高、另一个方向降低，不是真正的「峰」）",
    "ridge": "岭线/脊线（至少有一个方向几乎没有曲率，可能有一整条线同样好，不存在单一最优点）",
    "flat": "近似平坦（拟合面在活跃变量范围内几乎没有曲率，模型给不出「哪里最优」的有效信息）",
}

def _pichia_canonical_classification_label(classification: str) -> str:
    return _PICHIA_CANONICAL_LABELS.get(classification, classification)

PICHIA_CANONICAL_HELP: dict[str, str] = {
    "分类": "经典响应面 canonical analysis 判别：不看网格搜索给出的「测试范围内最好的点」，只看拟合公式本身的曲率，判断这个曲面真实的形状是不是真的有峰。",
    "无约束理论最优点": "不考虑测试范围限制、只看拟合公式本身梯度为零的点——鞍点/岭线情形下这个点的意义有限，请配合下面两列一起看。",
    "该点在测试范围内": "理论最优点是否落在本轮实际测试范围内；「否」时它更多是数学参考点，不能直接当作可以验证的条件。",
    "结果可靠": "「否」表示分类是鞍点以外的退化情形（岭线/近似平坦），此时「理论最优点」不唯一或没有实际意义，不建议照搬这个点去验证。",
    "该点预测产量": "把理论最优点代入拟合模型算出的预测产量，同样只在「结果可靠」为「是」时才有直接参考价值。",
}

def _pichia_canonical_table(canonical: Any) -> pd.DataFrame:
    point_text = "、".join(
        f"{PICHIA_VARIABLE_LABELS.get(name, name)}={_num(value)}" for name, value in canonical.stationary_point.items()
    )
    row = {
        "分类": _pichia_canonical_classification_label(canonical.classification),
        "无约束理论最优点": point_text,
        "该点在测试范围内": "是" if canonical.stationary_point_in_tested_range else "否",
        "结果可靠": "是" if canonical.stationary_point_reliable else "否",
        "该点预测产量": _num(canonical.predicted_at_stationary_point),
    }
    return pd.DataFrame([row])

PICHIA_DESIRABILITY_HELP: dict[str, str] = {
    "方案": "两个候选点的对比：只看产量能拿到多少，vs. 同时兼顾产量和 OD600 可行性能拿到多少。",
    "预测产量": "该点代入产量响应面拟合算出的预测产量。",
    "预测OD600": "该点代入 OD600 响应面拟合算出的预测生长量；纯产量最优点这一行留空，因为选它时完全没考虑 OD600。",
    "满意度(0~1)": "Derringer-Suich 满意度：产量和 OD600 可行性各自打 0~1 分后取几何平均——只要有一项是 0（例如 OD600 完全不可行），总分就是 0，不会被另一项的高分「平均」掉。",
}

def _pichia_desirability_table(result: Any) -> pd.DataFrame:
    def _row(
        label: str, point: dict[str, float], yield_value: float, od_value: float | None, desirability: float | None
    ) -> dict[str, Any]:
        row: dict[str, Any] = {"方案": label}
        for name, value in point.items():
            row[PICHIA_VARIABLE_LABELS.get(name, name)] = _num(value)
        row["预测产量"] = _num(yield_value)
        # None (not a "-" placeholder string) so the column stays a clean
        # float dtype -- mixing floats and a literal placeholder string in
        # one column makes it an "object" dtype, which pyarrow's Streamlit
        # serialization only handles via a slow, warning-logging fallback
        # path; a plain missing value renders as an empty cell either way.
        row["预测OD600"] = _num(od_value) if od_value is not None else None
        row["满意度(0~1)"] = _num(desirability) if desirability is not None else None
        return row

    rows = [
        _row("纯产量最优点（不考虑OD600）", result.pure_yield_optimum, result.pure_yield_optimum_value, None, None),
        _row("联合满意度最优点", result.point, result.predicted_yield, result.predicted_od600, result.composite_desirability),
    ]
    return pd.DataFrame(rows)

def _pichia_desirability_column_config(result: Any) -> dict[str, Any]:
    config = dict(PICHIA_DESIRABILITY_HELP)
    for name in result.point:
        label = PICHIA_VARIABLE_LABELS.get(name, name)
        config.setdefault(label, f"「{label}」在该方案下的取值。")
    return {column: st.column_config.Column(help=text) for column, text in config.items()}

def _pichia_render_response_surface_deep_dive(
    fit: Any,
    ccd_rows: pd.DataFrame,
    case: str,
    od_fit: Any | None,
    od_threshold: float | None,
) -> None:
    """The three deeper analyses beyond the optimum's own CI (already shown
    inline on the optimum table): sensitivity/plateau width, canonical
    (saddle/ridge/peak) classification, and joint yield+OD600 desirability.
    Framed differently depending on `case` (classify_response_surface_case's
    output, same priority order as summarize_response_surface's own verdict)
    since the same numbers mean different things depending on whether the
    fit/data are already trustworthy or not -- e.g. a narrow sensitivity
    plateau is reassuring in the "normal" case but beside the point when the
    model has significant lack-of-fit in the first place."""
    if case == "lack_of_fit":
        st.warning(
            "⚠️ 存在显著失拟：下面的灵敏度/鞍点判别/联合优化都基于当前这个（形式可能不对的）二次模型，"
            "结论只能参考，不能替代重新检查模型形式。"
        )

    with st.expander("灵敏度分析（最优点附近多「皮实」）", expanded=False):
        st.caption("固定其余活跃变量在联合最优点，单独扫一个变量：预测产量掉到峰值 95% 以下之前，这个变量还有多大的浮动空间。")
        sensitivity = sensitivity_analysis(fit, ccd_rows)
        st.dataframe(
            _pichia_sensitivity_table(sensitivity),
            width="stretch",
            hide_index=True,
            column_config={column: st.column_config.Column(help=text) for column, text in PICHIA_SENSITIVITY_HELP.items()},
        )
        if case == "boundary" and any(result.touches_lower_bound or result.touches_upper_bound for result in sensitivity.values()):
            st.caption("留意「触及测试范围边界」为「是」的行——这些变量的真实容许范围可能比表里算出的更宽，只是被本轮测试范围卡住了。")

    with st.expander("鞍点/岭线判别（曲面到底是不是真的有峰）", expanded=(case == "boundary")):
        st.caption("经典响应面 canonical analysis：抛开网格搜索给出的「测试范围内最好的点」，只看拟合公式本身的曲率，判断曲面真实的形状。")
        canonical = canonical_analysis(fit, ccd_rows)
        st.dataframe(
            _pichia_canonical_table(canonical),
            width="stretch",
            hide_index=True,
            column_config={column: st.column_config.Column(help=text) for column, text in PICHIA_CANONICAL_HELP.items()},
        )
        if case == "boundary":
            if canonical.classification in ("ridge", "flat"):
                st.info(
                    "这和上面「最优点卡边界」的现象是一致的：曲面在活跃变量的某个方向上几乎没有曲率（呈岭线/近似平坦），"
                    "网格搜索自然找不到真正的封顶，只能停在测试范围的边缘——不是这个点真的最优，是没有曲率能定出唯一最优点。"
                )
            elif not canonical.stationary_point_in_tested_range:
                point_text = "、".join(
                    f"{PICHIA_VARIABLE_LABELS.get(name, name)}={_num(value)}" for name, value in canonical.stationary_point.items()
                )
                st.info(f"无约束理论最优点在 {point_text}，已经超出本轮测试范围——和「最优点卡边界」的现象一致，建议下一轮把测试范围扩大到能覆盖这个点。")
        elif case == "normal" and canonical.classification == "maximum" and canonical.stationary_point_in_tested_range:
            st.caption("曲面呈真极大值、理论最优点也落在测试范围内——和「一切正常」的结论一致，这个联合最优点值得安排验证批次。")

    if od_fit is not None and od_threshold is not None:
        with st.expander("产量 + OD600 联合优化（满意度）", expanded=(case == "od_infeasible")):
            st.caption("不把 OD600 当成硬性的「是/否」过滤，而是把产量和 OD600 可行性都打分（0~1），找同时兼顾两者的折中点，并和「只看产量」的最优点对比。")
            desirability = optimize_joint_desirability(fit, od_fit, ccd_rows, od_threshold)
            st.dataframe(
                _pichia_desirability_table(desirability),
                width="stretch",
                hide_index=True,
                column_config=_pichia_desirability_column_config(desirability),
            )
            if case == "od_infeasible":
                yield_cost = (
                    1.0 - desirability.predicted_yield / desirability.pure_yield_optimum_value
                    if desirability.pure_yield_optimum_value
                    else 0.0
                )
                feasibility_note = "可行" if desirability.od_desirability >= 1.0 - 1e-6 else "更接近可行（仍未完全达标）"
                st.success(
                    f"这就是「折中点」：把最优条件从纯产量最优点换成联合满意度最优点，预测产量下降约 {yield_cost:.1%}，"
                    f"换来预测 OD600 从不可行变为{feasibility_note}——建议下一轮把这个点也纳入验证批次，而不是只验证纯产量最优点。"
                )
            elif desirability.od_desirability >= 1.0 - 1e-6:
                st.caption("纯产量最优点本身已经满足 OD600 可行性，联合满意度最优点和它基本重合，不存在实质性权衡。")

    if case == "normal":
        st.caption("以上分析（含上方最优点的置信区间）互相印证：模型形式没问题、最优点不在边界、（若已检验）OD600 可行——建议按前面「下一步建议」安排验证批次。")

def _pichia_predicted_vs_actual_chart(
    actual: np.ndarray,
    predicted: np.ndarray,
    *,
    title: str,
    x_title: str,
    y_title: str,
    point_name: str,
    hovertemplate: str,
    height: int = 340,
    text: pd.Series | None = None,
    error_y: np.ndarray | None = None,
) -> go.Figure:
    """Shared predicted-vs-actual scatter + y=x reference line -- the same
    diagnostic shape (a fitted model's own predictions on points it's
    evaluated against, however those predictions were obtained) whether
    the caller is the CCD fit's plain residual chart or the GP leave-one-out
    CV's held-out-prediction chart; only the labels/data/error bars differ."""
    lo = float(min(actual.min(), predicted.min()))
    hi = float(max(actual.max(), predicted.max()))
    fig = go.Figure()
    fig.add_trace(
        go.Scatter(x=[lo, hi], y=[lo, hi], mode="lines", line=dict(color=PICHIA_MUTED_COLOR, dash="dash"), name="y=x", hoverinfo="skip")
    )
    fig.add_trace(
        go.Scatter(
            x=actual,
            y=predicted,
            mode="markers",
            marker=dict(color=PICHIA_ACCENT_COLOR, size=8),
            text=text,
            error_y=dict(type="data", array=error_y, visible=True, color=PICHIA_MUTED_COLOR) if error_y is not None else None,
            hovertemplate=hovertemplate,
            name=point_name,
        )
    )
    fig.update_layout(
        title=title,
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title=x_title,
        yaxis_title=y_title,
        margin=dict(t=50, b=40, l=60, r=30),
        height=height,
    )
    return fig

def _pichia_response_surface_residual_chart(fit: Any, df: pd.DataFrame) -> go.Figure:
    predicted = evaluate_response_surface(fit, df)
    actual = df[PICHIA_TARGET_COL].to_numpy(dtype=float)
    return _pichia_predicted_vs_actual_chart(
        actual,
        predicted,
        title="预测值 vs 实测值（残差诊断，越贴近虚线越好）",
        x_title="实测产量",
        y_title="预测产量",
        point_name="CCD 样本",
        hovertemplate="%{text}<br>实测: %{x:.4g}<br>预测: %{y:.4g}<extra></extra>",
        text=df["run_id"] if "run_id" in df.columns else None,
    )

def _pichia_response_surface_curve_chart(fit: Any, df: pd.DataFrame, variable: str) -> go.Figure:
    x_values = np.linspace(float(df[variable].min()), float(df[variable].max()), 61)
    predicted = evaluate_response_surface(fit, pd.DataFrame({variable: x_values}))
    fig = go.Figure()
    fig.add_trace(go.Scatter(x=x_values, y=predicted, mode="lines", line=dict(color=PICHIA_ACCENT_COLOR, width=2), name="拟合曲线"))
    fig.add_trace(
        go.Scatter(
            x=df[variable], y=df[PICHIA_TARGET_COL], mode="markers", marker=dict(color="#e6e4da", size=8), name="实测点"
        )
    )
    fig.update_layout(
        title=f"{PICHIA_VARIABLE_LABELS.get(variable, variable)} 响应曲线",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title=PICHIA_VARIABLE_LABELS.get(variable, variable),
        yaxis_title="产量",
        margin=dict(t=50, b=40, l=60, r=30),
        height=340,
    )
    return fig

def _pichia_response_surface_3d_chart(
    fit: Any, df: pd.DataFrame, x_variable: str, y_variable: str, colorscale: list[list[Any]]
) -> go.Figure:
    """3D twin of _pichia_response_surface_contour_chart, same underlying
    response_surface_grid data -- a surface gives the curvature/peak shape
    away at a glance (is it a sharp peak, a broad plateau, a ridge?), which a
    2D contour states just as correctly but less immediately; shown side by
    side with the contour rather than replacing it, since reading exact
    values back off a rotated 3D surface is much harder than off a contour's
    color scale and axis lines."""
    grid = response_surface_grid(fit, df, x_variable, y_variable)
    x_label = PICHIA_VARIABLE_LABELS.get(x_variable, x_variable)
    y_label = PICHIA_VARIABLE_LABELS.get(y_variable, y_variable)
    fig = go.Figure(
        go.Surface(
            x=grid["x_values"],
            y=grid["y_values"],
            z=grid["z"],
            colorscale=colorscale,
            showscale=False,
            opacity=0.92,
            hovertemplate=f"{x_label}: %{{x:.3g}}<br>{y_label}: %{{y:.3g}}<br>预测产量: %{{z:.4g}}<extra></extra>",
        )
    )
    fig.add_trace(
        go.Scatter3d(
            x=df[x_variable],
            y=df[y_variable],
            z=df[PICHIA_TARGET_COL],
            mode="markers",
            marker=dict(color="#e6e4da", size=4),
            name="实测点",
        )
    )
    fig.add_trace(
        go.Scatter3d(
            x=[fit.optimum[x_variable]],
            y=[fit.optimum[y_variable]],
            z=[fit.predicted_optimum],
            mode="markers",
            marker=dict(color="#e66767", size=6, symbol="diamond"),
            name="联合最优点",
        )
    )
    fig.update_layout(
        title=f"{x_label} × {y_label}（3D）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        scene=dict(
            xaxis_title=x_label,
            yaxis_title=y_label,
            zaxis_title="预测产量",
            bgcolor="rgba(0,0,0,0)",
        ),
        margin=dict(t=50, b=10, l=10, r=10),
        height=420,
    )
    return fig

def _pichia_response_surface_contour_chart(
    fit: Any,
    df: pd.DataFrame,
    x_variable: str,
    y_variable: str,
    colorscale: list[list[Any]],
    od_fit: Any | None = None,
    od_threshold: float | None = None,
) -> go.Figure:
    grid = response_surface_grid(fit, df, x_variable, y_variable)
    x_label = PICHIA_VARIABLE_LABELS.get(x_variable, x_variable)
    y_label = PICHIA_VARIABLE_LABELS.get(y_variable, y_variable)
    fig = go.Figure(
        go.Contour(
            x=grid["x_values"],
            y=grid["y_values"],
            z=grid["z"],
            colorscale=colorscale,
            colorbar=dict(title="预测产量"),
            hovertemplate=f"{x_label}: %{{x:.3g}}<br>{y_label}: %{{y:.3g}}<br>预测产量: %{{z:.4g}}<extra></extra>",
        )
    )
    if od_fit is not None and od_threshold is not None:
        # binarize OD600's own fitted surface at the feasibility threshold and
        # contour *that* at level 0.5 -- draws exactly the feasible/infeasible
        # boundary regardless of OD600's actual value range, rather than
        # fighting Plotly's contour-level picking to land a line precisely on
        # od_threshold in the raw OD600 units.
        od_grid = response_surface_grid(od_fit, df, x_variable, y_variable, fixed_at=grid["fixed_at"])
        feasible_mask = (od_grid["z"] >= od_threshold).astype(float)
        fig.add_trace(
            go.Contour(
                x=od_grid["x_values"],
                y=od_grid["y_values"],
                z=feasible_mask,
                contours=dict(start=0.5, end=0.5, size=0.5, coloring="lines"),
                line=dict(color="#ffffff", width=2.5, dash="dot"),
                showscale=False,
                name=f"OD600≥{od_threshold:.3g} 可行边界",
                hovertemplate=f"OD600 可行边界（阈值 {od_threshold:.3g}）<extra></extra>",
            )
        )
    fig.add_trace(
        go.Scatter(
            x=df[x_variable],
            y=df[y_variable],
            mode="markers",
            marker=dict(color="#e6e4da", size=7, symbol="x"),
            name="实测点",
            hoverinfo="skip",
        )
    )
    title = f"{x_label} × {y_label}"
    other_note = "、".join(f"{PICHIA_VARIABLE_LABELS.get(name, name)}={value:g}" for name, value in grid["fixed_at"].items())
    if other_note:
        title += f"（其余变量固定：{other_note}）"
    fig.update_layout(
        title=title,
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title=x_label,
        yaxis_title=y_label,
        margin=dict(t=60, b=40, l=60, r=30),
        height=380,
    )
    return fig

def _pichia_gp_pdp_chart(model: Any, feature_cols: list[str], variable: str, anchor: dict[str, Any]) -> go.Figure:
    """Partial-dependence read of an already-fitted BO GP (see
    round2_design.gp_partial_dependence -- never refits): sweeps `variable`
    across its full original bounds holding every other feature at anchor
    (the recommendation this chart is attached to), so the shaded band shows
    where the GP is actually confident (near training points) versus
    guessing (gaps between them)."""
    lower, upper = PICHIA_CONTINUOUS_BOUNDS[variable]
    pdp = gp_partial_dependence(model, feature_cols, variable, anchor, lower=lower, upper=upper, resolution=41)
    x_values = list(pdp["x"])
    mean = pdp["mean"]
    band = 1.96 * pdp["std"]  # ~95% band under the GP's own Gaussian posterior
    upper_band = list(mean + band)
    lower_band = list(mean - band)

    fig = go.Figure()
    fig.add_trace(
        go.Scatter(
            x=x_values + x_values[::-1],
            y=upper_band + lower_band[::-1],
            fill="toself",
            fillcolor="rgba(57,135,229,0.2)",
            line=dict(color="rgba(0,0,0,0)"),
            name="95% 置信区间",
            hoverinfo="skip",
        )
    )
    fig.add_trace(go.Scatter(x=x_values, y=mean, mode="lines", line=dict(color=PICHIA_ACCENT_COLOR, width=2), name="GP 预测均值"))
    if variable in anchor:
        fig.add_trace(
            go.Scatter(
                x=[anchor[variable]],
                y=[anchor.get("predicted_yield")],
                mode="markers",
                marker=dict(color="#e66767", size=11, symbol="star"),
                name="当前推荐点",
            )
        )
    fig.update_layout(
        title=f"{PICHIA_VARIABLE_LABELS.get(variable, variable)} 偏依赖（其余变量固定在推荐点）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title=PICHIA_VARIABLE_LABELS.get(variable, variable),
        yaxis_title="预测产量",
        margin=dict(t=50, b=40, l=60, r=30),
        height=340,
    )
    return fig

def _pichia_render_bo_gp_pdp(bo_result: dict[str, Any], active_variables: list[str]) -> None:
    model = bo_result.get("yield_model")
    feature_cols = bo_result.get("feature_cols")
    if model is None or not feature_cols or not active_variables:
        return
    anchor = bo_result["recommendations"][0]
    with st.expander("GP 偏依赖图（模型在没测过的区域有多不确定）", expanded=False):
        st.caption("阴影是 GP 自己算出的 95% 置信区间，不是另外估的——区间越宽，说明这一段离训练数据越远、模型越没底。")
        pdp_cols = st.columns(2)
        for index, variable in enumerate(active_variables):
            with pdp_cols[index % 2]:
                st.plotly_chart(_pichia_gp_pdp_chart(model, feature_cols, variable, anchor), width="stretch")

def _pichia_render_bo_verdicts(bo_result: dict[str, Any], cv_result: dict[str, Any] | None) -> None:
    """Narrative read of a BO recommendation batch (summarize_bo_recommendation's
    output), analogous to _pichia_render_response_surface_verdicts for the CCD
    fit -- unlike that one, the messages here are already fully Chinese (no
    bare variable-name placeholders to translate), since round2_design.py
    already writes them that way for this function (see its own docstring)."""
    yield_cv = cv_result["yield"] if cv_result else None
    od_cv = cv_result["od"] if cv_result else None
    verdicts = summarize_bo_recommendation(bo_result, yield_cv=yield_cv, od_cv=od_cv)
    _pichia_render_verdicts(verdicts)

def _pichia_bo_cv_residual_chart(cv: dict[str, Any], label: str) -> go.Figure:
    return _pichia_predicted_vs_actual_chart(
        cv["actual"],
        cv["predicted"],
        title=f"{label}：留一法预测 vs 实测（误差线是 GP 自己的预测标准差）",
        x_title="实测",
        y_title="留一法预测",
        point_name="留一法预测",
        hovertemplate="实测: %{x:.4g}<br>留一法预测: %{y:.4g}<extra></extra>",
        height=320,
        error_y=cv["predicted_std"],
    )

PICHIA_BO_CV_HELP: dict[str, str] = {
    "Q²": "留一法交叉验证的预测能力：每次把一个样本从训练集里拿掉、用剩下的重新拟合再预测这个点，和实测值比较算出的类 R² 指标。"
    "Q²≥0.5 一般认为有实用的预测能力；Q²<0 说明比直接猜训练集均值还差，模型目前不可信。",
    "留一法 RMSE": "留一法预测误差的均方根，单位和该指标本身一致——越小说明模型对没见过的新条件预测得越准。",
}

PICHIA_BO_CV_TARGETS: list[tuple[str, str]] = [("产量", "yield"), ("OD600", "od")]

def _pichia_render_bo_cv_result(cv_result: dict[str, Any]) -> None:
    cols = st.columns(2)
    for col, (label, key) in zip(cols, PICHIA_BO_CV_TARGETS):
        cv = cv_result[key]
        with col:
            st.metric("Q²", _num(cv["q_squared"]), help=PICHIA_BO_CV_HELP["Q²"])
            st.metric("留一法 RMSE", _num(cv["rmse"]), help=PICHIA_BO_CV_HELP["留一法 RMSE"])
            st.plotly_chart(_pichia_bo_cv_residual_chart(cv, label), width="stretch")

def _pichia_bo_cv_training_rows(train_df: pd.DataFrame) -> pd.DataFrame:
    """Rows with both target columns present -- the same row selection
    recommend_round2_bo_batch itself uses to train yield_model/od_model.
    Cross-validation must filter on both, not just whichever single target
    it's validating this call: whenever backfill is partial (yield filled
    but OD600 not, or vice versa), filtering on only one column would
    validate against a different, larger row set than the one that
    actually produced the deployed model."""
    return train_df.dropna(subset=[PICHIA_TARGET_COL, PICHIA_OD_COL])

def _pichia_render_bo_recommendation_section(
    bo_result: dict[str, Any],
    active_variables: list[str],
    train_df: pd.DataFrame,
    cv_session_key: str,
) -> None:
    """Recommendation table + narrative read + GP partial-dependence + an
    on-demand leave-one-out cross-validation -- shared by both BO entry
    points (Round-1-only and the combined Round1+Round2 dataset), which
    otherwise built an identical table from an identical result-dict shape.
    Consolidating this avoids the two call sites silently drifting apart on
    formatting once there's new shared content (the verdicts/CV) to add."""
    rows = []
    for rec in bo_result["recommendations"]:
        row = _pichia_variable_display(rec)
        row["预测产量"] = _num(rec.get("predicted_yield"))
        row["预测产量标准差"] = _num(rec.get("predicted_yield_std"))
        row["预测 OD600"] = _num(rec.get("predicted_od600"))
        rows.append(row)
    st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)

    with st.expander("模型校验（留一法交叉验证，判断这个 GP 能信到什么程度）", expanded=False):
        st.caption(
            "对训练数据逐个样本做「留一法」：把这个样本从训练集里拿掉，用剩下的重新拟合 GP，预测被拿掉的这个点，"
            "再和它的实测值比较——这样得到的误差才是真正的「预测」误差，不是模型看过这个点之后回头报的拟合误差。"
            "样本量不大，但每次点击都会重新拟合多次模型，需要几秒到几十秒，不会自动运行。"
        )
        if st.button("运行交叉验证", key=f"{cv_session_key}_run_button"):
            cv_train_df = _pichia_bo_cv_training_rows(train_df)
            try:
                yield_cv = gp_leave_one_out_cv(cv_train_df, bo_result["feature_cols"], PICHIA_TARGET_COL)
                od_cv = gp_leave_one_out_cv(cv_train_df, bo_result["feature_cols"], PICHIA_OD_COL)
            except ImportError:
                st.warning("当前环境未安装 torch/botorch/gpytorch，无法运行交叉验证。")
            except ValueError as exc:
                st.warning(f"交叉验证暂时无法运行：{exc}")
            except Exception as exc:
                st.warning(f"交叉验证运行失败（样本量很小时，某一折的模型拟合可能数值不稳定）：{exc}")
            else:
                st.session_state[cv_session_key] = {"yield": yield_cv, "od": od_cv}

        # read session_state here, after the button block above may have
        # just written to it -- lets a fresh result show immediately on the
        # same click without needing st.rerun() (which would otherwise
        # re-execute this whole page, including refitting the CCD model for
        # the combined-data call site, just to re-read a variable that was
        # bound too early).
        cv_result = st.session_state.get(cv_session_key)
        if cv_result:
            _pichia_render_bo_cv_result(cv_result)

    _pichia_render_bo_verdicts(bo_result, st.session_state.get(cv_session_key))
    _pichia_render_bo_gp_pdp(bo_result, active_variables)

def _pichia_round2_results_analysis_section(plan: Round2Plan, round1_df: pd.DataFrame) -> None:
    """What becomes analyzable once the Round 2 sheet above is actually
    backfilled -- generating that sheet only produces conditions to test, not
    conclusions, and without this section there is nowhere the CCD block's
    fitted response surface, the interval-interaction verdict, or a BO
    recommendation informed by the combined dataset would ever show up."""
    full_design = st.session_state.get("round2_full_design_df")
    if full_design is None or full_design.empty:
        return
    numeric_full = _pichia_numeric_results(full_design)
    filled_mask = numeric_full[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1)
    if not filled_mask.any():
        return

    st.markdown("#### Round 2 结果分析")
    filled_design = numeric_full.loc[filled_mask]

    if plan.active_variables:
        st.markdown("##### 响应面 (CCD) 拟合")
        ccd_rows = filled_design[filled_design["run_type"] == "ccd"]
        try:
            fit = fit_ccd_response_surface(ccd_rows, plan.active_variables)
        except ValueError as exc:
            st.info(f"CCD 拟合暂时无法运行：{exc}")
        else:
            od_fit: Any | None = None
            try:
                od_fit = fit_ccd_response_surface(ccd_rows, plan.active_variables, target_col=PICHIA_OD_COL)
            except ValueError:
                pass  # OD600 fit is a supplementary feasibility overlay -- yield analysis still stands without it
            od_threshold = plan.od_threshold["threshold"]

            fit_cols = st.columns(3)
            fit_cols[0].metric(
                "R²", _num(fit.r_squared), help="拟合优度：模型解释了多少比例的产量变化，1.0=完美拟合，越低说明二次模型解释力越弱。"
            )
            fit_cols[1].metric(
                "样本数 / 参数数",
                f"{fit.n_points} / {fit.n_params}",
                help="用于拟合的 CCD 样本数，和模型需要估计的参数个数（截距+一次项+二次项+交互项）；样本数远大于参数数时，估计才稳。",
            )
            if fit.lack_of_fit is None:
                fit_cols[2].metric("失拟检验", "样本不足，无法判断")
            else:
                verdict = "存在失拟" if fit.lack_of_fit["significant_lack_of_fit"] else "无显著失拟"
                fit_cols[2].metric(
                    "失拟检验",
                    verdict,
                    help=f"p={_num(fit.lack_of_fit['p_value'])}。检验「二次模型这个形式对不对」，和某一项系数是否显著是两件独立的事。",
                )

            optimum_point = {**plan.fixed_values, **fit.optimum}
            optimum_display = {PICHIA_VARIABLE_LABELS.get(name, name): _num(value) for name, value in optimum_point.items()}
            optimum_display["预测产量"] = _num(fit.predicted_optimum)
            optimum_ci = predict_with_confidence_interval(fit, pd.DataFrame([fit.optimum]))
            optimum_display["预测产量 95% CI"] = f"[{_num(optimum_ci.loc[0, 'ci_low'])}, {_num(optimum_ci.loc[0, 'ci_high'])}]"
            if od_fit is not None:
                predicted_od_at_optimum = float(evaluate_response_surface(od_fit, pd.DataFrame([fit.optimum]))[0])
                optimum_display["预测OD600"] = _num(predicted_od_at_optimum)
                optimum_display["OD600可行"] = "是" if predicted_od_at_optimum >= od_threshold else "否"
            optimum_help: dict[str, str] = {
                PICHIA_VARIABLE_LABELS.get(name, name): f"响应面拟合在本轮实际测试范围内搜索得到的「{PICHIA_VARIABLE_LABELS.get(name, name)}」最优取值。"
                for name in optimum_point
            }
            optimum_help["预测产量"] = "把最优点代入拟合模型算出的预测产量——模型估计值，不是实测值。"
            optimum_help["预测产量 95% CI"] = (
                "这个预测值本身的 95% 置信区间：区间越宽，说明这个「最优点」的产量估计越不确定，"
                "换一批数据重新做同样的实验，结果可能有明显差异；区间很窄才说明这个数字站得住。"
            )
            if od_fit is not None:
                optimum_help["预测OD600"] = "把最优点代入 OD600 的响应面拟合算出的预测生长量，同样是模型估计值。"
                optimum_help["OD600可行"] = "预测OD600 是否达到可行阈值；「否」表示这个产量最优点在生长可行性上不成立，不能直接采用。"
            st.dataframe(
                pd.DataFrame([optimum_display]),
                width="stretch",
                hide_index=True,
                column_config={col: st.column_config.Column(help=text) for col, text in optimum_help.items()},
            )
            st.caption("预测最优点的搜索范围限制在 CCD 实际测试过的区间内，不外推到 round 1 原始边界之外。")

            _pichia_render_response_surface_verdicts(
                summarize_response_surface(fit, ccd_rows, od_fit=od_fit, od_threshold=od_threshold if od_fit is not None else None)
            )

            with st.expander("系数明细（每一项是否显著）", expanded=False):
                st.caption(
                    "失拟检验看的是「模型整体形式对不对」；这里的显著性看的是「这一项本身是不是噪声」，两者是独立的两件事。"
                    "显著性：*** p<0.001，** p<0.01，* p<0.05，· p<0.1，空白=不显著——星号只是 p 值的快速参考，具体数值看 p 值列。"
                )
                st.dataframe(
                    _pichia_coefficient_table(fit),
                    width="stretch",
                    hide_index=True,
                    column_config=_pichia_coefficient_column_config(),
                )

            st.plotly_chart(_pichia_response_surface_residual_chart(fit, ccd_rows), width="stretch")

            if len(plan.active_variables) == 1:
                st.plotly_chart(
                    _pichia_response_surface_curve_chart(fit, ccd_rows, plan.active_variables[0]), width="stretch"
                )
            else:
                st.caption(
                    "每组图固定其余活跃变量在预测最优点，○是 CCD 实测点，等高线图上的×同理，红色菱形/星标是模型给出的联合最优点；"
                    + ("白色虚线内是预测 OD600 达到可行阈值的区域。" if od_fit is not None else "OD600 拟合暂时无法运行，没有可行边界线。")
                )
                colorscale_name = st.selectbox(
                    "曲面配色",
                    list(PICHIA_SEQUENTIAL_COLORSCALE_OPTIONS),
                    index=list(PICHIA_SEQUENTIAL_COLORSCALE_OPTIONS).index(PICHIA_SEQUENTIAL_COLORSCALE_DEFAULT),
                    help="推荐选项在保持深色主题蓝调的前提下拉宽了明度跨度，比原配色更容易分清相邻等高线档位；"
                    "越靠后的选项辨识度越高，但和界面主题的贴合度也越低。",
                    key="pichia_response_surface_colorscale",
                )
                colorscale = PICHIA_SEQUENTIAL_COLORSCALE_OPTIONS[colorscale_name]
                pairs = list(itertools.combinations(plan.active_variables, 2))
                for x_variable, y_variable in pairs:
                    view_cols = st.columns(2)
                    with view_cols[0]:
                        st.plotly_chart(
                            _pichia_response_surface_3d_chart(fit, ccd_rows, x_variable, y_variable, colorscale),
                            width="stretch",
                        )
                    with view_cols[1]:
                        st.plotly_chart(
                            _pichia_response_surface_contour_chart(
                                fit, ccd_rows, x_variable, y_variable, colorscale, od_fit=od_fit, od_threshold=od_threshold
                            ),
                            width="stretch",
                        )

            st.markdown("##### 深入分析")
            case = classify_response_surface_case(
                fit, ccd_rows, od_fit=od_fit, od_threshold=od_threshold if od_fit is not None else None
            )
            _pichia_render_response_surface_deep_dive(fit, ccd_rows, case, od_fit, od_threshold if od_fit is not None else None)
    else:
        st.caption("本轮没有活跃变量进入响应面，跳过 CCD 拟合。")

    interaction_rows = full_design[full_design["run_type"] == "interval_interaction"]
    if not interaction_rows.empty:
        st.markdown("##### 补料间隔交互检验")
        interaction_variable = interaction_rows["changed_variable"].iloc[0]
        extra_interval_level = float(interaction_rows["interval_h"].iloc[0])
        interaction_levels = sorted(interaction_rows[interaction_variable].dropna().unique().tolist())
        result = analyze_interval_interaction(
            full_design, round1_df, PICHIA_ROUND1_BASELINE, interaction_variable, extra_interval_level, interaction_levels
        )
        label = PICHIA_VARIABLE_LABELS.get(interaction_variable, interaction_variable)
        if result["noise_sd"] is None:
            st.info(f"新区域噪声参考样本不足（{result['noise_n']} 个），暂时无法判断显著性，仅展示效应量。")
        rows = [
            {
                "对比": f"{label}={result['low_level']:g}",
                "新间隔-旧间隔差值": _num(result["effect_at_low"]),
                "显著": {True: "是", False: "否", None: "样本不足"}[result["interval_effect_significant_at_low"]],
            },
            {
                "对比": f"{label}={result['high_level']:g}",
                "新间隔-旧间隔差值": _num(result["effect_at_high"]),
                "显著": {True: "是", False: "否", None: "样本不足"}[result["interval_effect_significant_at_high"]],
            },
            {
                "对比": "交互效应（两者差值）",
                "新间隔-旧间隔差值": _num(result["interaction_effect"]),
                "显著": {True: "是", False: "否", None: "样本不足"}[result["interaction_significant"]],
            },
        ]
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)
        if result["interaction_significant"] is True:
            st.caption(f"「补料间隔 {extra_interval_level:g}h」的效应随「{label}」水平变化——不是一个固定的主效应，用它时要连着「{label}」一起看。")
        elif result["interaction_significant"] is False:
            st.caption(f"「补料间隔 {extra_interval_level:g}h」的效应在两个「{label}」水平下差不多，可以当一个不依赖「{label}」的主效应来看。")

    st.markdown("##### 合并数据后的贝叶斯优化建议")
    st.caption("用 Round 1 + 本轮已回填的 Round 2 数据一起重新训练 GP——数据量和覆盖面都比只用 Round 1 时更完整。")
    combined_columns = ["run_id", *PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL]
    combined_df = pd.concat(
        [round1_df[combined_columns], filled_design[combined_columns]],
        ignore_index=True,
    )
    n_batch_combined = st.slider("建议批次大小", min_value=3, max_value=15, value=9, key="round2_combined_bo_batch_size")
    if st.button("用合并数据生成贝叶斯优化建议", type="primary", key="run_combined_bo"):
        try:
            bo_result = recommend_round2_bo_batch(
                combined_df,
                fixed_values=plan.fixed_values,
                active_variables=plan.active_variables or list(PICHIA_CONTINUOUS_BOUNDS),
                od_threshold=plan.od_threshold["threshold"],
                n_batch=int(n_batch_combined),
            )
        except ImportError:
            st.warning("当前环境未安装 torch/botorch/gpytorch，无法运行贝叶斯优化建议。")
        except ValueError as exc:
            st.warning(f"未生成建议：{exc}")
        except Exception as exc:
            st.warning(f"贝叶斯优化建议生成失败（样本量很小时，GP 拟合可能数值不稳定）：{exc}")
        else:
            st.session_state["round2_combined_bo_result"] = bo_result
            st.success(
                f"已用 {len(combined_df)} 行合并数据生成 {len(bo_result['recommendations'])} 个建议"
                f"（候选池 {bo_result['n_candidates']} 个，满足约束 {bo_result['n_feasible']} 个）。"
            )

    combined_bo_result = st.session_state.get("round2_combined_bo_result")
    if combined_bo_result:
        _pichia_render_bo_recommendation_section(
            combined_bo_result, plan.active_variables, combined_df, "round2_combined_bo_cv_result"
        )

def _pichia_round2_tab() -> None:
    st.markdown("### Round 2：显著性分析 → 响应面（CCD）+ 约束贝叶斯优化")

    results_df = st.session_state.get("round1_results_df")
    if results_df is None or results_df.empty:
        st.info("请先在「Round 1：实验设计」页签生成设计并回填实测结果。")
        return

    run_df = _pichia_numeric_results(results_df)
    filled = int(run_df[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1).sum())
    st.caption(f"当前已回填 {filled}/{len(run_df)} 行完整结果。")
    if filled < len(run_df):
        st.warning("还有行没有回填完整结果；下面的分析只会用到已经有产量和 OD600 的行，结论可能随着数据补全而变化。")

    with st.expander("分析参数", expanded=False):
        param_cols = st.columns(3)
        with param_cols[0]:
            max_active_variables = st.number_input(
                "最多活跃变量数 (K 上限)",
                min_value=1,
                max_value=6,
                value=3,
                help="超过这个数量的显著变量会按效应量排序，排不进的先固定在 round 1 最优水平。",
                key="round2_max_active_variables",
            )
        with param_cols[1]:
            ccd_step_fraction = st.slider(
                "CCD 步长比例",
                min_value=0.1,
                max_value=1.0,
                value=0.5,
                step=0.05,
                help="响应面设计的轴向步长 = 该比例 x 单变量法(OFAT)的水平间距。",
                key="round2_ccd_step_fraction",
            )
        with param_cols[2]:
            od_threshold_fraction = st.slider(
                "OD600 约束比例",
                min_value=0.1,
                max_value=1.0,
                value=0.7,
                step=0.05,
                help="OD600 约束阈值 = 该比例 x 基线 OD600 均值，工程默认值，非生物学判断。",
                key="round2_od_threshold_fraction",
            )

    try:
        plan = plan_round2(
            run_df,
            PICHIA_ROUND1_BASELINE,
            od_threshold_fraction=float(od_threshold_fraction),
            ccd_step_fraction=float(ccd_step_fraction),
            max_active_variables=int(max_active_variables),
        )
    except ValueError as exc:
        st.error(f"Round 2 分析暂时无法运行：{exc}")
        return

    st.markdown("#### 噪声与显著性阈值")
    noise = plan.noise
    cols = st.columns(3)
    cols[0].metric("基线重复数", noise["n_replicates"])
    cols[1].metric("基线均值", _num(noise["baseline_mean"]))
    cols[2].metric("显著性阈值", _num(noise["threshold"]))

    technical_noise = _pichia_pooled_technical_noise(run_df)
    if technical_noise:
        st.caption(
            f"另一种噪声：同一样本测 2 次（技术重复）的差异。汇总全部 {technical_noise['n_runs']} 个有重复的编号后，"
            f"技术重复噪声(SD)≈{_num(technical_noise['pooled_sd'])}（自由度 {technical_noise['dof']}，"
            f"比上面基线批次噪声的自由度 {noise['n_replicates'] - 1} 高很多，估得更稳）。"
            "有意不并入上面的显著性阈值：每个 round 1 样本统一测 2 次技术重复，"
            "批次噪声本身已经是「run 间会差多少」的正确尺度；这里只作为独立的数据质量参考——"
            "谁的技术重复差异远超其他编号（下面会标出来），值得单独复核，但不改变谁进入响应面设计。"
        )
        if technical_noise["outliers"]:
            outlier_text = "、".join(f"{run_id}(SD≈{_num(sd)})" for run_id, sd in technical_noise["outliers"])
            st.warning(f"这些编号的技术重复差异明显超过其余编号的普遍水平（>2x 汇总 SD）：{outlier_text}，建议重点核实。")

    st.markdown("#### 固定变量（不显著，或离散档位已测完）")
    fixed_rows = [
        {"变量": PICHIA_VARIABLE_LABELS.get(name, name), "固定取值": _num(value)}
        for name, value in plan.fixed_values.items()
    ]
    st.dataframe(pd.DataFrame(fixed_rows), width="stretch", hide_index=True)

    st.markdown("#### 活跃变量（进入响应面设计）")
    st.metric("活跃变量数 (K)", len(plan.active_variables))
    st.plotly_chart(
        _pichia_effect_magnitude_chart(plan.effects, noise["threshold"]),
        width="stretch",
    )
    if not plan.active_variables:
        st.info("当前数据下没有变量的效应大到需要响应面细化；可以直接看下面的贝叶斯优化建议，或考虑扩大探索范围重新走一轮。")
    for variable, note in plan.untested_notes.items():
        st.warning(f"⚠️ 「{PICHIA_VARIABLE_LABELS.get(variable, variable)}」本轮未测，不是「测过发现不显著」：{note}")
    for variable, note in plan.boundary_notes.items():
        st.warning(f"{PICHIA_VARIABLE_LABELS.get(variable, variable)}：{note}")
    for variable, note in plan.overflow_notes.items():
        st.info(f"{PICHIA_VARIABLE_LABELS.get(variable, variable)}：{note}")

    if plan.combo_interactions:
        st.markdown("#### 联合探索点的可加性检查")
        st.caption("检查每个联合探索样本的实测值和「按单变量效应线性叠加」的预测值之间的差距；差距超过阈值提示可能存在变量间交互。")
        interaction_rows = [
            {
                "样本": item["row_id"],
                "改变的变量": "、".join(PICHIA_VARIABLE_LABELS.get(name, name) for name in item["changed_variables"]),
                "实测": _num(item["observed"]),
                "可加性预测": _num(item["additive_prediction"]),
                "残差": _num(item["residual"]),
                "可能存在交互": "是" if item["possible_interaction"] else "否",
            }
            for item in plan.combo_interactions
        ]
        st.dataframe(pd.DataFrame(interaction_rows), width="stretch", hide_index=True)

    if plan.active_variables:
        st.markdown("#### Round 2 响应面设计（Face-centered CCD）")
        ccd_rows = []
        for row in plan.design_rows:
            display = _pichia_variable_display(row)
            display["复用 Round1 样本"] = row.get("reused_from_round1") or ""
            ccd_rows.append(display)
        st.dataframe(pd.DataFrame(ccd_rows), width="stretch", hide_index=True)
        reused_count = sum(1 for row in plan.design_rows if row.get("reused_from_round1"))
        st.caption(f"共 {len(plan.design_rows)} 个设计点，其中 {reused_count} 个和 Round 1 已有数据重合，可以直接复用、不用重新做。")

    _pichia_round2_design_and_backfill_section(plan, run_df)
    _pichia_round2_results_analysis_section(plan, run_df)

    st.markdown("#### 约束贝叶斯优化建议批次")
    st.caption(
        f"OD600 约束阈值：{_num(plan.od_threshold['threshold'])}"
        f"（= 基线 OD600 均值 x {plan.od_threshold['fraction']}，工程默认值，请研发组结合实际需求确认）。"
    )
    n_batch = st.slider("建议批次大小", min_value=3, max_value=15, value=9, key="round2_bo_batch_size")
    if st.button("生成 Round 2 贝叶斯优化建议", type="primary", key="run_round2_bo"):
        try:
            bo_result = recommend_round2_bo_batch(
                run_df,
                fixed_values=plan.fixed_values,
                active_variables=plan.active_variables or list(PICHIA_CONTINUOUS_BOUNDS),
                od_threshold=plan.od_threshold["threshold"],
                n_batch=int(n_batch),
            )
        except ImportError:
            st.warning("当前环境未安装 torch/botorch/gpytorch，无法运行贝叶斯优化建议；以上响应面(CCD)部分不受影响。")
        except ValueError as exc:
            st.warning(f"未生成建议：{exc}")
        except Exception as exc:
            st.warning(f"贝叶斯优化建议生成失败（样本量很小时，GP 拟合可能数值不稳定）：{exc}")
        else:
            st.session_state["round2_bo_result"] = bo_result
            st.success(
                f"已生成 {len(bo_result['recommendations'])} 个建议"
                f"（候选池 {bo_result['n_candidates']} 个，满足约束 {bo_result['n_feasible']} 个）。"
            )

    bo_result = st.session_state.get("round2_bo_result")
    if bo_result:
        _pichia_render_bo_recommendation_section(bo_result, plan.active_variables, run_df, "round2_bo_cv_result")

        if st.button("保存本次 Round 2 分析到历史记录", key="save_round2_snapshot"):
            created_at = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            record_id = f"r2_{hashlib.sha256(created_at.encode('utf-8')).hexdigest()[:10]}"
            _pichia_ui_records().append(
                {
                    "record_id": record_id,
                    "record_name": f"Round2 {created_at}",
                    "created_at": created_at,
                    "fixed_values": plan.fixed_values,
                    "active_variables": plan.active_variables,
                    "od_threshold": plan.od_threshold,
                    "bo_recommendations": bo_result["recommendations"],
                }
            )
            st.session_state["pichia_ui_record_notice"] = "已保存本次 Round 2 分析"
            _remember_ui_cache()
            st.rerun()

def _pichia_history_tab() -> None:
    st.markdown("### 历史记录（当前会话）")
    st.caption(
        "这里保存的是本次会话里生成的 Round 2 分析快照；刷新页面会恢复，重启应用后清空，不写入文件。"
        "Round 1 的实测结果请用「Round 1」页签里的下载/保存按钮处理，不依赖这里的缓存。"
    )
    notice = st.session_state.pop("pichia_ui_record_notice", None)
    if notice:
        st.success(str(notice))

    clear_cols = st.columns([1, 1, 3])
    with clear_cols[0]:
        clear_confirm = st.checkbox("确认清空界面缓存", key="confirm_clear_pichia_ui_cache")
    with clear_cols[1]:
        if st.button("清空界面缓存", disabled=not clear_confirm, key="clear_pichia_ui_cache"):
            _clear_ui_cache()
            st.rerun()

    records = _pichia_ui_records()
    if not records:
        st.info("暂无历史记录。在「Round 2」页签生成贝叶斯优化建议后可以保存快照。")
        return

    summary_rows = [
        {
            "记录名": record.get("record_name"),
            "时间": record.get("created_at"),
            "活跃变量": "、".join(PICHIA_VARIABLE_LABELS.get(name, name) for name in record.get("active_variables", [])),
            "建议数": len(record.get("bo_recommendations", [])),
        }
        for record in records
    ]
    st.dataframe(pd.DataFrame(summary_rows), width="stretch", hide_index=True)

    labels = [f"{record.get('created_at', '')} | {record.get('record_name', '')}" for record in records]
    selected_label = st.selectbox("查看记录", labels, key="pichia_record_selector")
    selected_index = labels.index(selected_label)
    record = records[selected_index]

    with st.expander("选中记录详情", expanded=False):
        fixed_rows = [
            {"变量": PICHIA_VARIABLE_LABELS.get(name, name), "固定取值": _num(value)}
            for name, value in (record.get("fixed_values") or {}).items()
        ]
        st.dataframe(pd.DataFrame(fixed_rows), width="stretch", hide_index=True)
        if record.get("bo_recommendations"):
            rec_rows = []
            for rec in record["bo_recommendations"]:
                row = _pichia_variable_display(rec)
                row["预测产量"] = _num(rec.get("predicted_yield"))
                rec_rows.append(row)
            st.dataframe(pd.DataFrame(rec_rows), width="stretch", hide_index=True)

        selected_id = str(record.get("record_id"))
        delete_confirm = st.checkbox("确认删除选中记录", key=f"confirm_delete_{selected_id}")
        if st.button("删除选中记录", disabled=not delete_confirm, key=f"delete_record_{selected_id}"):
            del records[selected_index]
            st.session_state["pichia_ui_record_notice"] = "已删除选中记录"
            _remember_ui_cache()
            st.rerun()

def _pichia_hlf_page() -> None:
    _ensure_pichia_data_area()
    st.caption(
        "当前模式：毕赤酵母 hLF 摇瓶实验设计。Round 1 是可配置的基线+单变量+联合探索设计构建器，"
        "Round 2 基于 Round 1 实测结果做显著性分析、响应面(CCD)设计和约束贝叶斯优化。"
    )

    round1_tab, round2_tab, history_tab = st.tabs(["Round 1：实验设计", "Round 2：响应面 + 贝叶斯优化", "历史记录"])
    with round1_tab:
        _pichia_round1_tab()
    with round2_tab:
        _pichia_round2_tab()
    with history_tab:
        _pichia_history_tab()
