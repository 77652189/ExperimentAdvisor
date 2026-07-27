from __future__ import annotations

import copy
import hashlib
import importlib
import json
import sys
from io import BytesIO
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from matplotlib import font_manager
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from experiment_advisor.ingestion.run_level import TARGET_COL, add_model_derived_features, training_view
from experiment_advisor.optimizer.search_space import SearchSpace
from experiment_advisor.recommendation.round1_design import (
    BASELINE as PICHIA_ROUND1_BASELINE,
    OFAT_LEVELS as PICHIA_OFAT_LEVELS,
    count_round1_rows,
    generate_round1_design,
    round1_template_columns,
)
from experiment_advisor.recommendation.round2_design import (
    ALL_VARIABLES as PICHIA_VARIABLES,
    CONTINUOUS_BOUNDS as PICHIA_CONTINUOUS_BOUNDS,
    FIXED_LEVELS as PICHIA_FIXED_LEVELS,
    OD_COL as PICHIA_OD_COL,
    TARGET_COL as PICHIA_TARGET_COL,
    FactorEffect,
    plan_round2,
    recommend_round2_bo_batch,
)
import experiment_advisor.recommendation.service as recommendation_service
from experiment_advisor.recommendation.quality import evaluate_recommendation_quality
from experiment_advisor.report import generate_recommendation_report


def _configure_plot_fonts() -> None:
    """Prefer CJK-capable fonts so Chinese plot labels render correctly."""
    preferred_fonts = [
        "Microsoft YaHei",
        "SimHei",
        "SimSun",
        "Noto Sans CJK SC",
        "Source Han Sans SC",
        "Arial Unicode MS",
    ]
    available_fonts = {font.name for font in font_manager.fontManager.ttflist}
    for font_name in preferred_fonts:
        if font_name in available_fonts:
            plt.rcParams["font.sans-serif"] = [font_name, *preferred_fonts]
            plt.rcParams["font.family"] = "sans-serif"
            break
    plt.rcParams["axes.unicode_minus"] = False


_configure_plot_fonts()

METHOD_LABELS = {
    "standard_bo_qnei": "标准 GP-BO（qNEI）",
}

METHOD_EXPLANATIONS = {
    "standard_bo_qnei": "BoTorch SingleTaskGP 直接拟合产量，并用 qNEI 联合优化下一批推荐。",
}

UNCERTAINTY_LABELS = {
    "gp_posterior_std": "GP 后验标准差",
}

RISK_LABELS = {"low": "低", "medium": "中", "high": "高"}
FLAG_LABELS = {
    "far_from_history": "远离历史实验",
    "near_search_boundary": "接近参数边界",
    "high_residual_uncertainty": "不确定性较高",
}

RECOMMENDATION_CACHE_VERSION = "raw-bo-v1"
RECOMMENDATION_CACHE_KEY = "recommendation_result_cache"
UI_CACHE_VERSION = "ui-cache-v1"

PICHIA_DATA_DIR = PROJECT_ROOT / "data" / "pichia"
PICHIA_UPLOAD_DIR = PICHIA_DATA_DIR / "uploads"
PICHIA_FINAL_DIR = PICHIA_DATA_DIR / "final"
PICHIA_TEMPLATE_DIR = PICHIA_DATA_DIR / "templates"
PICHIA_DEFAULT_DATASET_PATH = PICHIA_FINAL_DIR / "pichia_run_level_dataset.csv"
PICHIA_TEMPLATE_PATH = PICHIA_TEMPLATE_DIR / "pichia_run_level_template.csv"
PICHIA_VARIABLE_LABELS = {
    "seed_od": "种子初始 OD600",
    "glucose_pct": "补料葡萄糖浓度 (%)",
    "ph": "培养基 pH",
    "volume_ml": "初始装液量 (mL)",
    "temp_c": "发酵温度 (℃)",
    "interval_h": "补料时间间隔 (h)",
}

# dataviz palette (dark-mode slots) -- see .claude/plans note on chart conventions:
# fixed categorical order for run_type identity, one diverging pair for correlation,
# accent-vs-muted "emphasis" pair for the effect-magnitude significance cutoff.
PICHIA_RUN_TYPE_COLORS = {"baseline": "#3987e5", "ofat": "#d95926", "combo": "#199e70"}
PICHIA_RUN_TYPE_LABELS = {"baseline": "基线重复", "ofat": "单变量(OFAT)", "combo": "联合探索"}
PICHIA_DIVERGING_COLORSCALE = [[0, "#3987e5"], [0.5, "#383835"], [1, "#e66767"]]
PICHIA_ACCENT_COLOR = "#3987e5"
PICHIA_MUTED_COLOR = "#6b6a64"

PICHIA_UI_CACHE_KEYS = {
    "recommendation_mode",
    "round2_bo_batch_size",
    "round2_max_active_variables",
    "round2_ccd_step_fraction",
    "round2_od_threshold_fraction",
    "pichia_ui_design_records",
}

PICHIA_UI_CACHE_PREFIXES: tuple[str, ...] = ()
# Note: round1_builder_* widget keys deliberately are NOT in the cross-session
# cache -- they pass explicit value=/default=/index= defaults on every render,
# which Streamlit forbids combining with a pre-populated session_state entry
# (raises "widget created with a default value but also had its value set via
# the Session State API"). They're normal per-session widget state instead;
# the thing that actually needs to survive is the generated round1_results_df,
# which is a plain session_state entry, not a cached widget key.


def _load_field_labels() -> dict[str, str]:
    dictionary_path = PROJECT_ROOT / "summary" / "supporting_reports" / "field_dictionary.csv"
    if not dictionary_path.exists():
        return {}
    dictionary = pd.read_csv(dictionary_path)
    return {
        str(row["field"]): str(row["zh_name"])
        for _, row in dictionary.iterrows()
        if pd.notna(row.get("field")) and pd.notna(row.get("zh_name"))
    }


FIELD_LABELS = _load_field_labels()


def _compare_recommenders(df: pd.DataFrame, top_k: int, seed: int, method: str = "ei") -> dict[str, Any]:
    service = importlib.reload(recommendation_service)
    return service.compare_recommenders(df, top_k=top_k, seed=seed, method=method)


def _recommendation_pool_size(top_k: int, multiplier: int = 3) -> int:
    return min(max(top_k * multiplier, top_k), 40)


def _dataset_fingerprint(df: pd.DataFrame) -> str:
    frame = df.copy()
    frame.columns = frame.columns.map(str)
    frame = frame.reindex(sorted(frame.columns), axis=1)

    digest = hashlib.sha256()
    digest.update(RECOMMENDATION_CACHE_VERSION.encode("utf-8"))
    digest.update(str(frame.shape).encode("utf-8"))
    digest.update("\0".join(frame.columns).encode("utf-8"))
    try:
        row_hash = pd.util.hash_pandas_object(frame, index=True).values
    except TypeError:
        row_hash = pd.util.hash_pandas_object(frame.astype(str), index=True).values
    digest.update(row_hash.tobytes())
    return digest.hexdigest()


@st.cache_resource
def _ui_cache_store() -> dict[str, Any]:
    return {"version": UI_CACHE_VERSION, "session_state": {}}


def _is_pichia_ui_cache_key(key: str) -> bool:
    if key in PICHIA_UI_CACHE_KEYS:
        return True
    return any(key.startswith(prefix) for prefix in PICHIA_UI_CACHE_PREFIXES)


def _restore_ui_cache_to_session() -> None:
    store = _ui_cache_store()
    if store.get("version") != UI_CACHE_VERSION:
        store.clear()
        store.update({"version": UI_CACHE_VERSION, "session_state": {}})
        return
    cached = store.get("session_state") or {}
    for key, value in cached.items():
        if key in st.session_state:
            continue
        try:
            st.session_state[key] = copy.deepcopy(value)
        except Exception:
            st.session_state[key] = value


def _remember_ui_cache() -> None:
    store = _ui_cache_store()
    cached: dict[str, Any] = {}
    for key, value in st.session_state.items():
        if not _is_pichia_ui_cache_key(str(key)):
            continue
        try:
            cached[str(key)] = copy.deepcopy(value)
        except Exception:
            cached[str(key)] = value
    store["version"] = UI_CACHE_VERSION
    store["session_state"] = cached


def _clear_ui_cache() -> None:
    store = _ui_cache_store()
    store.clear()
    store.update({"version": UI_CACHE_VERSION, "session_state": {}})
    for key in list(st.session_state.keys()):
        if _is_pichia_ui_cache_key(str(key)):
            del st.session_state[key]


def _read_recommendation_cache(state: Any, dataset_fingerprint: str) -> dict[str, Any] | None:
    cache = state.get(RECOMMENDATION_CACHE_KEY)
    if not isinstance(cache, dict):
        return None
    if cache.get("version") != RECOMMENDATION_CACHE_VERSION:
        return None
    if cache.get("dataset_fingerprint") != dataset_fingerprint:
        return None
    if not cache.get("comparison"):
        return None
    return cache


def _write_recommendation_cache(
    state: Any,
    *,
    dataset_fingerprint: str,
    comparison: dict[str, Any],
    report_md: str,
    report_path: Path,
    run_settings: dict[str, Any],
) -> None:
    state[RECOMMENDATION_CACHE_KEY] = {
        "version": RECOMMENDATION_CACHE_VERSION,
        "dataset_fingerprint": dataset_fingerprint,
        "comparison": comparison,
        "report_md": report_md,
        "report_path": str(report_path),
        "run_settings": run_settings,
    }


def _ensure_strategy_quality(comparison: dict[str, Any], df: pd.DataFrame, feature_cols: list[str]) -> dict[str, Any]:
    quality = comparison.get("strategy_quality_pool") or comparison.get("strategy_quality") or {}
    if quality:
        return quality

    selected = comparison.get("unfiltered_selected_recommendations") or comparison.get("selected_recommendations", [])
    search_bounds = comparison.get("search_space") or {}
    features = feature_cols or list(search_bounds)
    if not selected or not search_bounds or not features:
        return {}

    space = SearchSpace(bounds={name: tuple(bounds) for name, bounds in search_bounds.items()})
    quality = evaluate_recommendation_quality(
        selected,
        _training_data(df),
        space,
        feature_cols=features,
        target_col=TARGET_COL,
    )
    comparison["strategy_quality"] = quality
    return quality


def _select_without_soft_filters(
    comparison: dict[str, Any],
    df: pd.DataFrame,
    feature_cols: list[str],
    target_count: int,
) -> dict[str, Any]:
    selected_method = comparison.get("selected_method", "standard_bo_qnei")
    base = (
        comparison.get("unfiltered_selected_recommendations")
        or comparison.get("recommendations", {}).get(selected_method, [])
        or comparison.get("selected_recommendations", [])
    )
    selected = []
    for item in base[:target_count]:
        tagged = item.copy()
        tagged.pop("soft_filter_status", None)
        tagged.pop("history_range_violations", None)
        tagged.pop("history_range_violation_features", None)
        selected.append(tagged)

    comparison["unfiltered_selected_recommendations"] = base
    comparison["selected_recommendations"] = selected
    comparison["soft_filter"] = {
        "enabled": False,
        "n_before": len(base),
        "n_after": len(selected),
        "target_count": target_count,
    }

    search_bounds = comparison.get("search_space") or {}
    if search_bounds:
        space = SearchSpace(bounds={name: tuple(bounds) for name, bounds in search_bounds.items()})
        comparison["strategy_quality"] = evaluate_recommendation_quality(
            selected,
            _training_data(df),
            space,
            feature_cols=feature_cols,
            target_col=TARGET_COL,
        )
    return comparison


def _history_sigma_ranges(df: pd.DataFrame, feature_cols: list[str], sigma: float) -> dict[str, tuple[float, float, float]]:
    train = _training_data(df)
    ranges: dict[str, tuple[float, float, float]] = {}
    for feature in feature_cols:
        if feature not in train.columns:
            continue
        values = pd.to_numeric(train[feature], errors="coerce").dropna()
        if len(values) < 2:
            continue
        mean = float(values.mean())
        std = float(values.std(ddof=0))
        if std <= 1e-12:
            continue
        ranges[feature] = (mean - sigma * std, mean + sigma * std, std)
    return ranges


def _history_range_violation(
    recommendation: dict[str, Any],
    ranges: dict[str, tuple[float, float, float]],
) -> tuple[int, float, list[str]]:
    count = 0
    total_excess = 0.0
    features = []
    for feature, value in recommendation.get("params", {}).items():
        if feature not in ranges or value is None:
            continue
        low, high, std = ranges[feature]
        numeric_value = float(value)
        if numeric_value < low:
            count += 1
            total_excess += (low - numeric_value) / std
            features.append(feature)
        elif numeric_value > high:
            count += 1
            total_excess += (numeric_value - high) / std
            features.append(feature)
    return count, total_excess, features


def _apply_soft_filters(
    comparison: dict[str, Any],
    df: pd.DataFrame,
    feature_cols: list[str],
    max_nearest_history_distance: float,
    max_boundary_risk: float,
    history_sigma: float,
    target_count: int | None = None,
) -> dict[str, Any]:
    quality = _ensure_strategy_quality(comparison, df, feature_cols)
    per_items = quality.get("per_recommendation") or []
    if not per_items:
        return comparison

    base = comparison.get("unfiltered_selected_recommendations") or comparison.get("selected_recommendations", [])
    per_by_rank = {item.get("rank"): item for item in per_items}
    sigma_ranges = _history_sigma_ranges(df, feature_cols, history_sigma)

    def failure_reasons(rec: dict[str, Any]) -> dict[str, Any]:
        quality_item = per_by_rank.get(rec.get("rank"), {})
        nearest = quality_item.get("nearest_history_distance")
        boundary = quality_item.get("boundary_risk")
        sigma_violation_count, _, sigma_features = _history_range_violation(rec, sigma_ranges)
        return {
            "nearest": nearest is not None and nearest > max_nearest_history_distance,
            "boundary": boundary is not None and boundary > max_boundary_risk,
            "history_range": sigma_violation_count > 0,
            "history_range_features": sigma_features,
        }

    def passes(rec: dict[str, Any]) -> bool:
        reasons = failure_reasons(rec)
        return not (reasons["nearest"] or reasons["boundary"] or reasons["history_range"])

    passed = [item for item in base if passes(item)]
    failed = [item for item in base if item.get("rank") not in {rec.get("rank") for rec in passed}]
    keep_count = target_count if target_count is not None else len(base)

    selected = []
    for item in passed[:keep_count]:
        tagged = item.copy()
        tagged["soft_filter_status"] = "通过"
        violation_count, _, violation_features = _history_range_violation(item, sigma_ranges)
        tagged["history_range_violations"] = violation_count
        tagged["history_range_violation_features"] = violation_features
        selected.append(tagged)

    failed_sigma = [
        item.get("rank")
        for item in failed
        if failure_reasons(item)["history_range"]
    ]
    failed_nearest = [
        item.get("rank")
        for item in failed
        if failure_reasons(item)["nearest"]
    ]
    failed_boundary = [
        item.get("rank")
        for item in failed
        if failure_reasons(item)["boundary"]
    ]
    comparison["unfiltered_selected_recommendations"] = base
    comparison["selected_recommendations"] = selected
    comparison["strategy_quality_pool"] = quality
    search_bounds = comparison.get("search_space") or {}
    if search_bounds:
        space = SearchSpace(bounds={name: tuple(bounds) for name, bounds in search_bounds.items()})
        comparison["strategy_quality"] = evaluate_recommendation_quality(
            selected,
            _training_data(df),
            space,
            feature_cols=feature_cols,
            target_col=TARGET_COL,
        )
    comparison["soft_filter"] = {
        "enabled": True,
        "max_nearest_history_distance": max_nearest_history_distance,
        "max_boundary_risk": max_boundary_risk,
        "history_sigma": history_sigma,
        "n_before": len(base),
        "n_passed": len(passed),
        "n_after": len(selected),
        "target_count": keep_count,
        "failed_ranks": [item.get("rank") for item in failed],
        "failed_nearest_history_ranks": failed_nearest,
        "failed_boundary_risk_ranks": failed_boundary,
        "failed_history_range_ranks": failed_sigma,
        "failure_counts": {
            "nearest_history_distance": len(failed_nearest),
            "boundary_risk": len(failed_boundary),
            "history_range": len(failed_sigma),
        },
    }
    return comparison


def _load_default_dataset() -> pd.DataFrame:
    dataset_path = PROJECT_ROOT / "data" / "final" / "run_level_modeling_dataset.csv"
    if not dataset_path.exists():
        raise FileNotFoundError(f"默认数据集不存在：{dataset_path}")
    return pd.read_csv(dataset_path)


def _training_data(df: pd.DataFrame) -> pd.DataFrame:
    if "exclude_from_training" in df.columns:
        return training_view(df, TARGET_COL)
    return df.dropna(subset=[TARGET_COL])


def _name(value: str) -> str:
    return FIELD_LABELS.get(value, value)


def _display_name(value: str) -> str:
    zh_name = _name(value)
    if zh_name == value:
        return value
    return f"{zh_name} ({value})"


def _display_dataframe(frame: pd.DataFrame, *, keep_english: bool = False) -> pd.DataFrame:
    if keep_english:
        return frame.rename(columns={column: _display_name(str(column)) for column in frame.columns})
    return frame.rename(columns={column: _name(str(column)) for column in frame.columns})


def _num(value: Any) -> Any:
    return round(float(value), 4) if isinstance(value, int | float) else value


def _flags(flags: list[str] | None) -> str:
    if not flags:
        return "无明显风险标记"
    return "；".join(FLAG_LABELS.get(flag, flag) for flag in flags)


def _is_empty_value(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and pd.isna(value):
        return True
    return False


def _drop_empty_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty:
        return frame
    keep = []
    for column in frame.columns:
        values = frame[column]
        if not values.map(_is_empty_value).all():
            keep.append(column)
    return frame[keep]


def _deduplicate_columns(frame: pd.DataFrame) -> pd.DataFrame:
    if frame.empty or not frame.columns.duplicated().any():
        return frame
    counts: dict[str, int] = {}
    columns = []
    for column in frame.columns:
        name = str(column)
        counts[name] = counts.get(name, 0) + 1
        columns.append(name if counts[name] == 1 else f"{name}_{counts[name]}")
    result = frame.copy()
    result.columns = columns
    return result


def _candidate_table(method: str, items: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for item in items:
        row = {"排序": item.get("rank"), "预测产量": _num(item.get("predicted_yield"))}
        if item.get("soft_filter_status"):
            row["软过滤状态"] = item.get("soft_filter_status")
        if item.get("history_range_violations") is not None:
            row["历史范围超限数"] = item.get("history_range_violations")
        if method == "standard_bo_qnei":
            row.update(
                {
                    "GP 后验标准差": _num(item.get("model_uncertainty")),
                    "qNEI 批量推荐得分": _num(item.get("acquisition_score")),
                }
            )
        for key, value in item.get("params", {}).items():
            row[_name(key)] = _num(value)
        rows.append(row)
    return _deduplicate_columns(_drop_empty_columns(pd.DataFrame(rows)))


def _overview(df: pd.DataFrame) -> None:
    train_df = _training_data(df)
    cols = st.columns(4)
    cols[0].metric("总 run 数", len(df))
    cols[1].metric("可训练 run 数", len(train_df))
    cols[2].metric("排除 run 数", len(df) - len(train_df))
    cols[3].metric("目标字段", _display_name(TARGET_COL))

    if not train_df.empty:
        y = pd.to_numeric(train_df[TARGET_COL], errors="coerce").dropna()
        stats = st.columns(4)
        stats[0].metric("历史最低产量", f"{y.min():.3g}")
        stats[1].metric("历史中位产量", f"{y.median():.3g}")
        stats[2].metric("历史最高产量", f"{y.max():.3g}")
        stats[3].metric("历史平均产量", f"{y.mean():.3g}")

    with st.expander("训练数据筛选说明", expanded=False):
        st.write("缺少产量、污染、异常或失败备注的 run 会保留在数据表中用于审计，但不会参与模型训练。")
        if "exclusion_reason" in df.columns:
            counts = df["exclusion_reason"].fillna("").replace("", "可训练").value_counts().reset_index()
            counts.columns = ["原因", "数量"]
            st.dataframe(counts, width="stretch", hide_index=True)

    with st.expander("数据预览", expanded=False):
        st.dataframe(_display_dataframe(df.head(30), keep_english=True), width="stretch")

    with st.expander("字段中英对照", expanded=False):
        dictionary_path = PROJECT_ROOT / "summary" / "supporting_reports" / "field_dictionary.csv"
        if dictionary_path.exists():
            dictionary = pd.read_csv(dictionary_path)
            run_dictionary = dictionary.loc[dictionary["table"] == "run_level_modeling_dataset"].copy()
            run_dictionary = run_dictionary.fillna("")
            run_dictionary = run_dictionary.rename(
                columns={
                    "field": "英文字段",
                    "zh_name": "中文名称",
                    "unit": "单位",
                    "role": "角色",
                    "description": "说明",
                }
            )
            st.dataframe(run_dictionary[["英文字段", "中文名称", "单位", "角色", "说明"]], width="stretch", hide_index=True)
        else:
            st.info("尚未生成字段字典，可运行 `python data/scripts/generate_field_dictionary.py`。")


def _nearest_history(df: pd.DataFrame, params: dict[str, float], top_n: int = 3) -> pd.DataFrame:
    features = [name for name in params if name in df.columns]
    history = _training_data(df).dropna(subset=features)
    if history.empty or not features:
        return pd.DataFrame()
    means = history[features].mean()
    stds = history[features].std(ddof=0).replace(0, 1.0)
    candidate = pd.Series(params)[features]
    distances = (((history[features] - means) / stds - ((candidate - means) / stds)) ** 2).sum(axis=1) ** 0.5
    nearest = history.assign(相似距离=distances).sort_values("相似距离").head(top_n)
    columns = [col for col in ["fermenter_run_id", "experiment_date", TARGET_COL, "相似距离", *features] if col in nearest.columns]
    renamed = nearest[columns].rename(columns={col: _name(col) for col in columns})
    return _deduplicate_columns(renamed)


def _method_block(method: str, items: list[dict[str, Any]]) -> None:
    st.subheader(METHOD_LABELS.get(method, method))
    st.caption(METHOD_EXPLANATIONS.get(method, ""))
    if items:
        st.dataframe(_candidate_table(method, items), width="stretch", hide_index=True)
    else:
        st.info("暂无结果。")


def _standard_bo_summary(top: dict[str, Any]) -> None:
    rows = [
        {"项目": "预测产量", "数值": _num(top.get("predicted_yield")), "说明": "标准 GP 模型对候选点的产量均值预测。"},
        {"项目": "GP 后验标准差", "数值": _num(top.get("model_uncertainty")), "说明": "标准 GP 对该候选点预测不确定性的估计。"},
        {"项目": "qNEI 批量推荐得分", "数值": _num(top.get("acquisition_score")), "说明": "qNEI 联合优化后候选点的模型均值展示值，批内点由联合采集函数生成。"},
    ]
    st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)


def _standard_gp_slice_frame(
    df: pd.DataFrame,
    params: dict[str, float],
    search_space: dict[str, tuple[float, float]] | dict[str, list[float]],
    feature: str,
    fitted_gp: Any,
    feature_cols: list[str],
    n_points: int = 120,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    try:
        import numpy as np
    except Exception as exc:
        raise ImportError("绘制标准 GP 图需要 numpy。") from exc

    train = _training_data(df)[[*feature_cols, TARGET_COL]].dropna()
    if len(train) < 5:
        raise ValueError("至少需要 5 条完整训练数据才能绘制标准 GP 图。")

    # Fix other features at historical mean — makes the slice deterministic
    # (independent of which recommendation or seed was used).
    hist_means = {f: float(train[f].mean()) for f in feature_cols if f in train.columns}
    anchor = {**hist_means, **{k: v for k, v in params.items() if k not in feature_cols}}

    # X-axis spans the historical observed range, not the full search-space bounds.
    # If the recommended value falls outside, extend the range to include it.
    hist_min = float(train[feature].min())
    hist_max = float(train[feature].max())
    rec_val = float(params.get(feature, hist_means.get(feature, hist_min)))
    x_min = min(hist_min, rec_val)
    x_max = max(hist_max, rec_val)

    grid = pd.DataFrame([anchor] * n_points)
    grid[feature] = np.linspace(x_min, x_max, n_points)
    mean, std = fitted_gp.predict(grid[feature_cols], return_std=True)
    curve = pd.DataFrame(
        {
            feature: grid[feature],
            "posterior_mean": mean,
            "posterior_std": std,
            "lower_95": mean - 1.96 * std,
            "upper_95": mean + 1.96 * std,
        }
    )
    return curve, train


def _standard_gp_plot(
    df: pd.DataFrame,
    items: list[dict[str, Any]] | dict[str, Any],
    search_space: dict[str, Any],
    fitted_gp: Any,
    feature_cols: list[str],
) -> None:
    if fitted_gp is None:
        st.info("标准 GP 模型不可用，无法绘制后验切片图。")
        return

    # Accept either a single recommendation dict or a list
    if isinstance(items, dict):
        items = [items]
    if not items:
        st.info("当前推荐缺少可绘图参数。")
        return

    # Build rank labels for the switcher
    rank_options = {
        f"推荐 #{item.get('rank', i + 1)}（预测产量 {_num(item.get('predicted_yield', ''))} g/L）": item
        for i, item in enumerate(items)
    }

    st.markdown("### 标准 GP 后验切片图")
    st.caption(
        "其它参数固定在历史均值，只沿一个参数变化，横轴为历史实测范围。"
        "红虚线为所选推荐点的参数值。此图曲线与随机种子和采集函数无关，反映模型学到的稳定规律。"
    )

    col_rec, col_feat = st.columns([2, 3])
    with col_rec:
        selected_label = st.radio(
            "查看推荐",
            list(rank_options.keys()),
            index=0,
            key="standard_gp_slice_rank",
        )
    active_item = rank_options[selected_label]
    params = active_item.get("params", {})

    features = [f for f in feature_cols if f in search_space and f in params and f in df.columns]
    if not features:
        st.info("当前推荐缺少可绘图参数。")
        return

    with col_feat:
        feature = st.selectbox(
            "选择横轴参数",
            features,
            format_func=_display_name,
            key="standard_gp_slice_feature",
        )

    try:
        curve, train = _standard_gp_slice_frame(df, params, search_space, feature, fitted_gp, feature_cols)
    except Exception as exc:
        st.warning(f"无法绘制标准 GP 图：{exc}")
        return

    fig, ax = plt.subplots(figsize=(8, 4.5))
    ax.plot(curve[feature], curve["posterior_mean"], color="#2563eb", label="GP posterior mean")
    ax.fill_between(
        curve[feature],
        curve["lower_95"],
        curve["upper_95"],
        color="#93c5fd",
        alpha=0.35,
        label="95% posterior interval",
    )
    ax.scatter(train[feature], train[TARGET_COL], color="#111827", s=24, alpha=0.65, label="history")
    rec_val = float(params[feature])
    ax.axvline(rec_val, color="#dc2626", linestyle="--", linewidth=1.5, label=f"recommended value ({rec_val:.2f})")
    ax.set_xlabel(_display_name(feature))
    ax.set_ylabel(_display_name(TARGET_COL))
    ax.grid(alpha=0.2)
    ax.legend(loc="best")
    st.pyplot(fig, clear_figure=True)


def _pdp_curve(fitted_gp: Any, train: pd.DataFrame, feature_cols: list[str], feature: str, n_points: int = 50) -> pd.DataFrame:
    """Calculate a simple 1D PDP curve for explanation tables."""
    try:
        import numpy as np
    except Exception as exc:
        raise ImportError("计算偏依赖摘要需要 numpy。") from exc

    low = float(train[feature].quantile(0.05))
    high = float(train[feature].quantile(0.95))
    grid = np.linspace(low, high, n_points)
    values = []
    for value in grid:
        batch = train.copy()
        batch[feature] = value
        predictions = fitted_gp.predict(batch[feature_cols])
        values.append(float(np.mean(predictions)))
    return pd.DataFrame({feature: grid, "mean_prediction": values})


def _pdp_direction(curve: pd.DataFrame, feature: str) -> tuple[str, str]:
    """Convert a PDP curve into a plain-language trend label."""
    values = curve["mean_prediction"]
    x_values = curve[feature]
    effect = float(values.max() - values.min())
    if effect < 1.0:
        return "影响很弱", "当前数据下模型几乎不随该参数变化"

    start = float(values.iloc[0])
    end = float(values.iloc[-1])
    if end - start > effect * 0.35:
        return "偏高更好", "参数升高时，模型平均预测产量上升"
    if start - end > effect * 0.35:
        return "偏低更好", "参数升高时，模型平均预测产量下降"

    best_position = (float(x_values.iloc[values.idxmax()]) - float(x_values.min())) / max(float(x_values.max() - x_values.min()), 1e-9)
    if best_position < 0.35:
        return "低区间较好", "模型偏好的区域靠近历史低值端"
    if best_position > 0.65:
        return "高区间较好", "模型偏好的区域靠近历史高值端"
    return "中间区间较好", "模型偏好的区域靠近历史中间值"


def _pdp_summary(fitted_gp: Any, train: pd.DataFrame, feature_cols: list[str]) -> pd.DataFrame:
    """Build a concise PDP interpretation table for non-ML readers."""
    rows = []
    for feature in feature_cols:
        curve = _pdp_curve(fitted_gp, train, feature_cols, feature)
        values = curve["mean_prediction"]
        x_values = curve[feature]
        best_index = int(values.idxmax())
        direction, explanation = _pdp_direction(curve, feature)
        rows.append(
            {
                "参数": _display_name(feature),
                "模型倾向": direction,
                "模型偏好值": _num(float(x_values.iloc[best_index])),
                "影响幅度(g/L)": _num(float(values.max() - values.min())),
                "直观解释": explanation,
            }
        )
    return pd.DataFrame(rows).sort_values("影响幅度(g/L)", ascending=False)


def _gp_pdp(df: pd.DataFrame, fitted_gp: Any, feature_cols: list[str]) -> None:
    """使用偏依赖图展示 GP 各特征的平均边际效应。"""
    import math as _math
    import numpy as np

    if fitted_gp is None:
        st.info("标准 GP 模型不可用，无法绘制偏依赖图。")
        return
    if not feature_cols:
        st.info("无可用特征列，无法绘制偏依赖图。")
        return

    train = _training_data(df)[feature_cols].dropna()
    if len(train) < 5:
        st.info("训练数据不足，无法绘制偏依赖图。")
        return

    with st.expander("这张图怎么读", expanded=True):
        st.write(
            "偏依赖图不是单次实验的真实曲线，而是模型在历史参数分布上做平均后的趋势。"
            "曲线向上表示该参数增大时，模型平均预测产量更高；曲线向下表示更低。"
            "黑色短线表示历史实验实际覆盖的位置，远离短线密集区域的结论要谨慎。"
        )
        st.write("下面的表把每条曲线翻译成更直接的工艺判断，按模型认为影响幅度从大到小排序。")
        try:
            st.dataframe(_pdp_summary(fitted_gp, train, feature_cols), width="stretch", hide_index=True)
        except Exception as exc:
            st.warning(f"偏依赖摘要计算失败：{exc}")

    st.markdown("### 各特征平均边际效应（1D 偏依赖图）")
    st.caption(
        "对每个特征值，将其余特征在训练数据的实际分布上取平均，"
        "反映该特征的边际效应。与切片图不同，平均过程自然携带特征间的协变关系。"
    )

    n_features = len(feature_cols)
    n_cols = min(3, n_features)
    n_rows = _math.ceil(n_features / n_cols)
    fig1, axes = plt.subplots(n_rows, n_cols, figsize=(5 * n_cols, 4 * n_rows))
    axes_flat = np.array(axes).flatten()

    for i, feature in enumerate(feature_cols):
        curve = _pdp_curve(fitted_gp, train, feature_cols, feature, n_points=50)
        ax = axes_flat[i]
        ax.plot(curve[feature], curve["mean_prediction"], color="#2563eb", linewidth=1.8)
        y_min = float(curve["mean_prediction"].min())
        y_max = float(curve["mean_prediction"].max())
        margin = (y_max - y_min) * 0.15 if y_max > y_min else 1.0
        rug_y = y_min - margin * 0.6
        ax.plot(train[feature].values, np.full(len(train), rug_y),
                "|", color="black", markersize=8, alpha=0.6)
        ax.set_ylim(rug_y - margin * 0.2, y_max + margin * 0.2)
        ax.set_title(_display_name(feature), fontsize=10)
        ax.set_xlabel(_display_name(feature))
        ax.set_ylabel("平均预测产量（g/L）")
        ax.grid(alpha=0.2)

    for ax in axes_flat[n_features:]:
        ax.set_visible(False)

    fig1.suptitle("GP 偏依赖图（平均边际效应）", fontsize=13)
    plt.tight_layout()
    st.pyplot(fig1, clear_figure=True)

    key_a = "temperature_shift_time_h"
    key_b = "lactose_total_ml"
    if key_a in feature_cols and key_b in feature_cols:
        st.markdown("### 升温时机 × 乳糖总量 联动效应（2D 偏依赖图）")
        st.caption(
            "两个相关性最强的特征（EDA Spearman r=0.91）的联合偏依赖图。"
            "颜色越深表示该参数组合下 GP 预测产量越高。"
        )
        a_low = float(train[key_a].quantile(0.05))
        a_high = float(train[key_a].quantile(0.95))
        b_low = float(train[key_b].quantile(0.05))
        b_high = float(train[key_b].quantile(0.95))
        grid_a = np.linspace(a_low, a_high, 20)
        grid_b = np.linspace(b_low, b_high, 20)
        Z = np.zeros((len(grid_b), len(grid_a)))
        for i, val_a in enumerate(grid_a):
            for j, val_b in enumerate(grid_b):
                batch = train.copy()
                batch[key_a] = val_a
                batch[key_b] = val_b
                Z[j, i] = float(np.mean(fitted_gp.predict(batch[feature_cols])))

        in_window = (
            train[key_a].between(a_low, a_high)
            & train[key_b].between(b_low, b_high)
        )
        n_outside = int((~in_window).sum())
        if n_outside:
            st.caption(
                f"该 2D 图只显示 5%-95% 历史分位窗口，避免把少数边缘点拉大坐标轴并造成外推误读；"
                f"{n_outside} 个窗口外历史点未参与坐标轴缩放。"
            )

        fig2, ax2 = plt.subplots(figsize=(7, 5))
        filled = ax2.contourf(grid_a, grid_b, Z, levels=10, cmap="viridis")
        plt.colorbar(filled, ax=ax2, label="预测产量（g/L）")
        contour_lines = ax2.contour(grid_a, grid_b, Z, levels=10,
                                    colors="black", linewidths=0.5, alpha=0.6)
        ax2.clabel(contour_lines, fmt="%.2f", fontsize=8)
        ax2.scatter(train.loc[in_window, key_a], train.loc[in_window, key_b],
                    marker="|", color="black", alpha=0.5, s=60, zorder=5)
        ax2.set_xlim(a_low, a_high)
        ax2.set_ylim(b_low, b_high)
        ax2.set_xlabel(_display_name(key_a))
        ax2.set_ylabel(_display_name(key_b))
        ax2.set_title("GP预测产量：升温时机 × 乳糖总量")
        plt.tight_layout()
        st.pyplot(fig2, clear_figure=True)


def _loocv_scatter(df: pd.DataFrame, feature_cols: list[str]) -> None:
    """LOO-CV 预测 vs 实测散点图，评估 GP 模型的泛化能力。"""
    try:
        import warnings
        import numpy as np
        from sklearn.exceptions import ConvergenceWarning
        from sklearn.gaussian_process import GaussianProcessRegressor
        from sklearn.gaussian_process.kernels import ConstantKernel, Matern, WhiteKernel
        from sklearn.metrics import mean_absolute_error, r2_score
        from sklearn.model_selection import LeaveOneOut, cross_val_predict
        from sklearn.pipeline import make_pipeline
        from sklearn.preprocessing import StandardScaler
    except ImportError as exc:
        st.warning(f"LOO-CV 需要 scikit-learn：{exc}")
        return

    if not feature_cols:
        st.info("无可用特征列。")
        return

    feature_df = add_model_derived_features(_training_data(df))
    train = feature_df[feature_cols + [TARGET_COL]].dropna()
    if len(train) < 8:
        st.info("训练数据不足（至少需要 8 条），无法计算 LOO-CV。")
        return

    x = train[feature_cols]
    y = train[TARGET_COL].astype(float).values
    n_features = len(feature_cols)

    kernel = (
        ConstantKernel(1.0, constant_value_bounds=(1e-3, 1e3))
        * Matern(
            nu=2.5,
            length_scale=[1.0] * n_features,
            length_scale_bounds=[(1e-2, 1e2)] * n_features,
        )
        + WhiteKernel(noise_level=1.0, noise_level_bounds=(1e-3, 1e2))
    )
    pipeline = make_pipeline(
        StandardScaler(),
        GaussianProcessRegressor(kernel=kernel, normalize_y=True, random_state=42),
    )

    with st.spinner("正在计算 LOO-CV（约需 5-15 秒）..."):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", ConvergenceWarning)
            y_pred = cross_val_predict(pipeline, x, y, cv=LeaveOneOut())

    r2 = r2_score(y, y_pred)
    mae = mean_absolute_error(y, y_pred)

    fig, ax = plt.subplots(figsize=(6, 5))
    ax.scatter(y, y_pred, color="#2563eb", alpha=0.7, s=40, label="历史 run")
    lims = [float(min(np.min(y), np.min(y_pred)) - 5), float(max(np.max(y), np.max(y_pred)) + 5)]
    ax.plot(lims, lims, color="#dc2626", linestyle="--", linewidth=1.2, label="完美预测线")
    ax.set_xlabel("实测产量 (g/L)")
    ax.set_ylabel("LOO-CV 预测产量 (g/L)")
    ax.set_title(f"GP 泛化能力：LOO-CV  |  R²={r2:.3f}  |  MAE={mae:.2f} g/L")
    ax.legend()
    ax.grid(alpha=0.2)
    plt.tight_layout()
    st.pyplot(fig, clear_figure=True)

    if r2 >= 0.6:
        st.success(f"R²={r2:.3f}，GP 对留出实验有较好的预测能力，推荐方向可参考。")
    elif r2 >= 0.3:
        st.warning(f"R²={r2:.3f}，GP 有一定预测能力，但不确定性较高，推荐应结合人工判断。")
    else:
        st.error(f"R²={r2:.3f}，GP 泛化能力较弱，推荐参数仅供参考，建议扩充数据后再优化。")

    st.caption(
        f"每个点代表一条历史 run：用其余 {len(train) - 1} 条数据训练 GP，预测该 run 的产量。"
        f"MAE={mae:.2f} g/L 表示平均预测误差。"
    )


def _nearest_history_validation(
    df: pd.DataFrame,
    params: dict[str, float],
    feature_cols: list[str],
    top_n: int = 5,
) -> None:
    """展示与推荐参数最相似的历史实验及其实际产量。"""
    try:
        import numpy as np
    except ImportError as exc:
        st.warning(f"最近邻验证需要 numpy：{exc}")
        return

    features = [feature for feature in feature_cols if feature in df.columns and feature in params]
    history = _training_data(df).dropna(subset=features + [TARGET_COL])
    if history.empty or not features:
        st.info("无足够历史数据进行最近邻对比。")
        return

    means = history[features].mean()
    stds = history[features].std(ddof=0).replace(0, 1.0)
    candidate = pd.Series(params)[features]
    norm_candidate = (candidate - means) / stds
    norm_history = (history[features] - means) / stds
    distances = ((norm_history - norm_candidate) ** 2).sum(axis=1) ** 0.5
    nearest = history.assign(_dist=distances).sort_values("_dist").head(top_n)

    actual_yields = nearest[TARGET_COL].values
    recommended_yield = params.get("predicted_yield")

    fig, ax = plt.subplots(figsize=(7, 4))
    run_labels = [
        str(run_id)[:12] if not pd.isna(run_id) else f"Run {index + 1}"
        for index, run_id in enumerate(nearest.get("fermenter_run_id", nearest.index))
    ]
    ax.bar(run_labels, actual_yields, color="#6b7280", alpha=0.8, label="历史实测产量")
    if recommended_yield is not None:
        ax.axhline(
            float(recommended_yield),
            color="#2563eb",
            linestyle="--",
            linewidth=1.5,
            label=f"推荐预测产量 ({float(recommended_yield):.1f} g/L)",
        )
    ax.set_ylabel("产量 (g/L)")
    ax.set_title(f"与推荐参数最相似的 {top_n} 条历史实验")
    ax.legend()
    ax.grid(axis="y", alpha=0.2)
    plt.xticks(rotation=15, ha="right")
    plt.tight_layout()
    st.pyplot(fig, clear_figure=True)

    mean_nearby = float(np.mean(actual_yields))
    max_nearby = float(np.max(actual_yields))
    dist_nearest = float(nearest["_dist"].iloc[0])

    if dist_nearest > 2.0:
        st.warning(
            f"最近邻距离={dist_nearest:.2f}（标准化），推荐参数远离历史实验区域，属于外推，需谨慎。"
        )
    else:
        st.success(
            f"最近邻距离={dist_nearest:.2f}，推荐参数有历史数据支撑。"
            f"相似实验平均产量 {mean_nearby:.1f} g/L，最高 {max_nearby:.1f} g/L。"
        )

    display_cols = [feature for feature in features if feature in nearest.columns] + [TARGET_COL]
    table = nearest[display_cols].copy().rename(columns={column: _display_name(column) for column in display_cols})
    table.insert(0, "相似距离", nearest["_dist"].round(3).values)
    if "fermenter_run_id" in nearest.columns:
        table.insert(0, "Run ID", nearest["fermenter_run_id"].values)
    st.dataframe(table, width="stretch", hide_index=True)


def _metric_value(metrics: dict[str, Any], key: str) -> Any:
    value = metrics.get(key)
    return _num(value) if isinstance(value, int | float) else value


def _strategy_quality_block(comparison: dict[str, Any], df: pd.DataFrame, selected: list[dict[str, Any]], feature_cols: list[str]) -> None:
    quality = comparison.get("strategy_quality") or {}
    if not quality:
        search_bounds = comparison.get("search_space") or {}
        features = feature_cols or list(search_bounds)
        if selected and search_bounds and features:
            try:
                space = SearchSpace(bounds={name: tuple(bounds) for name, bounds in search_bounds.items()})
                quality = evaluate_recommendation_quality(
                    selected,
                    _training_data(df),
                    space,
                    feature_cols=features,
                    target_col=TARGET_COL,
                )
                comparison["strategy_quality"] = quality
                st.caption("当前结果来自旧缓存，已基于现有推荐现场补算策略质量指标。")
            except Exception as exc:
                st.info(f"当前结果没有推荐策略质量指标，现场补算也失败：{exc}")
                return
        else:
            st.info("当前结果没有推荐策略质量指标。请重新运行推荐生成完整诊断。")
            return

    st.markdown("### Batch 多样性")
    diversity = quality.get("batch_diversity", {})
    diversity_cols = st.columns(4)
    diversity_cols[0].metric("最小两两距离", _metric_value(diversity, "min_pairwise_distance"))
    diversity_cols[1].metric("平均两两距离", _metric_value(diversity, "mean_pairwise_distance"))
    diversity_cols[2].metric("簇数量", diversity.get("cluster_count_threshold_0_10"))
    diversity_cols[3].metric("平均特征覆盖", _metric_value(diversity, "mean_feature_range_coverage"))

    coverage = diversity.get("feature_range_coverage") or {}
    if coverage:
        coverage_rows = [
            {"参数": _display_name(feature), "字段名": feature, "覆盖比例": _num(value)}
            for feature, value in coverage.items()
        ]
        st.dataframe(pd.DataFrame(coverage_rows), width="stretch", hide_index=True)

    st.markdown("### 历史支撑与边界风险")
    support = quality.get("history_support", {})
    boundary = quality.get("boundary_risk", {})
    risk_cols = st.columns(4)
    risk_cols[0].metric("平均最近邻距离", _metric_value(support, "mean_nearest_history_distance"))
    risk_cols[1].metric("最远最近邻距离", _metric_value(support, "max_nearest_history_distance"))
    risk_cols[2].metric("平均边界风险", _metric_value(boundary, "mean_boundary_risk"))
    risk_cols[3].metric("高边界风险数量", boundary.get("n_near_boundary_gt_0_8"))

    per_items = quality.get("per_recommendation") or []
    if per_items:
        rows = []
        for item in per_items:
            rows.append(
                {
                    "排序": item.get("rank"),
                    "类型": item.get("recommendation_type", "—"),
                    "预测产量": _num(item.get("predicted_yield")),
                    "GP 后验标准差": _num(item.get("model_uncertainty")),
                    "最近邻距离": _num(item.get("nearest_history_distance")),
                    "最近邻 Run": item.get("nearest_run_id", "—"),
                    "最近邻产量": _num(item.get("nearest_run_yield")),
                    "边界风险": _num(item.get("boundary_risk")),
                }
            )
        st.dataframe(pd.DataFrame(rows), width="stretch", hide_index=True)

    st.markdown("### 推荐参数的历史近邻")
    st.caption("为每个推荐组合分别展示最相似的历史实验实际产量，用于判断整批推荐是否落在有数据支撑的区域。")
    if selected:
        for item in selected:
            rank = item.get("rank", "?")
            predicted = item.get("predicted_yield")
            title = f"推荐 #{rank}"
            if predicted is not None:
                title = f"{title}｜预测产量 {_num(predicted)} g/L"
            with st.expander(title, expanded=rank == 1):
                selected_params = {**item.get("params", {}), "predicted_yield": predicted}
                _nearest_history_validation(df, selected_params, feature_cols)
    else:
        st.info("暂无推荐点，无法做历史近邻对比。")


def _metric_explanations() -> None:
    st.markdown("### 主推荐：标准 GP-BO（qNEI）")
    standard_rows = [
        {"指标": "预测产量", "如何产生": "GP 直接拟合历史产量后，对候选点输出 posterior mean。", "如何解读": "模型预测值，不是实测值。"},
        {"指标": "GP 后验标准差", "如何产生": "GP 对候选点预测分布的 posterior std。", "如何解读": "越高表示模型在该区域越不确定。"},
        {"指标": "qNEI 批量推荐", "如何产生": "BoTorch qNoisyExpectedImprovement 联合优化整批候选点。", "如何解读": "批内候选会考虑边际收益递减，减少 top-k 聚集。"},
        {"指标": "观测噪声", "如何产生": "qNEI 基于 noisy baseline 建模已观测数据。", "如何解读": "相比直接使用历史最大值的 EI，更不容易被噪声高点牵引。"},
    ]
    st.dataframe(pd.DataFrame(standard_rows), width="stretch", hide_index=True)


def _ensure_pichia_data_area() -> None:
    for directory in [PICHIA_UPLOAD_DIR, PICHIA_FINAL_DIR, PICHIA_TEMPLATE_DIR]:
        directory.mkdir(parents=True, exist_ok=True)
    if not PICHIA_TEMPLATE_PATH.exists():
        pd.DataFrame(columns=round1_template_columns()).to_csv(PICHIA_TEMPLATE_PATH, index=False, encoding="utf-8-sig")


def _pichia_numeric_results(df: pd.DataFrame) -> pd.DataFrame:
    numeric = df.copy()
    for column in (PICHIA_TARGET_COL, PICHIA_OD_COL):
        if column in numeric.columns:
            numeric[column] = pd.to_numeric(numeric[column], errors="coerce")
    return numeric


def _pichia_ui_records() -> list[dict[str, Any]]:
    records = st.session_state.setdefault("pichia_ui_design_records", [])
    if not isinstance(records, list):
        records = []
        st.session_state["pichia_ui_design_records"] = records
    return records


def _pichia_variable_display(row: dict[str, Any]) -> dict[str, Any]:
    return {PICHIA_VARIABLE_LABELS.get(name, name): _num(row.get(name)) for name in PICHIA_VARIABLES if name in row}


def _pichia_round1_builder() -> None:
    st.markdown("#### 方案配置")
    st.caption(
        "基线重复、单变量法(OFAT)、联合探索(LHS) 三个模块可以独立开关和配置——某个数量填 0，"
        "或不选任何变量，就相当于关掉那个模块，可以拼出纯 LHS、纯 OFAT 或任意组合。"
        "表单打开时的默认值就是已经和研发组确认过的方案（基线×3 + OFAT×11 + 联合探索×4，共 18 个样本）。"
    )

    st.markdown("##### 基线取值")
    baseline_config: dict[str, float] = {}
    continuous_vars = list(PICHIA_CONTINUOUS_BOUNDS)
    baseline_cols = st.columns(len(continuous_vars))
    for index, variable in enumerate(continuous_vars):
        lower, upper = PICHIA_CONTINUOUS_BOUNDS[variable]
        with baseline_cols[index]:
            baseline_config[variable] = float(
                st.number_input(
                    PICHIA_VARIABLE_LABELS.get(variable, variable),
                    min_value=float(lower),
                    max_value=float(upper),
                    value=float(PICHIA_ROUND1_BASELINE[variable]),
                    key=f"round1_builder_baseline_{variable}",
                )
            )
    fixed_vars = list(PICHIA_FIXED_LEVELS)
    fixed_cols = st.columns(len(fixed_vars))
    for index, variable in enumerate(fixed_vars):
        options = PICHIA_FIXED_LEVELS[variable]
        default_value = PICHIA_ROUND1_BASELINE[variable]
        with fixed_cols[index]:
            baseline_config[variable] = float(
                st.selectbox(
                    PICHIA_VARIABLE_LABELS.get(variable, variable),
                    options,
                    index=options.index(default_value) if default_value in options else 0,
                    key=f"round1_builder_baseline_{variable}",
                )
            )

    st.markdown("##### 基线重复")
    n_baseline = st.number_input(
        "基线重复次数（0 = 不生成基线重复行）",
        min_value=0,
        max_value=10,
        value=3,
        key="round1_builder_n_baseline",
    )

    st.markdown("##### 单变量法 (OFAT)")
    ofat_vars = st.multiselect(
        "参与 OFAT 的变量（不选 = 不生成 OFAT 行）",
        list(PICHIA_VARIABLES),
        default=list(PICHIA_VARIABLES),
        format_func=lambda v: PICHIA_VARIABLE_LABELS.get(v, v),
        key="round1_builder_ofat_vars",
    )
    ofat_levels_config: dict[str, list[float]] = {}
    ofat_errors: list[str] = []
    for variable in ofat_vars:
        default_levels = PICHIA_OFAT_LEVELS.get(variable, [])
        selected = st.multiselect(
            f"{PICHIA_VARIABLE_LABELS.get(variable, variable)} 的测试水平",
            options=default_levels,
            default=default_levels,
            accept_new_options=True,
            help="可以直接勾掉/加回推荐水平，也可以在框里输入自定义数值后回车添加。",
            key=f"round1_builder_ofat_levels_{variable}",
        )
        try:
            ofat_levels_config[variable] = sorted({float(value) for value in selected})
        except (TypeError, ValueError):
            ofat_errors.append(PICHIA_VARIABLE_LABELS.get(variable, variable))
    if ofat_errors:
        st.error(f"以下变量输入了无法识别成数字的自定义水平，请检查：{', '.join(ofat_errors)}")

    st.markdown("##### 联合探索 (LHS)")
    n_combo = st.number_input(
        "联合探索点数（0 = 不生成联合探索行）",
        min_value=0,
        max_value=50,
        value=4,
        key="round1_builder_n_combo",
    )
    combo_vars = st.multiselect(
        "参与联合探索的连续变量",
        continuous_vars,
        default=continuous_vars,
        format_func=lambda v: PICHIA_VARIABLE_LABELS.get(v, v),
        key="round1_builder_combo_vars",
    )

    counts = count_round1_rows(ofat_levels_config, int(n_baseline), int(n_combo), baseline=baseline_config)
    st.caption(
        f"预计生成：基线 {counts['baseline']} + OFAT {counts['ofat']} + 联合探索 {counts['combo']} "
        f"= 共 {counts['total']} 行"
    )

    generate_col, reset_col = st.columns(2)
    with generate_col:
        if st.button(
            "生成 Round 1 设计",
            type="primary",
            width="stretch",
            disabled=bool(ofat_errors),
            key="generate_round1_design_button",
        ):
            st.session_state["round1_results_df"] = generate_round1_design(
                baseline=baseline_config,
                ofat_levels=ofat_levels_config or None,
                n_baseline_replicates=int(n_baseline),
                n_combo_points=int(n_combo),
                combo_variables=combo_vars or None,
            )
            st.session_state.pop("round2_bo_result", None)
            st.success(f"已生成 {counts['total']} 行 Round 1 设计。")
    with reset_col:
        if st.button("直接使用已验证方案（18 样本，不改上面的表单）", width="stretch", key="round1_use_preset"):
            st.session_state["round1_results_df"] = generate_round1_design()
            st.session_state.pop("round2_bo_result", None)
            st.success("已生成已验证的 18 样本方案。")


def _pichia_yield_scatter_chart(df: pd.DataFrame, value_col: str, title: str) -> go.Figure:
    """Box+jitter chart of `value_col` grouped by run_type: shows both the
    per-group distribution/spread and every individual point (run_id on hover),
    covering the "distribution" and "error/noise" asks in one figure."""

    numeric = df.copy()
    numeric[value_col] = pd.to_numeric(numeric[value_col], errors="coerce")
    plot_df = numeric.dropna(subset=[value_col])

    fig = go.Figure()
    for run_type in ["baseline", "ofat", "combo"]:
        subset = plot_df[plot_df["run_type"] == run_type]
        if subset.empty:
            continue
        fig.add_trace(
            go.Box(
                y=subset[value_col],
                x=[PICHIA_RUN_TYPE_LABELS[run_type]] * len(subset),
                boxpoints="all",
                jitter=0.5,
                pointpos=0,
                marker_color=PICHIA_RUN_TYPE_COLORS[run_type],
                line_color=PICHIA_RUN_TYPE_COLORS[run_type],
                text=subset["run_id"],
                hovertemplate="%{text}<br>" + value_col + "=%{y}<extra></extra>",
            )
        )
    fig.update_layout(
        title=title,
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        showlegend=False,
        yaxis_title=value_col,
        margin=dict(t=40, b=30, l=50, r=20),
        height=380,
    )
    return fig


def _pichia_correlation_heatmap(df: pd.DataFrame) -> go.Figure:
    columns = [*PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL]
    numeric = df[columns].apply(pd.to_numeric, errors="coerce")
    corr = numeric.corr(method="spearman", min_periods=3)
    labels = [PICHIA_VARIABLE_LABELS.get(column, column) for column in columns]
    fig = go.Figure(
        data=go.Heatmap(
            z=corr.to_numpy(),
            x=labels,
            y=labels,
            colorscale=PICHIA_DIVERGING_COLORSCALE,
            zmid=0,
            zmin=-1,
            zmax=1,
            hovertemplate="%{y} vs %{x}: %{z:.2f}<extra></extra>",
            colorbar=dict(title="Spearman r"),
        )
    )
    fig.update_layout(
        title="变量与产量/OD600 相关性（Spearman，仅 18 个样本，供参考，非结论性）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        margin=dict(t=40, b=40, l=110, r=20),
        height=450,
    )
    return fig


def _pichia_baseline_lookup(df: pd.DataFrame) -> dict[str, Any]:
    """Most-common value per variable column. Stands in for "the baseline"
    even when baseline-replicate rows are disabled (n_baseline_replicates=0),
    since every row that doesn't deliberately vary a given variable leaves it
    at that value -- so the mode across the whole design is still correct."""
    lookup: dict[str, Any] = {}
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            mode = df[variable].mode()
            if not mode.empty:
                lookup[variable] = mode.iloc[0]
    return lookup


def _pichia_format_value(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def _pichia_type_label(run_type: str, changed_variable: str | None) -> str:
    """"类型" column text. OFAT rows name the specific variable being swept
    ("单变量-种子初始OD600") rather than a generic "单变量(OFAT)" label shared
    by all 11 OFAT rows -- otherwise the type column alone can't tell them
    apart at a glance. Baseline/combo stay generic since every row of that
    type really is the same thing."""
    if run_type == "ofat":
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable or "")
        return f"单变量-{label}"
    return PICHIA_RUN_TYPE_LABELS.get(run_type, run_type)


def _pichia_row_note(
    run_type: str,
    changed_variable: str | None,
    row: pd.Series | None = None,
    baseline_lookup: dict[str, Any] | None = None,
) -> str:
    if run_type == "baseline":
        return "当前摇瓶标准条件重复，用于估计批次噪声（纯误差）"
    if run_type == "ofat":
        label = PICHIA_VARIABLE_LABELS.get(changed_variable, changed_variable or "")
        if row is not None and baseline_lookup and changed_variable in baseline_lookup:
            old = _pichia_format_value(baseline_lookup[changed_variable])
            new = _pichia_format_value(row.get(changed_variable))
            return f"单变量法：「{label}」由基线 {old} 改为 {new}，其余变量维持基线不变"
        return f"单变量法：只改变「{label}」，其余变量维持基线不变"
    if run_type == "combo":
        return "联合探索：连续变量按拉丁超立方(LHS)联合取值，温度/补料间隔取固定档位组合"
    return ""


def _pichia_design_display_frame(df: pd.DataFrame) -> pd.DataFrame:
    """Chinese-labeled, note-annotated preview of the design -- used for the
    plain in-app st.dataframe. No row coloring here: st.dataframe's Styler
    support renders inconsistently for this table in practice, so the actual
    "pretty, colored" output is the Excel workbook below instead of trying to
    force parity inside Streamlit's canvas-rendered grid."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    display = pd.DataFrame(index=df.index)
    display["编号"] = df["run_id"]
    display["类型"] = [
        _pichia_type_label(run_type.loc[idx], changed_variable.loc[idx]) for idx in df.index
    ]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            display[PICHIA_VARIABLE_LABELS.get(variable, variable)] = df[variable]
    display["收获时OD600"] = df.get(PICHIA_OD_COL)
    display["hLF产量"] = df.get(PICHIA_TARGET_COL)
    display["备注/目的"] = [
        _pichia_row_note(run_type.loc[idx], changed_variable.loc[idx], df.loc[idx], baseline_lookup)
        for idx in df.index
    ]
    return display


def _pichia_round1_workbook_bytes(df: pd.DataFrame) -> bytes:
    """Polished .xlsx twin of the design: colored rows by run_type, Chinese
    headers, an auto-generated notes column, a legend, and a frozen header --
    the same look already approved from an earlier one-off script, now built
    dynamically from whatever design was actually generated (any mix of
    baseline/OFAT/LHS) instead of a fixed example."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    font_name = "Calibri"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    baseline_fill = PatternFill("solid", fgColor="E2EFDA")
    combo_fill = PatternFill("solid", fgColor="DDEBF7")
    fillin_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # only "type" and "note" hold variable-length sentences -- everything else
    # is a short id/number, so it stays single-line and the table stays compact
    wrap_keys = {"run_type", "note"}

    columns: list[tuple[str, str, float]] = [("run_id", "编号", 9), ("run_type", "类型", 18)]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            columns.append((variable, PICHIA_VARIABLE_LABELS.get(variable, variable), 13))
    columns.append((PICHIA_OD_COL, "收获时OD600(待填)", 12))
    columns.append((PICHIA_TARGET_COL, "hLF产量(待填)", 12))
    columns.append(("note", "备注/目的", 34))

    wb = Workbook()
    ws = wb.active
    ws.title = "Round1设计"
    ws.sheet_view.showGridLines = False

    for col_idx, (_key, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # deliberately no explicit row_dimensions[1].height: leaving customHeight
    # unset lets Excel auto-fit each row to its own wrapped content on open,
    # which is what actually works when note length varies this much between
    # row types -- any single fixed height is either too tall for short
    # baseline notes or clipped for the longer combo/ofat notes.
    ws.freeze_panes = "A2"

    row_fill_by_type = {"baseline": baseline_fill, "combo": combo_fill}

    for offset, idx in enumerate(df.index):
        row_idx = offset + 2
        rt = run_type.loc[idx]
        row = df.loc[idx]
        note = _pichia_row_note(rt, changed_variable.loc[idx], row, baseline_lookup)
        values: dict[str, Any] = {
            "run_id": row.get("run_id"),
            "run_type": _pichia_type_label(rt, changed_variable.loc[idx]),
            PICHIA_OD_COL: row.get(PICHIA_OD_COL),
            PICHIA_TARGET_COL: row.get(PICHIA_TARGET_COL),
            "note": note,
        }
        for variable in PICHIA_VARIABLES:
            if variable in df.columns:
                values[variable] = row.get(variable)

        row_fill = row_fill_by_type.get(rt)
        for col_idx, (key, _label, _width) in enumerate(columns, start=1):
            value = values.get(key)
            if isinstance(value, float) and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            if key == "note":
                cell.alignment = wrap_left
            elif key in wrap_keys:
                cell.alignment = wrap_center
            else:
                cell.alignment = center
            if key in (PICHIA_OD_COL, PICHIA_TARGET_COL):
                cell.fill = fillin_fill
            elif row_fill is not None:
                cell.fill = row_fill

    legend_row = len(df) + 3
    legend_title = ws.cell(row=legend_row, column=1, value="图例：")
    legend_title.font = Font(name=font_name, size=10, bold=True)
    legend_items = [
        (baseline_fill, "基线重复（估计批次噪声）"),
        (combo_fill, "联合探索点（LHS，多变量同时变化）"),
        (fillin_fill, "需要填写的结果列"),
    ]
    for offset, (fill, text) in enumerate(legend_items):
        marker_row = legend_row + 1 + offset
        marker = ws.cell(row=marker_row, column=1, value=" ")
        marker.fill = fill
        marker.border = border
        label_cell = ws.cell(row=marker_row, column=2, value=text)
        label_cell.font = Font(name=font_name, size=10)
        ws.merge_cells(start_row=marker_row, start_column=2, end_row=marker_row, end_column=len(columns))

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _pichia_round1_tab() -> None:
    _ensure_pichia_data_area()
    st.markdown("### Round 1：实验设计构建器")
    st.caption(
        "温度只能是 20/25/30℃，补料间隔只能是 12/24h——这两个变量的档位一旦被单变量法测全，"
        "就已经穷尽了离散选项，Round 2 不会再对它们做响应面细化。"
    )

    with st.expander("方案配置", expanded="round1_results_df" not in st.session_state):
        _pichia_round1_builder()

    if "round1_results_df" not in st.session_state:
        st.info("配置好方案后点击「生成 Round 1 设计」，或直接用「直接使用已验证方案」一键生成。")
        return

    st.markdown("#### 上传已回填结果（可选）")
    uploaded = st.file_uploader("上传已回填 yield_g_per_l / od600 的 CSV", type=["csv"], key="round1_results_upload")
    if uploaded is not None:
        try:
            payload = uploaded.getvalue()
            uploaded_df = pd.read_csv(BytesIO(payload))
        except Exception as exc:
            st.error(f"读取失败：{exc}")
        else:
            missing = [column for column in ["run_id", *PICHIA_VARIABLES, PICHIA_TARGET_COL, PICHIA_OD_COL] if column not in uploaded_df.columns]
            if missing:
                st.error(f"上传文件缺少字段：{', '.join(missing)}")
            else:
                _ensure_pichia_data_area()
                archive_name = Path(uploaded.name).name or "pichia_round1_uploaded.csv"
                (PICHIA_UPLOAD_DIR / archive_name).write_bytes(payload)
                st.session_state["round1_results_df"] = uploaded_df
                st.success("已加载并归档上传的 Round 1 结果。")

    working_df = st.session_state["round1_results_df"]

    st.markdown("#### 设计总览")
    excel_bytes = _pichia_round1_workbook_bytes(working_df)
    st.download_button(
        "下载设计表格（Excel，配色+备注，和示例一致）",
        data=excel_bytes,
        file_name="pichia_round1_design.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key="download_round1_excel",
    )
    st.caption("配色、图例和备注列在 Excel 里最完整；下面是网页内的快速预览（不带颜色）。")
    st.dataframe(_pichia_design_display_frame(working_df), width="stretch", hide_index=True)

    st.markdown("#### 回填实测结果")
    st.caption("按编号对应到上面总览表里的样本，这里只需要填产量和 OD600 这两列。")
    entry_view = working_df[["run_id", PICHIA_TARGET_COL, PICHIA_OD_COL]].reset_index(drop=True)
    edited_entries = st.data_editor(
        entry_view,
        width="stretch",
        num_rows="fixed",
        hide_index=True,
        key="round1_data_editor",
        disabled=["run_id"],
        column_config={
            "run_id": "编号",
            PICHIA_TARGET_COL: "hLF产量",
            PICHIA_OD_COL: "收获时OD600",
        },
    )
    edited = working_df.copy()
    edited[PICHIA_TARGET_COL] = edited_entries[PICHIA_TARGET_COL].to_numpy()
    edited[PICHIA_OD_COL] = edited_entries[PICHIA_OD_COL].to_numpy()
    st.session_state["round1_results_df"] = edited

    numeric = _pichia_numeric_results(edited)
    filled = int(numeric[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1).sum())
    st.caption(f"已回填 {filled}/{len(edited)} 行完整结果（yield_g_per_l 和 od600 都有值才算完成）。")

    action_cols = st.columns(2)
    with action_cols[0]:
        csv_bytes = edited.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "下载当前表格",
            data=csv_bytes,
            file_name="pichia_round1_results.csv",
            mime="text/csv",
            key="download_round1_results",
        )
    with action_cols[1]:
        if st.button("保存到 data/pichia/final/（供 Round 2 和下次打开使用）", key="save_round1_to_final"):
            _ensure_pichia_data_area()
            edited.to_csv(PICHIA_DEFAULT_DATASET_PATH, index=False, encoding="utf-8-sig")
            st.success(f"已保存到 {PICHIA_DEFAULT_DATASET_PATH}")

    if int(numeric[PICHIA_TARGET_COL].notna().sum()) >= 3:
        st.markdown("#### 结果可视化")
        chart_cols = st.columns(2)
        with chart_cols[0]:
            st.plotly_chart(
                _pichia_yield_scatter_chart(numeric, PICHIA_TARGET_COL, "产量分布（按样本类型）"),
                width="stretch",
            )
        with chart_cols[1]:
            st.plotly_chart(
                _pichia_yield_scatter_chart(numeric, PICHIA_OD_COL, "OD600 分布（按样本类型）"),
                width="stretch",
            )
        st.plotly_chart(_pichia_correlation_heatmap(numeric), width="stretch")
    else:
        st.caption("至少回填 3 行产量数据后，这里会显示分布图和相关性热力图。")


def _pichia_effect_magnitude_chart(effects: dict[str, FactorEffect], threshold: float) -> go.Figure:
    """Horizontal bar of each variable's effect_magnitude, descending, with a
    confidence-interval error bar and a reference line at the significance
    threshold -- this is what actually decides K (how many variables become
    "active" in resolve_round2_variables). Emphasis coloring (accent vs muted)
    rather than a full categorical palette, since the only distinction that
    matters here is "crossed the line" vs "didn't"."""

    ordered = sorted(effects.values(), key=lambda effect: effect.effect_magnitude, reverse=True)
    labels = [PICHIA_VARIABLE_LABELS.get(effect.variable, effect.variable) for effect in ordered]
    magnitudes = [effect.effect_magnitude for effect in ordered]
    colors = [PICHIA_ACCENT_COLOR if effect.significant else PICHIA_MUTED_COLOR for effect in ordered]
    error_plus = [max(effect.ci_high - effect.effect_magnitude, 0.0) for effect in ordered]
    error_minus = [max(effect.effect_magnitude - effect.ci_low, 0.0) for effect in ordered]

    fig = go.Figure(
        go.Bar(
            x=magnitudes,
            y=labels,
            orientation="h",
            marker_color=colors,
            error_x=dict(type="data", symmetric=False, array=error_plus, arrayminus=error_minus, color="#c3c2b7"),
            hovertemplate="%{y}: %{x:.3g}<extra></extra>",
        )
    )
    fig.add_vline(
        x=threshold,
        line_dash="dash",
        line_color="#c3c2b7",
        annotation_text=f"显著性阈值 {threshold:.3g}",
        annotation_position="top",
    )
    fig.update_layout(
        title="各变量效应量（误差棒=置信区间，仅 3 次基线重复，df=2，区间偏宽属实情）",
        template="plotly_dark",
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        xaxis_title="效应量 |偏离基线|",
        margin=dict(t=50, b=40, l=140, r=30),
        height=340,
    )
    return fig


def _pichia_round2_tab() -> None:
    st.markdown("### Round 2：显著性分析 → 响应面（CCD）+ 约束贝叶斯优化")

    results_df = st.session_state.get("round1_results_df")
    if results_df is None or results_df.empty:
        st.info("请先在「Round 1：实验设计」页签生成设计并回填实测结果。")
        return

    run_df = _pichia_numeric_results(results_df)
    filled = int(run_df[[PICHIA_TARGET_COL, PICHIA_OD_COL]].notna().all(axis=1).sum())
    st.caption(f"当前已回填 {filled}/{len(run_df)} 行完整结果。")
    if filled < len(run_df):
        st.warning("还有行没有回填完整结果；下面的分析只会用到已经有产量和 OD600 的行，结论可能随着数据补全而变化。")

    with st.expander("分析参数", expanded=False):
        param_cols = st.columns(3)
        with param_cols[0]:
            max_active_variables = st.number_input(
                "最多活跃变量数 (K 上限)",
                min_value=1,
                max_value=6,
                value=3,
                help="超过这个数量的显著变量会按效应量排序，排不进的先固定在 round 1 最优水平。",
                key="round2_max_active_variables",
            )
        with param_cols[1]:
            ccd_step_fraction = st.slider(
                "CCD 步长比例",
                min_value=0.1,
                max_value=1.0,
                value=0.5,
                step=0.05,
                help="响应面设计的轴向步长 = 该比例 x 单变量法(OFAT)的水平间距。",
                key="round2_ccd_step_fraction",
            )
        with param_cols[2]:
            od_threshold_fraction = st.slider(
                "OD600 约束比例",
                min_value=0.1,
                max_value=1.0,
                value=0.7,
                step=0.05,
                help="OD600 约束阈值 = 该比例 x 基线 OD600 均值，工程默认值，非生物学判断。",
                key="round2_od_threshold_fraction",
            )

    try:
        plan = plan_round2(
            run_df,
            PICHIA_ROUND1_BASELINE,
            od_threshold_fraction=float(od_threshold_fraction),
            ccd_step_fraction=float(ccd_step_fraction),
            max_active_variables=int(max_active_variables),
        )
    except ValueError as exc:
        st.error(f"Round 2 分析暂时无法运行：{exc}")
        return

    st.markdown("#### 噪声与显著性阈值")
    noise = plan.noise
    cols = st.columns(3)
    cols[0].metric("基线重复数", noise["n_replicates"])
    cols[1].metric("基线均值", _num(noise["baseline_mean"]))
    cols[2].metric("显著性阈值", _num(noise["threshold"]))

    st.markdown("#### 固定变量（不显著，或离散档位已测完）")
    fixed_rows = [
        {"变量": PICHIA_VARIABLE_LABELS.get(name, name), "固定取值": _num(value)}
        for name, value in plan.fixed_values.items()
    ]
    st.dataframe(pd.DataFrame(fixed_rows), width="stretch", hide_index=True)

    st.markdown("#### 活跃变量（进入响应面设计）")
    st.metric("活跃变量数 (K)", len(plan.active_variables))
    st.plotly_chart(
        _pichia_effect_magnitude_chart(plan.effects, noise["threshold"]),
        width="stretch",
    )
    if not plan.active_variables:
        st.info("当前数据下没有变量的效应大到需要响应面细化；可以直接看下面的贝叶斯优化建议，或考虑扩大探索范围重新走一轮。")
    for variable, note in plan.boundary_notes.items():
        st.warning(f"{PICHIA_VARIABLE_LABELS.get(variable, variable)}：{note}")
    for variable, note in plan.overflow_notes.items():
        st.info(f"{PICHIA_VARIABLE_LABELS.get(variable, variable)}：{note}")

    if plan.combo_interactions:
        st.markdown("#### 联合探索点的可加性检查")
        st.caption("检查每个联合探索样本的实测值和「按单变量效应线性叠加」的预测值之间的差距；差距超过阈值提示可能存在变量间交互。")
        interaction_rows = [
            {
                "样本": item["row_id"],
                "改变的变量": "、".join(PICHIA_VARIABLE_LABELS.get(name, name) for name in item["changed_variables"]),
                "实测": _num(item["observed"]),
                "可加性预测": _num(item["additive_prediction"]),
                "残差": _num(item["residual"]),
                "可能存在交互": "是" if item["possible_interaction"] else "否",
            }
            for item in plan.combo_interactions
        ]
        st.dataframe(pd.DataFrame(interaction_rows), width="stretch", hide_index=True)

    if plan.active_variables:
        st.markdown("#### Round 2 响应面设计（Face-centered CCD）")
        ccd_rows = []
        for row in plan.design_rows:
            display = _pichia_variable_display(row)
            display["复用 Round1 样本"] = row.get("reused_from_round1") or ""
            ccd_rows.append(display)
        st.dataframe(pd.DataFrame(ccd_rows), width="stretch", hide_index=True)
        reused_count = sum(1 for row in plan.design_rows if row.get("reused_from_round1"))
        st.caption(f"共 {len(plan.design_rows)} 个设计点，其中 {reused_count} 个和 Round 1 已有数据重合，可以直接复用、不用重新做。")

    st.markdown("#### 约束贝叶斯优化建议批次")
    st.caption(
        f"OD600 约束阈值：{_num(plan.od_threshold['threshold'])}"
        f"（= 基线 OD600 均值 x {plan.od_threshold['fraction']}，工程默认值，请研发组结合实际需求确认）。"
    )
    n_batch = st.slider("建议批次大小", min_value=3, max_value=15, value=9, key="round2_bo_batch_size")
    if st.button("生成 Round 2 贝叶斯优化建议", type="primary", key="run_round2_bo"):
        try:
            bo_result = recommend_round2_bo_batch(
                run_df,
                fixed_values=plan.fixed_values,
                active_variables=plan.active_variables or list(PICHIA_CONTINUOUS_BOUNDS),
                od_threshold=plan.od_threshold["threshold"],
                n_batch=int(n_batch),
            )
        except ImportError:
            st.warning("当前环境未安装 torch/botorch/gpytorch，无法运行贝叶斯优化建议；以上响应面(CCD)部分不受影响。")
        except ValueError as exc:
            st.warning(f"未生成建议：{exc}")
        else:
            st.session_state["round2_bo_result"] = bo_result
            st.success(
                f"已生成 {len(bo_result['recommendations'])} 个建议"
                f"（候选池 {bo_result['n_candidates']} 个，满足约束 {bo_result['n_feasible']} 个）。"
            )

    bo_result = st.session_state.get("round2_bo_result")
    if bo_result:
        bo_rows = []
        for rec in bo_result["recommendations"]:
            row = _pichia_variable_display(rec)
            row["预测产量"] = _num(rec.get("predicted_yield"))
            row["预测产量标准差"] = _num(rec.get("predicted_yield_std"))
            row["预测 OD600"] = _num(rec.get("predicted_od600"))
            bo_rows.append(row)
        st.dataframe(pd.DataFrame(bo_rows), width="stretch", hide_index=True)

        if st.button("保存本次 Round 2 分析到历史记录", key="save_round2_snapshot"):
            created_at = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")
            record_id = f"r2_{hashlib.sha256(created_at.encode('utf-8')).hexdigest()[:10]}"
            _pichia_ui_records().append(
                {
                    "record_id": record_id,
                    "record_name": f"Round2 {created_at}",
                    "created_at": created_at,
                    "fixed_values": plan.fixed_values,
                    "active_variables": plan.active_variables,
                    "od_threshold": plan.od_threshold,
                    "bo_recommendations": bo_result["recommendations"],
                }
            )
            st.session_state["pichia_ui_record_notice"] = "已保存本次 Round 2 分析"
            _remember_ui_cache()
            st.rerun()


def _pichia_history_tab() -> None:
    st.markdown("### 历史记录（当前会话）")
    st.caption(
        "这里保存的是本次会话里生成的 Round 2 分析快照；刷新页面会恢复，重启应用后清空，不写入文件。"
        "Round 1 的实测结果请用「Round 1」页签里的下载/保存按钮处理，不依赖这里的缓存。"
    )
    notice = st.session_state.pop("pichia_ui_record_notice", None)
    if notice:
        st.success(str(notice))

    clear_cols = st.columns([1, 1, 3])
    with clear_cols[0]:
        clear_confirm = st.checkbox("确认清空界面缓存", key="confirm_clear_pichia_ui_cache")
    with clear_cols[1]:
        if st.button("清空界面缓存", disabled=not clear_confirm, key="clear_pichia_ui_cache"):
            _clear_ui_cache()
            st.rerun()

    records = _pichia_ui_records()
    if not records:
        st.info("暂无历史记录。在「Round 2」页签生成贝叶斯优化建议后可以保存快照。")
        return

    summary_rows = [
        {
            "记录名": record.get("record_name"),
            "时间": record.get("created_at"),
            "活跃变量": "、".join(PICHIA_VARIABLE_LABELS.get(name, name) for name in record.get("active_variables", [])),
            "建议数": len(record.get("bo_recommendations", [])),
        }
        for record in records
    ]
    st.dataframe(pd.DataFrame(summary_rows), width="stretch", hide_index=True)

    labels = [f"{record.get('created_at', '')} | {record.get('record_name', '')}" for record in records]
    selected_label = st.selectbox("查看记录", labels, key="pichia_record_selector")
    selected_index = labels.index(selected_label)
    record = records[selected_index]

    with st.expander("选中记录详情", expanded=False):
        fixed_rows = [
            {"变量": PICHIA_VARIABLE_LABELS.get(name, name), "固定取值": _num(value)}
            for name, value in (record.get("fixed_values") or {}).items()
        ]
        st.dataframe(pd.DataFrame(fixed_rows), width="stretch", hide_index=True)
        if record.get("bo_recommendations"):
            rec_rows = []
            for rec in record["bo_recommendations"]:
                row = _pichia_variable_display(rec)
                row["预测产量"] = _num(rec.get("predicted_yield"))
                rec_rows.append(row)
            st.dataframe(pd.DataFrame(rec_rows), width="stretch", hide_index=True)

        selected_id = str(record.get("record_id"))
        delete_confirm = st.checkbox("确认删除选中记录", key=f"confirm_delete_{selected_id}")
        if st.button("删除选中记录", disabled=not delete_confirm, key=f"delete_record_{selected_id}"):
            del records[selected_index]
            st.session_state["pichia_ui_record_notice"] = "已删除选中记录"
            _remember_ui_cache()
            st.rerun()


def _pichia_hlf_page() -> None:
    _ensure_pichia_data_area()
    st.caption(
        "当前模式：毕赤酵母 hLF 摇瓶实验设计。Round 1 是固定的基线+单变量+联合探索设计，"
        "Round 2 基于 Round 1 实测结果做显著性分析、响应面(CCD)设计和约束贝叶斯优化。"
    )

    round1_tab, round2_tab, history_tab = st.tabs(["Round 1：实验设计", "Round 2：响应面 + 贝叶斯优化", "历史记录"])
    with round1_tab:
        _pichia_round1_tab()
    with round2_tab:
        _pichia_round2_tab()
    with history_tab:
        _pichia_history_tab()


def main() -> None:
    st.set_page_config(
        page_title="发酵工艺优化推荐系统",
        page_icon=r"C:\Users\63097\Documents\LauncherIcons\experimentadvisor_8505.png",
        layout="wide",
    )
    _restore_ui_cache_to_session()
    st.title("发酵工艺优化推荐系统")
    with st.sidebar:
        mode = st.radio("推荐模式", ["毕赤酵母 hLF", "大肠杆菌 BO（历史，数据已作废）"], key="recommendation_mode")
    _remember_ui_cache()

    if mode == "毕赤酵母 hLF":
        _pichia_hlf_page()
        _remember_ui_cache()
        return

    st.caption("主方法：标准 GP-BO（qNEI），使用 BoTorch 联合优化批量推荐。")

    with st.sidebar:
        st.header("数据入口")
        source = st.radio("选择数据来源", ["使用 data/final/run_level_modeling_dataset.csv", "上传 run-level CSV"])
        bo_method = st.radio(
            "推荐方法",
            options=["EI（顺序贪心）", "qNEI（批量联合）"],
            index=0,
            help=(
                "EI：单点期望改善，顺序贪心生成批次，结果稳定可解释，推荐作为主方法。\n\n"
                "qNEI：联合批量优化，理论上多样性更好，但对随机种子敏感，推荐值偶有偏离历史分布。"
            ),
            key="bo_method",
        )
        _method_arg = "ei" if bo_method.startswith("EI") else "qnei"
        top_k = st.slider("推荐数量", min_value=1, max_value=10, value=5)
        seed = st.number_input(
            "随机种子",
            min_value=0,
            max_value=9999,
            value=0,
            step=1,
            help="固定种子保证每次运行结果一致。修改为不同整数可探索多组推荐方案。",
            key="bo_seed",
        )
        enable_soft_filter = st.checkbox(
            "启用软过滤",
            value=False,
            help="开启后先生成更大的候选池，再按最近邻距离、边界风险和历史合理范围筛出主推荐。",
            key="enable_soft_filter",
        )
        candidate_pool_multiplier = 1
        max_nearest_history_distance = 2.0
        max_boundary_risk = 0.8
        history_sigma = 2.0
        if enable_soft_filter:
            candidate_pool_multiplier = st.slider(
                "候选池倍数",
                min_value=1,
                max_value=5,
                value=3,
                help="先生成 推荐数量×倍数 的候选池，再用软过滤筛出主推荐。候选池越大越可能凑够通过项，但运行会更慢。",
                key="candidate_pool_multiplier",
            )
            max_nearest_history_distance = st.number_input(
                "最大最近邻距离",
                min_value=0.0,
                max_value=10.0,
                value=2.0,
                step=0.1,
                help="软过滤阈值。推荐点到最近历史实验的标准化距离超过该值时，不进入主推荐。",
                key="max_nearest_history_distance",
            )
            max_boundary_risk = st.number_input(
                "最大边界风险",
                min_value=0.0,
                max_value=1.0,
                value=0.8,
                step=0.05,
                help="软过滤阈值。推荐点过于靠近搜索空间边界时，不进入主推荐。",
                key="max_boundary_risk",
            )
            history_sigma = st.number_input(
                "历史合理范围 σ 倍数",
                min_value=0.5,
                max_value=5.0,
                value=2.0,
                step=0.5,
                help="软过滤阈值。每个参数以历史均值±k个标准差作为合理范围，比绝对min/max更能抵抗异常批次。",
                key="history_sigma",
            )
        run_button = st.button("运行推荐", type="primary", width="stretch")
        st.divider()
        st.markdown("### 默认决策")
        st.write("默认采用标准 GP-BO（qNEI）作为主推荐；qNEI 联合优化整批候选点并处理观测噪声。")

    uploaded_file = st.file_uploader("上传已整理好的 run-level CSV", type=["csv"]) if source == "上传 run-level CSV" else None

    try:
        df = _load_default_dataset() if source == "使用 data/final/run_level_modeling_dataset.csv" else pd.read_csv(uploaded_file) if uploaded_file else None
    except Exception as exc:
        st.error(f"数据加载失败：{exc}")
        return

    if df is None:
        st.info("请上传 run-level CSV，或选择默认数据目录。")
        return
    if TARGET_COL not in df.columns:
        st.error(f"数据缺少目标字段 `{TARGET_COL}`。")
        return

    _overview(df)
    dataset_fingerprint = _dataset_fingerprint(df)
    report_path = PROJECT_ROOT / "summary" / "recommendation_report.md"
    cache_message = ""
    if run_button:
        with st.spinner("正在训练标准 GP-BO（qNEI）并生成推荐..."):
            pool_size = _recommendation_pool_size(top_k, candidate_pool_multiplier) if enable_soft_filter else int(top_k)
            comparison = _compare_recommenders(df, top_k=pool_size, seed=int(seed), method=_method_arg)
            comparison["requested_top_k"] = int(top_k)
            comparison["recommendation_pool_size"] = int(pool_size)
            comparison["soft_filter_enabled"] = bool(enable_soft_filter)
            feature_cols_for_filter = comparison.get("model_info", {}).get("standard_bo_feature_cols", [])
            if enable_soft_filter:
                comparison = _apply_soft_filters(
                    comparison,
                    df,
                    feature_cols_for_filter,
                    max_nearest_history_distance=float(max_nearest_history_distance),
                    max_boundary_risk=float(max_boundary_risk),
                    history_sigma=float(history_sigma),
                    target_count=int(top_k),
                )
            else:
                comparison = _select_without_soft_filters(
                    comparison,
                    df,
                    feature_cols_for_filter,
                    target_count=int(top_k),
                )
            run_settings = {
                "top_k": int(top_k),
                "pool_size": int(pool_size),
                "seed": int(seed),
                "method": _method_arg,
                "soft_filter_enabled": bool(enable_soft_filter),
                "candidate_pool_multiplier": int(candidate_pool_multiplier),
                "max_nearest_history_distance": float(max_nearest_history_distance),
                "max_boundary_risk": float(max_boundary_risk),
                "history_sigma": float(history_sigma),
            }
            comparison["run_settings"] = run_settings
            report_md = generate_recommendation_report(comparison, output_path=report_path)
            _write_recommendation_cache(
                st.session_state,
                dataset_fingerprint=dataset_fingerprint,
                comparison=comparison,
                report_md=report_md,
                report_path=report_path,
                run_settings=run_settings,
            )
        st.session_state["recommendation_comparison"] = comparison
        st.session_state["recommendation_report_md"] = report_md
        st.session_state["recommendation_report_path"] = str(report_path)
        cache_message = "已重新训练并更新缓存。"
    else:
        cached = _read_recommendation_cache(st.session_state, dataset_fingerprint)
        if cached:
            comparison = cached["comparison"]
            report_md = cached.get("report_md", "")
            report_path = Path(cached.get("report_path", str(report_path)))
            st.session_state["recommendation_comparison"] = comparison
            st.session_state["recommendation_report_md"] = report_md
            st.session_state["recommendation_report_path"] = str(report_path)
            cache_message = "已自动读取缓存结果；只有点击“运行推荐”才会重新训练。"
        else:
            st.info("当前数据没有可用推荐缓存。点击左侧“运行推荐”后，系统会训练模型、生成候选点，并缓存本次结果。")
            return

    selected_method = comparison.get("selected_method", "standard_bo_qnei")
    selected = comparison.get("selected_recommendations", [])
    decision = comparison.get("decision", {})
    fitted_gp = comparison.get("model_info", {}).get("fitted_standard_bo_gp")
    feature_cols = comparison.get("model_info", {}).get("standard_bo_feature_cols", [])
    model_feature_cols = comparison.get("model_info", {}).get("standard_bo_model_feature_cols", feature_cols)
    selected = comparison.get("selected_recommendations", [])
    if not report_md:
        report_md = generate_recommendation_report(comparison, output_path=report_path)
        st.session_state["recommendation_report_md"] = report_md

    st.success("推荐已生成")
    if cache_message:
        st.caption(cache_message)
    run_settings = comparison.get("run_settings") or {}
    if run_settings:
        st.caption(
            "当前结果参数：方法 {method}，推荐数 {top_k}，候选池 {pool_size}，随机种子 {seed}，软过滤 {soft_filter}。".format(
                method=run_settings.get("method"),
                top_k=run_settings.get("top_k"),
                pool_size=run_settings.get("pool_size"),
                seed=run_settings.get("seed"),
                soft_filter="启用" if run_settings.get("soft_filter_enabled") else "未启用",
            )
        )
    cols = st.columns(2)
    cols[0].metric("主推荐方法", METHOD_LABELS.get(selected_method, selected_method))
    cols[1].metric("训练 run 数", comparison.get("n_training_rows"))
    st.info(decision.get("reason", "默认采用标准 GP-BO（qNEI）。"))
    soft_filter = comparison.get("soft_filter") or {}
    if soft_filter and not soft_filter.get("enabled", True):
        st.caption(
            "软过滤未启用：直接显示 BO 候选池前 {after}/{target} 个推荐。".format(
                after=soft_filter.get("n_after"),
                target=soft_filter.get("target_count"),
            )
        )
    elif soft_filter:
        st.caption(
            "软过滤：最近邻距离 <= {dist}，边界风险 <= {risk}，历史合理范围 ±{sigma}σ；"
            "候选池 {before} 个，通过 {passed} 个，当前显示 {after}/{target} 个推荐。".format(
                dist=_num(soft_filter.get("max_nearest_history_distance")),
                risk=_num(soft_filter.get("max_boundary_risk")),
                sigma=_num(soft_filter.get("history_sigma")),
                passed=soft_filter.get("n_passed"),
                after=soft_filter.get("n_after"),
                before=soft_filter.get("n_before"),
                target=soft_filter.get("target_count"),
            )
        )
        failure_counts = soft_filter.get("failure_counts") or {}
        if failure_counts:
            st.caption(
                "软过滤失败原因计数（可重叠）：最近邻距离 {nearest}，边界风险 {boundary}，历史合理范围 {history_range}。".format(
                    nearest=failure_counts.get("nearest_history_distance", 0),
                    boundary=failure_counts.get("boundary_risk", 0),
                    history_range=failure_counts.get("history_range", 0),
                )
            )
        if soft_filter.get("failed_nearest_history_ranks"):
            st.caption(f"最近邻距离超限推荐排序：{soft_filter.get('failed_nearest_history_ranks')}")
        if soft_filter.get("failed_boundary_risk_ranks"):
            st.caption(f"边界风险超限推荐排序：{soft_filter.get('failed_boundary_risk_ranks')}")
        if soft_filter.get("failed_history_range_ranks"):
            st.caption(f"历史合理范围外推荐排序：{soft_filter.get('failed_history_range_ranks')}")
        if soft_filter.get("target_count") and soft_filter.get("n_after", 0) < soft_filter.get("target_count"):
            st.warning("当前候选池中通过软过滤的推荐不足目标数量；建议放宽阈值、提高候选池倍数，或更换随机种子。")

    tabs = st.tabs(["主推荐", "代理模型验证", "推荐策略质量", "GP 偏依赖图", "指标说明", "Markdown 报告"])
    with tabs[0]:
        _method_block(selected_method, selected)
        if selected:
            _standard_bo_summary(selected[0])
            _standard_gp_plot(df, selected, comparison.get("search_space", {}), fitted_gp, feature_cols)
    with tabs[1]:
        st.markdown("### LOO-CV 模型泛化能力")
        st.caption("验证 GP 模型是否能预测它没见过的实验结果。这是判断推荐可信度的核心依据。")
        _loocv_scatter(df, model_feature_cols)
    with tabs[2]:
        _strategy_quality_block(comparison, df, selected, feature_cols)
    with tabs[3]:
        st.caption("偏依赖图反映特征的平均边际效应，自然携带历史数据中的特征协变关系，优于固定其他参数的条件切片图。")
        _gp_pdp(df, fitted_gp, feature_cols)
    with tabs[4]:
        _metric_explanations()
    with tabs[5]:
        st.markdown(report_md)
        st.caption(f"报告已写入：{report_path}")


if __name__ == "__main__":
    main()
