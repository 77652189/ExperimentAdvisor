"""Pichia design-sheet I/O: Excel export for both rounds, and the parsing that
turns a returned sheet back into an analysable frame.

Shared by Round 1 and Round 2 on purpose -- both rounds hand out a sheet and
take the same sheet back, so the unit-annotated headers, "not detected" tokens
and same-run_id technical replicates have to be understood identically on both
paths. Pure functions only: nothing here touches st.* or session state.
"""
from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from App.pichia_common import (
    PICHIA_OD_COL,
    PICHIA_RUN_TYPE_COLORS,
    PICHIA_RUN_TYPE_LABELS,
    PICHIA_TARGET_COL,
    PICHIA_VARIABLE_LABELS,
    PICHIA_VARIABLES,
    _num,
    _pichia_baseline_lookup,
    _pichia_row_note,
    _pichia_type_label,
)

def _pichia_round1_workbook_bytes(df: pd.DataFrame) -> bytes:
    """Polished .xlsx twin of the design: colored rows by run_type, Chinese
    headers, an auto-generated notes column, a legend, and a frozen header --
    the same look already approved from an earlier one-off script, now built
    dynamically from whatever design was actually generated (any mix of
    baseline/OFAT/LHS) instead of a fixed example."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    font_name = "Calibri"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    baseline_fill = PatternFill("solid", fgColor="E2EFDA")
    combo_fill = PatternFill("solid", fgColor="DDEBF7")
    fillin_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    # only "type" and "note" hold variable-length sentences -- everything else
    # is a short id/number, so it stays single-line and the table stays compact
    wrap_keys = {"run_type", "note"}

    columns: list[tuple[str, str, float]] = [("run_id", "编号", 9), ("run_type", "类型", 18)]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            columns.append((variable, PICHIA_VARIABLE_LABELS.get(variable, variable), 13))
    columns.append((PICHIA_OD_COL, "收获时OD600(待填)", 12))
    columns.append((PICHIA_TARGET_COL, "hLF产量(待填)", 12))
    columns.append(("note", "备注/目的", 34))

    wb = Workbook()
    ws = wb.active
    ws.title = "Round1设计"
    ws.sheet_view.showGridLines = False

    for col_idx, (_key, label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    # deliberately no explicit row_dimensions[1].height: leaving customHeight
    # unset lets Excel auto-fit each row to its own wrapped content on open,
    # which is what actually works when note length varies this much between
    # row types -- any single fixed height is either too tall for short
    # baseline notes or clipped for the longer combo/ofat notes.
    ws.freeze_panes = "A2"

    row_fill_by_type = {"baseline": baseline_fill, "combo": combo_fill}

    for offset, idx in enumerate(df.index):
        row_idx = offset + 2
        rt = run_type.loc[idx]
        row = df.loc[idx]
        note = _pichia_row_note(rt, changed_variable.loc[idx], row, baseline_lookup)
        values: dict[str, Any] = {
            "run_id": row.get("run_id"),
            "run_type": _pichia_type_label(rt, changed_variable.loc[idx]),
            PICHIA_OD_COL: row.get(PICHIA_OD_COL),
            PICHIA_TARGET_COL: row.get(PICHIA_TARGET_COL),
            "note": note,
        }
        for variable in PICHIA_VARIABLES:
            if variable in df.columns:
                values[variable] = row.get(variable)

        row_fill = row_fill_by_type.get(rt)
        for col_idx, (key, _label, _width) in enumerate(columns, start=1):
            value = values.get(key)
            if isinstance(value, float) and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            if key == "note":
                cell.alignment = wrap_left
            elif key in wrap_keys:
                cell.alignment = wrap_center
            else:
                cell.alignment = center
            if key in (PICHIA_OD_COL, PICHIA_TARGET_COL):
                cell.fill = fillin_fill
            elif row_fill is not None:
                cell.fill = row_fill

    # legend lives on its own sheet, not appended below the data table --
    # otherwise pd.read_excel() on reimport reads it as extra data rows
    # (confirmed: it silently inflated an 18-row design to 23 rows).
    legend_ws = wb.create_sheet("图例说明")
    legend_ws.sheet_view.showGridLines = False
    legend_ws.column_dimensions["A"].width = 4
    legend_ws.column_dimensions["B"].width = 40
    legend_items = [
        (baseline_fill, "基线重复（估计批次噪声）"),
        (combo_fill, "联合探索点（LHS，多变量同时变化）"),
        (fillin_fill, "需要填写的结果列"),
    ]
    for row_idx, (fill, text) in enumerate(legend_items, start=1):
        marker = legend_ws.cell(row=row_idx, column=1, value=" ")
        marker.fill = fill
        marker.border = border
        label_cell = legend_ws.cell(row=row_idx, column=2, value=text)
        label_cell.font = Font(name=font_name, size=10)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

def _pichia_round2_workbook_bytes(df: pd.DataFrame) -> bytes:
    """Round 2 twin of _pichia_round1_workbook_bytes: same visual style
    (colored rows by run_type, Chinese headers, notes column, legend), but
    two differences instead of a from-scratch rewrite. (1) run_type/
    changed_variable ride along as hidden columns rather than being encoded
    into "类型" text and decoded back on reimport -- round 1's "类型" decoder
    only recognizes 基线重复/联合探索/单变量-X, and teaching it four more
    prefixes for a one-off Round 2 sheet isn't worth the risk of a subtly
    wrong decode; keeping the machine-readable columns literally present
    means _pichia_remap_uploaded_columns needs no Round-2-specific case at
    all. (2) row colors come from the now 7-entry PICHIA_RUN_TYPE_COLORS
    rather than the 2-entry round-1 subset."""
    run_type = df["run_type"] if "run_type" in df.columns else pd.Series("", index=df.index)
    changed_variable = df["changed_variable"] if "changed_variable" in df.columns else pd.Series(None, index=df.index)
    baseline_lookup = _pichia_baseline_lookup(df)

    font_name = "Calibri"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name=font_name, size=11, bold=True, color="FFFFFF")
    fillin_fill = PatternFill("solid", fgColor="FFF2CC")
    type_fills = {run_type_key: PatternFill("solid", fgColor=color.lstrip("#")) for run_type_key, color in PICHIA_RUN_TYPE_COLORS.items()}
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center = Alignment(horizontal="center", vertical="center")
    wrap_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    wrap_keys = {"run_type", "note"}

    columns: list[tuple[str, str, float]] = [("run_id", "编号", 9), ("run_type", "类型", 20)]
    for variable in PICHIA_VARIABLES:
        if variable in df.columns:
            columns.append((variable, PICHIA_VARIABLE_LABELS.get(variable, variable), 13))
    columns.append((PICHIA_OD_COL, "收获时OD600(待填)", 12))
    columns.append((PICHIA_TARGET_COL, "hLF产量(待填)", 12))
    columns.append(("note", "备注/目的", 36))
    hidden_columns: list[tuple[str, str, float]] = [("_run_type_raw", "run_type", 12), ("_changed_variable_raw", "changed_variable", 16)]

    wb = Workbook()
    ws = wb.active
    ws.title = "Round2设计"
    ws.sheet_view.showGridLines = False

    all_columns = columns + hidden_columns
    for col_idx, (_key, label, width) in enumerate(all_columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for col_idx in range(len(columns) + 1, len(all_columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    ws.freeze_panes = "A2"

    for offset, idx in enumerate(df.index):
        row_idx = offset + 2
        rt = run_type.loc[idx]
        cv = changed_variable.loc[idx]
        row = df.loc[idx]
        note = _pichia_row_note(rt, cv, row, baseline_lookup)
        values: dict[str, Any] = {
            "run_id": row.get("run_id"),
            "run_type": _pichia_type_label(rt, cv),
            PICHIA_OD_COL: row.get(PICHIA_OD_COL),
            PICHIA_TARGET_COL: row.get(PICHIA_TARGET_COL),
            "note": note,
            "_run_type_raw": rt,
            "_changed_variable_raw": None if pd.isna(cv) else cv,
        }
        for variable in PICHIA_VARIABLES:
            if variable in df.columns:
                values[variable] = row.get(variable)

        row_fill = type_fills.get(rt)
        for col_idx, (key, _label, _width) in enumerate(all_columns, start=1):
            value = values.get(key)
            if isinstance(value, float) and pd.isna(value):
                value = None
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=font_name, size=10)
            cell.border = border
            if key == "note":
                cell.alignment = wrap_left
            elif key in wrap_keys:
                cell.alignment = wrap_center
            else:
                cell.alignment = center
            if key in (PICHIA_OD_COL, PICHIA_TARGET_COL):
                cell.fill = fillin_fill
            elif row_fill is not None:
                cell.fill = row_fill

    legend_ws = wb.create_sheet("图例说明")
    legend_ws.sheet_view.showGridLines = False
    legend_ws.column_dimensions["A"].width = 4
    legend_ws.column_dimensions["B"].width = 40
    present_types = [key for key in PICHIA_RUN_TYPE_COLORS if key in set(run_type.dropna())]
    legend_items = [(type_fills[key], PICHIA_RUN_TYPE_LABELS.get(key, key)) for key in present_types]
    legend_items.append((fillin_fill, "需要填写的结果列"))
    for row_idx, (fill, text) in enumerate(legend_items, start=1):
        marker = legend_ws.cell(row=row_idx, column=1, value=" ")
        marker.fill = fill
        marker.border = border
        label_cell = legend_ws.cell(row=row_idx, column=2, value=text)
        label_cell.font = Font(name=font_name, size=10)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

_YIELD_HEADER_HINT = "产量"
_MG_PER_L_PATTERN = re.compile(r"mg\s*/\s*l", re.IGNORECASE)
_NOT_DETECTED_TOKENS = {"未检测到", "未检出", "未测出", "nd", "n.d.", "n/a"}


def _looks_not_detected(value: Any) -> bool:
    return isinstance(value, str) and value.strip().lower() in _NOT_DETECTED_TOKENS


def _coerce_result_numeric(series: pd.Series) -> pd.Series:
    """Real Round 1 results can hold a qualitative "below detection limit"
    token instead of a number in a result cell -- treated as 0 (a censored-low
    reading is evidence of "very little/none", not a missing measurement),
    everything else parsed as usual so a genuinely blank cell stays NaN."""
    cleaned = series.apply(lambda value: 0.0 if _looks_not_detected(value) else value)
    return pd.to_numeric(cleaned, errors="coerce")


def _average_duplicate_run_ids(df: pd.DataFrame) -> tuple[pd.DataFrame, list[str]]:
    """Collapse 2+ rows sharing a run_id (technical replicate measurements)
    into one row per run_id.

    The 6 design variables are expected to be *identical* across one run_id's
    replicate rows -- that's what makes them replicates. A mismatch there
    (data-entry typo, or a row mislabeled with the wrong run_id) is reported
    as a warning string rather than silently resolved by "whichever row came
    first", since silently picking one would hide a real data problem.

    Every other numeric column (yield, od600, and any future response
    variable such as UPR) is averaged, but the raw values are not discarded:
    "<col>_n" (replicate count), "<col>_spread" (max-min across replicates),
    and "<col>_reps" (the raw values, "|"-joined) ride alongside the mean --
    a large spread on a specific run is exactly the kind of thing averaging
    would otherwise hide, and the original numbers stay recoverable from this
    same table without reopening the raw upload archive. Single-replicate
    run_ids get the same three columns (n=1, spread/reps=None) so the schema
    is uniform whether or not that particular run happened to repeat.
    """
    order = df["run_id"].drop_duplicates().tolist()
    design_columns = [column for column in PICHIA_VARIABLES if column in df.columns]
    passthrough_columns = [column for column in ("run_type", "changed_variable") if column in df.columns]
    numeric_columns = [
        column
        for column in df.columns
        if column not in ("run_id", *design_columns, *passthrough_columns) and pd.api.types.is_numeric_dtype(df[column])
    ]
    leftover_columns = [
        column
        for column in df.columns
        if column not in ("run_id", *design_columns, *passthrough_columns, *numeric_columns)
    ]

    warnings: list[str] = []
    rows: list[dict[str, Any]] = []
    for run_id, group in df.groupby("run_id", sort=False):
        row: dict[str, Any] = {"run_id": run_id}
        for column in (*design_columns, *passthrough_columns, *leftover_columns):
            distinct = group[column].dropna()
            if column in design_columns:
                distinct = distinct.round(6)
            distinct = distinct.unique()
            if column in design_columns and len(distinct) > 1:
                warnings.append(
                    f"编号 {run_id} 的「{PICHIA_VARIABLE_LABELS.get(column, column)}」在重复测量行之间不一致："
                    f"{list(distinct)}，已取第一行的值，请核对是否录入有误"
                )
            row[column] = group[column].iloc[0]
        for column in numeric_columns:
            values = group[column].dropna()
            row[column] = float(values.mean()) if len(values) else float("nan")
            row[f"{column}_n"] = int(len(values))
            row[f"{column}_spread"] = float(values.max() - values.min()) if len(values) > 1 else None
            row[f"{column}_reps"] = "|".join(f"{value:g}" for value in values) if len(values) > 1 else None
        rows.append(row)

    averaged = pd.DataFrame(rows)
    averaged["_order"] = averaged["run_id"].apply(order.index)
    averaged = averaged.sort_values("_order").drop(columns="_order").reset_index(drop=True)
    return averaged, warnings


def _pichia_replicate_spread_report(df: pd.DataFrame) -> pd.DataFrame:
    """Display-friendly extract of the "<col>_n"/"<col>_spread" columns
    _average_duplicate_run_ids attaches, limited to run_ids that actually had
    2+ replicates for the yield or od600 column -- for the upload-time "look
    at this now" callout, sorted by the caller worst-spread-first."""
    records: list[dict[str, Any]] = []
    for _, row in df.iterrows():
        counts = [row.get(f"{column}_n") for column in (PICHIA_TARGET_COL, PICHIA_OD_COL) if f"{column}_n" in df.columns]
        if not any(pd.notna(count) and count > 1 for count in counts):
            continue
        record: dict[str, Any] = {"run_id": row["run_id"], "重复次数": int(max(count for count in counts if pd.notna(count)))}
        for column, label in ((PICHIA_TARGET_COL, "hLF产量"), (PICHIA_OD_COL, "OD600")):
            spread = row.get(f"{column}_spread")
            if pd.isna(spread):
                continue
            mean = row.get(column)
            record[f"{label}均值"] = _num(mean)
            record[f"{label}重复间差值"] = _num(spread)
            record[f"{label}重复原始值"] = row.get(f"{column}_reps")
        records.append(record)
    return pd.DataFrame(records)


def _pichia_pooled_technical_noise(df: pd.DataFrame, column: str = PICHIA_TARGET_COL) -> dict[str, Any] | None:
    """Pool within-run_id replicate variance across every run that has 2+
    technical replicates (parses "<column>_reps", see _average_duplicate_run_ids),
    rather than trusting any single run's own 1-degree-of-freedom spread.

    This is a genuinely different noise source from plan_round2's baseline_sd:
    that one is *between-run* (batch) noise, estimated from how much the 3
    baseline runs' means disagree with each other. This one is *within-run*
    (technical/assay) noise, estimated from how much one run's own repeated
    measurement wobbles. Standard pooled-variance treatment across k groups:
    s_pooled^2 = sum(SS_i) / sum(n_i - 1), which is why pooling across every
    run_id (not just baseline) matters here -- 16 runs x 1 df each gives a far
    more stable estimate than the 2 df available from 3 baseline runs alone.

    Deliberately kept separate from estimate_baseline_noise's significance
    threshold, not just left unfinished: every round 1 sample is measured with
    the same 2 technical replicates, so the between-run (batch) SD already IS
    the right noise scale for comparing one run's mean against another's --
    folding this in wouldn't change which variables clear the threshold (for
    2026-08 Y103 data, this pooled SD is smaller than baseline_sd anyway, so a
    "use whichever is stricter" rule would be a no-op), and the threshold only
    gates a soft, self-correcting choice (which variable gets a round-2 CCD
    slot) where the added rigor isn't worth the small-sample estimation risk a
    formal variance-components combination would carry. This function's value
    is as a data-quality flag on individual runs, independent of that decision.
    Returns None if the column has no replicate-detail data at all.
    """
    reps_column = f"{column}_reps"
    if reps_column not in df.columns:
        return None

    sum_of_squares = 0.0
    degrees_of_freedom = 0
    per_run_sd: list[tuple[str, float]] = []
    for _, row in df.iterrows():
        raw = row.get(reps_column)
        if not isinstance(raw, str) or not raw:
            continue
        values = [float(token) for token in raw.split("|")]
        if len(values) < 2:
            continue
        mean = sum(values) / len(values)
        run_ss = sum((value - mean) ** 2 for value in values)
        sum_of_squares += run_ss
        degrees_of_freedom += len(values) - 1
        per_run_sd.append((str(row["run_id"]), (run_ss / (len(values) - 1)) ** 0.5))

    if degrees_of_freedom == 0:
        return None

    pooled_sd = (sum_of_squares / degrees_of_freedom) ** 0.5
    outliers = sorted(
        ((run_id, sd) for run_id, sd in per_run_sd if sd > 2 * pooled_sd),
        key=lambda item: item[1],
        reverse=True,
    )
    return {
        "pooled_sd": pooled_sd,
        "dof": degrees_of_freedom,
        "n_runs": len(per_run_sd),
        "outliers": outliers,
    }


def _pichia_remap_uploaded_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    """Recognizes the polished Excel export's Chinese headers, in addition to
    the plain English round1_template_columns() schema, so a design that was
    downloaded, filled in by hand in Excel, and re-uploaded round-trips
    correctly. Also reconstructs run_type/changed_variable from the combined
    "类型" column (e.g. "单变量-发酵温度 (℃)") -- round2's significance
    analysis and the result charts both key off those two columns, so losing
    them on reimport would silently break Round 2, not just cosmetics.

    Two real-world variations seen in returned lab data (2026-08 Y103 round 1)
    are also handled: (1) the yield header can carry a unit annotation this
    app's own template doesn't include (e.g. "hLF产量（mg/L）(待填)") -- matched
    by "产量" appearing anywhere in a header rather than an exact string, and
    auto-converted to g/L when "mg/L" is present, since every yield_g_per_l
    consumer in this codebase assumes g/L; (2) a "未检测到" token in a result
    cell instead of a number.

    Returns (cleaned_df, replicate_spread_report, consistency_warnings).
    cleaned_df keeps one row per run_id -- analyze_round1_effects/plan_round2
    are built on that shape, and feeding raw un-collapsed replicate rows
    through them would silently double-count each run_id's evidence and
    conflate technical (within-run) noise with the between-run noise the
    baseline-replicate significance threshold is meant to estimate. Nothing
    from the original rows is discarded in exchange for that shape, though:
    see _average_duplicate_run_ids for the "<col>_n/_spread/_reps" columns
    that keep the raw values reachable from cleaned_df itself, and
    replicate_spread_report/consistency_warnings for what's worth a human
    looking at right away. Both are empty/[] unless some run_id had 2+ rows.
    """
    label_to_variable = {label: variable for variable, label in PICHIA_VARIABLE_LABELS.items()}
    rename_map = {"编号": "run_id", "收获时OD600(待填)": PICHIA_OD_COL}
    rename_map.update(label_to_variable)
    renamed = df.rename(columns=rename_map)

    yield_source = next(
        (
            column
            for column in renamed.columns
            if isinstance(column, str) and _YIELD_HEADER_HINT in column and column != PICHIA_TARGET_COL
        ),
        None,
    )
    if yield_source is not None:
        scale = 0.001 if _MG_PER_L_PATTERN.search(yield_source) else 1.0
        renamed[PICHIA_TARGET_COL] = _coerce_result_numeric(renamed[yield_source]) * scale
        renamed = renamed.drop(columns=[yield_source])
    elif PICHIA_TARGET_COL in renamed.columns:
        renamed[PICHIA_TARGET_COL] = _coerce_result_numeric(renamed[PICHIA_TARGET_COL])
    if PICHIA_OD_COL in renamed.columns:
        renamed[PICHIA_OD_COL] = _coerce_result_numeric(renamed[PICHIA_OD_COL])

    if "类型" in renamed.columns and "run_type" not in renamed.columns:
        run_types: list[str | None] = []
        changed_vars: list[str | None] = []
        for raw in renamed["类型"]:
            text = str(raw).strip()
            if text == PICHIA_RUN_TYPE_LABELS.get("baseline"):
                run_types.append("baseline")
                changed_vars.append(None)
            elif text == PICHIA_RUN_TYPE_LABELS.get("combo"):
                run_types.append("combo")
                changed_vars.append(None)
            elif text.startswith("单变量-"):
                run_types.append("ofat")
                changed_vars.append(label_to_variable.get(text[len("单变量-"):]))
            else:
                run_types.append(None)
                changed_vars.append(None)
        renamed["run_type"] = run_types
        renamed["changed_variable"] = changed_vars
        renamed = renamed.drop(columns=["类型"])

    renamed = renamed.drop(columns=["备注/目的"], errors="ignore")

    # defensive: drop any row missing a variable value -- every real design
    # row has all 6 populated, so this filters out a legend/blank row picked
    # up from an older export that still had the legend below the table (or
    # any stray manual annotation row) without needing to special-case it.
    variable_columns = [variable for variable in PICHIA_VARIABLES if variable in renamed.columns]
    if variable_columns:
        renamed = renamed[renamed[variable_columns].notna().all(axis=1)].reset_index(drop=True)

    if "run_id" in renamed.columns and renamed["run_id"].duplicated().any():
        renamed, consistency_warnings = _average_duplicate_run_ids(renamed)
        spread_report = _pichia_replicate_spread_report(renamed)
    else:
        spread_report = pd.DataFrame()
        consistency_warnings = []

    return renamed, spread_report, consistency_warnings
